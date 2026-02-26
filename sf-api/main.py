from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from simple_salesforce import Salesforce
from dotenv import load_dotenv
import openpyxl
from copy import copy as style_copy
from openpyxl.utils import get_column_letter
import base64
import datetime
import os
import json
import html
import io
import re # re is already imported but consistent with request
import requests
import os
import datetime # existing import is just 'import datetime', user snippet uses 'from datetime import datetime' but we can adapt or just import what's needed.
# existing imports...

# Load environment variables
load_dotenv()
groq_client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

# ================= CẤU HÌNH BASE.VN =================
SF_CASE_ID_DEFAULT = '500fD00000XSvMwQAL'

BASE_SERVICE_ID = "7204"
BASE_BLOCK_ID_CREATE = "7210"
BASE_USERNAME = "PhuongTran"

KEYS = {
    "MA_KH": "service_ma_khach_hang",
    "NGAY_PHAN_ANH": "service_ngay_phan_anh",
    "NOI_DUNG": "service_noi_dung_khieu_nai",
    "NGAY_XUAT": "service_ngay_xuat_ngay_tau",
    "SO_CONT": "service_so_container",
    "LSX": "service_so_lenh_san_xuat"
}

URL_GET_ALL = "https://service.base.vn/extapi/v1/ticket/get.all"
URL_GET_DETAIL = "https://service.base.vn/extapi/v1/ticket/get.detail"
URL_CREATE = "https://service.base.vn/extapi/v1/ticket/create"
URL_EDIT_CUSTOM = "https://service.base.vn/extapi/v1/ticket/edit.custom.fields"


app = FastAPI(title="Salesforce Packing List API")

# Add CORS Middleware
from fastapi.middleware.cors import CORSMiddleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

class ShipmentRequest(BaseModel):
    shipment_id: str

def expand_items_table(ws, template_row, n):
    """Expand the items table to accommodate n rows"""
    max_col = ws.max_column
    row_style = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=template_row, column=col)
        row_style.append(style_copy(cell._style) if cell.has_style else None)
    row_height = ws.row_dimensions[template_row].height
    add_rows = max(0, n - 1)
    
    # Handle merged cells
    merges_to_shift = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > template_row:
            merges_to_shift.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))
    
    for mr in merges_to_shift:
        rng = f"{get_column_letter(mr[2])}{mr[0]}:{get_column_letter(mr[3])}{mr[1]}"
        ws.unmerge_cells(rng)
    
    # Insert rows
    if add_rows > 0:
        ws.insert_rows(template_row + 1, amount=add_rows)
        for offset in range(1, add_rows + 1):
            r = template_row + offset
            for col in range(1, max_col + 1):
                dst = ws.cell(row=r, column=col)
                dst.value = None
                st = row_style[col - 1]
                if st is not None:
                    dst._style = style_copy(st)
            if row_height is not None:
                ws.row_dimensions[r].height = row_height
    
    # Re-merge shifted cells
    for mr in merges_to_shift:
        new_min_row = mr[0] + add_rows
        new_max_row = mr[1] + add_rows
        rng = f"{get_column_letter(mr[2])}{new_min_row}:{get_column_letter(mr[3])}{new_max_row}"
        ws.merge_cells(rng)
    
    # Update total formulas
    total_header_row = None
    for r in range(template_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Total":
            total_header_row = r
            break
    
    if total_header_row is None:
        raise ValueError("Total row not found")
    
    first_data_row = template_row
    last_data_row = template_row + n - 1
    ws[f"H{total_header_row}"] = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws[f"J{total_header_row}"] = f"=SUM(J{first_data_row}:J{last_data_row})"
    ws[f"K{total_header_row}"] = f"=COUNTA(K{first_data_row}:K{last_data_row})"

def get_salesforce_connection():
    """Initialize Salesforce connection"""
    username = os.getenv('SALESFORCE_USERNAME')
    password = os.getenv('SALESFORCE_PASSWORD')
    security_token = os.getenv('SALESFORCE_SECURITY_TOKEN')
    consumer_key = os.getenv('SALESFORCE_CONSUMER_KEY')
    consumer_secret = os.getenv('SALESFORCE_CONSUMER_SECRET')
    
    if None in (username, password, security_token, consumer_key, consumer_secret):
        raise ValueError("Salesforce credentials missing in environment variables")
    
    return Salesforce(
        username=username,
        password=password,
        security_token=security_token,
        consumer_key=consumer_key,
        consumer_secret=consumer_secret
    )

def get_picklist_values(sf, object_name: str, field_name: str) -> list[str]:
    """
    Get picklist values dynamically from Salesforce for any object and field.
    
    Args:
        sf: Salesforce connection instance
        object_name: API name of the Salesforce object (e.g., 'Shipment__c')
        field_name: API name of the picklist field (e.g., 'Freight__c')
    
    Returns:
        List of picklist option values
    """
    try:
        sobject = getattr(sf, object_name)
        description = sobject.describe()
        
        for field in description['fields']:
            if field['name'] == field_name:
                if field['type'] == 'picklist' or field['type'] == 'multipicklist':
                    return [option['value'] for option in field['picklistValues'] if option['active']]
                else:
                    print(f"⚠ Warning: {field_name} is not a picklist field (type: {field['type']})")
                    return []
        
        print(f"⚠ Warning: {field_name} field not found on {object_name}")
        return []
    except Exception as e:
        print(f"⚠ Warning: Could not fetch picklist values for {object_name}.{field_name}: {e}")
        return []

def get_output_directory() -> Path:
    """
    Get the appropriate output directory based on environment.
    Use /tmp for serverless environments (Vercel, AWS Lambda) where filesystem is read-only.
    Use ./output for local development.
    """
    # Check if we're in a serverless environment
    is_serverless = (
        os.getenv('VERCEL') is not None or  # Vercel
        os.getenv('AWS_LAMBDA_FUNCTION_NAME') is not None or  # AWS Lambda
        os.getenv('LAMBDA_TASK_ROOT') is not None  # AWS Lambda alternative
    )
    
    if is_serverless:
        output_dir = Path("/tmp")
    else:
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)
    
    return output_dir

def sanitize_filename(filename):
    """
    Sanitize the filename by removing or replacing invalid characters.
    """
    if not filename:
        return "Unknown"
    # Replace invalid characters with underscore
    return re.sub(r'[\\/*?:"<>|]', '_', str(filename)).strip()


def format_picklist_checkboxes(options, selected_value, uppercase=False):
    """
    Format picklist options as a checkbox list.
    Mark the selected value with ☑, others with ☐.
    """
    checked_box = '☑'
    unchecked_box = '☐'
    formatted_lines = []
    
    # Normalize selected value for comparison
    selected_value_norm = str(selected_value or '').strip().upper()
    
    for opt in options:
        opt_label = str(opt)
        # Compare uppercase to ensure matches work
        is_selected = opt_label.upper() == selected_value_norm
        
        if uppercase:
            opt_label = opt_label.upper()
            
        checkbox = checked_box if is_selected else unchecked_box
        formatted_lines.append(f"{checkbox} {opt_label}")
        
    return "\n".join(formatted_lines)

def generate_packing_list(shipment_id: str, template_path: str):
    """Generate packing list for a given shipment ID"""
    
    # Connect to Salesforce
    sf = get_salesforce_connection()
    
    # Get freight options dynamically from Salesforce
    freight_options = get_picklist_values(sf, 'Shipment__c', 'Freight__c')
    
    # Query shipment data
    shipment_query = f"""
    SELECT Name, Consignee__c, Invoice_Packing_list_no__c, Issued_date__c, Port_of_Origin__c,
    Final_Destination__c, Stockyard__c, Ocean_Vessel__c, B_L_No__c, Freight__c,
    Departure_Date_ETD__c, Arrival_Schedule_ETA__c, Remark_number_on_documents__c,
    Terms_of_Sales__c, Terms_of_Payment__c
    FROM Shipment__c
    WHERE Id = '{shipment_id}'
    """
    shipment_result = sf.query(shipment_query)
    if not shipment_result['records']:
        raise ValueError(f"No Shipment found with ID: {shipment_id}")
    shipment = shipment_result['records'][0]
    
    # Query account/consignee data
    if shipment['Consignee__c']:
        account_query = f"""
        SELECT Name, BillingStreet, BillingCity, BillingPostalCode, BillingCountry,
        Phone, Fax__c, VAT__c
        FROM Account
        WHERE Id = '{shipment['Consignee__c']}'
        """
        account_result = sf.query(account_query)
        account = account_result['records'][0] if account_result['records'] else {}
    else:
        account = {}
    
    # Query bookings
    bookings_query = f"""
    SELECT Id, Cont_Quantity__c
    FROM Booking__c
    WHERE Shipment__c = '{shipment_id}'
    """
    bookings_result = sf.query_all(bookings_query)
    bookings = bookings_result['records']
    total_containers_from_bookings = sum(booking.get('Cont_Quantity__c') or 0 for booking in bookings)
    
    # Query container items
    items_query = f"""
    SELECT Line_item_no_for_print__c, Product_Description__c, Length__c, Width__c, Height__c,
    Quantity_For_print__c, Unit_for_print__c, Crates__c, Packing__c, Order_No__c,
    Container__r.Name, Container__r.Container_Weight_Regulation__c
    FROM Container_Item__c
    WHERE Shipment__c = '{shipment_id}'
    """
    items_result = sf.query_all(items_query)
    items = items_result['records']
    
    # Load template
    wb = openpyxl.load_workbook(template_path)
    ws = wb['PackingList']
    
    # Replace placeholders (excluding Freight__c as it needs special handling)
    replacements = {
        '{{Shipment__c.Consignee__r.Name}}': account.get('Name') or '',
        '{{Shipment__c.Consignee__r.BillingStreet}}': account.get('BillingStreet') or '',
        '{{Shipment__c.Consignee__r.BillingCity}}': account.get('BillingCity') or '',
        '{{Shipment__c.Consignee__r.BillingPostalCode}}': account.get('BillingPostalCode') or '',
        '{{Shipment__c.Consignee__r.BillingCountry}}': account.get('BillingCountry') or '',
        '{{Shipment__c.Consignee__r.Phone}}': account.get('Phone') or '',
        '{{Shipment__c.Consignee__r.Fax__c}}': account.get('Fax__c') or '',
        '{{Shipment__c.Consignee__r.VAT__c}}': account.get('VAT__c') or '',
        '{{Shipment__c.Invoice_Packing_list_no__c}}': shipment.get('Invoice_Packing_list_no__c') or '',
        '{{Shipment__c.Issued_date__c}}': shipment.get('Issued_date__c') or '',
        '{{Shipment__c.Port_of_Origin__c}}': shipment.get('Port_of_Origin__c') or '',
        '{{Shipment__c.Final_Destination__c}}': shipment.get('Final_Destination__c') or '',
        '{{Shipment__c.Stockyard__c}}': shipment.get('Stockyard__c') or '',
        '{{Shipment__c.Ocean_Vessel__c}}': shipment.get('Ocean_Vessel__c') or '',
        '{{Shipment__c.B_L_No__c}}': shipment.get('B_L_No__c') or '',
        '{{Shipment__c.Departure_Date_ETD__c}}': shipment.get('Departure_Date_ETD__c') or '',
        '{{Shipment__c.Arrival_Schedule_ETA__c}}': shipment.get('Arrival_Schedule_ETA__c') or '',
        '{{Shipment__c.Remark_number_on_documents__c}}': shipment.get('Remark_number_on_documents__c') or '',
        '{{Shipment__c.Terms_of_Sales__c}}': shipment.get('Terms_of_Sales__c') or '',
        '{{Shipment__c.Terms_of_Payment__c}}': shipment.get('Terms_of_Payment__c') or '',
    }
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for placeholder, value in replacements.items():
                    cell.value = cell.value.replace(placeholder, str(value))
                if '{{TableStart:Shipment__c.r.Bookings__r}}' in cell.value:
                    cell.value = str(total_containers_from_bookings)
    
    # Remove "None" values
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and 'None' in cell.value:
                cell.value = cell.value.replace('None', '')
    
    # Handle freight checkboxes dynamically
    checked_box = '☑'
    unchecked_box = '☐'
    
    freight_value = (shipment.get('Freight__c') or '').strip()
    freight_upper = freight_value.upper()
    
    # Generate checkbox text with all options from Salesforce
    lines = []
    for opt in freight_options:
        mark = checked_box if opt.upper() == freight_upper else unchecked_box
        lines.append(f"{mark} {opt}")
    
    checkbox_text = "\n".join(lines)
    
    # Replace freight placeholder with checkbox text
    # for row in ws.iter_rows():
    #     for cell in row:
    #         if isinstance(cell.value, str) and '{{Shipment__c.Freight__c}}' in cell.value:
    #             cell.value = cell.value.replace('{{Shipment__c.Freight__c}}', checkbox_text)
    
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '{{Shipment__c.Freight__c}}' in cell.value:
                cell.value = cell.value.replace('{{Shipment__c.Freight__c}}', checkbox_text)
                if cell.alignment:
                    new_alignment = style_copy(cell.alignment)
                else:
                    from openpyxl.styles import Alignment
                    new_alignment = Alignment()
                new_alignment.wrap_text = True
                cell.alignment = new_alignment
    
    # Find table start row
    table_start_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=13):
        for cell in row:
            if cell.value and '{{TableStart:ContainerItems}}' in str(cell.value):
                table_start_row = cell.row
                break
        if table_start_row:
            break
    
    if not table_start_row:
        raise ValueError("No table start marker found in template")
    
    # Expand table
    expand_items_table(ws, table_start_row, len(items) if items else 1)
    
    # Fill in item data
    for idx, item in enumerate(items):
        row = table_start_row + idx
        container_r = item.get('Container__r', {})
        line_item_no = item.get('Line_item_no_for_print__c') or str(idx + 1)
        ws.cell(row, 1).value = line_item_no
        ws.cell(row, 2).value = item.get('Product_Description__c')
        ws.cell(row, 3).value = item.get('Length__c')
        ws.cell(row, 4).value = item.get('Width__c')
        ws.cell(row, 5).value = item.get('Height__c')
        ws.cell(row, 6).value = item.get('Quantity_For_print__c') or ''
        ws.cell(row, 7).value = item.get('Unit_for_print__c') or ''
        ws.cell(row, 8).value = item.get('Crates__c')
        ws.cell(row, 9).value = f"{item.get('Packing__c') or ''} pcs/crate"
        ws.cell(row, 10).value = container_r.get('Container_Weight_Regulation__c')
        ws.cell(row, 11).value = container_r.get('Name')
        ws.cell(row, 13).value = item.get('Order_No__c')
    
    # Save file
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"Packing_List_{shipment.get('Invoice_Packing_list_no__c', shipment['Name'])}_{timestamp}.xlsx"
    
    # Use appropriate output directory based on environment
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    
    wb.save(str(file_path))
    
    # Upload to Salesforce
    with open(file_path, "rb") as f:
        data = f.read()
    encoded = base64.b64encode(data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": shipment_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version['id'],
        "freight_options_used": freight_options
    }

@app.get("/")
async def root():
    """Root endpoint"""
    return {
        "message": "Salesforce Packing List API",
        "version": "1.0.0",
        "endpoints": {
            "GET /health": "Health check",
            "GET /generate-packing-list": "Generate packing list (test endpoint)",
            "POST /generate-packing-list": "Generate packing list (production endpoint)",
            "GET /generate_invoice/{shipment_id}": "Generate invoice for a shipment",
            "GET /generate-combined-export/{shipment_id}": "Generate combined packing list and invoice in one Excel file",
            "GET /download/{file_name}": "Download generated packing list file"
        }
    }

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    try:
        # Test Salesforce connection
        sf = get_salesforce_connection()
        freight_options = get_picklist_values(sf, 'Shipment__c', 'Freight__c')
        return {
            "status": "healthy",
            "salesforce_connected": True,
            "freight_options": freight_options,
            "timestamp": datetime.datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Service unhealthy: {str(e)}")

@app.get("/generate-packing-list")
async def generate_packing_list_get(shipment_id: str):
    """
    Generate packing list for a shipment (GET method for testing)
    
    Parameters:
    - shipment_id: Salesforce Shipment ID
    """
    try:
        template_path = os.getenv('TEMPLATE_PATH', 'templates/packing_list_template.xlsx')
        
        if not os.path.exists(template_path):
            raise HTTPException(
                status_code=404,
                detail=f"Template file not found at: {template_path}"
            )
        
        result = generate_packing_list(shipment_id, template_path)
        
        return {
            "status": "success",
            "message": "Packing list generated successfully",
            "data": result
        }
    
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating packing list: {str(e)}")

@app.post("/generate-packing-list")
async def generate_packing_list_post(request: ShipmentRequest):
    """
    Generate packing list for a shipment (POST method)
    
    Parameters:
    - shipment_id: Salesforce Shipment ID (in request body)
    """
    try:
        template_path = os.getenv('TEMPLATE_PATH', 'templates/packing_list_template.xlsx')
        
        if not os.path.exists(template_path):
            raise HTTPException(
                status_code=404,
                detail=f"Template file not found at: {template_path}"
            )
        
        result = generate_packing_list(request.shipment_id, template_path)
        
        return {
            "status": "success",
            "message": "Packing list generated successfully",
            "data": result
        }
    
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating packing list: {str(e)}")

# ================= HELPERS BASE.VN =================

def convert_html_to_richtext(raw_html):
    if not raw_html: return ""
    text = re.sub(r'<(br\s*/?|/p|/div|/tr)>', '\n', raw_html, flags=re.IGNORECASE)
    text = re.sub(r'<li.*?>', '\n- ', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    text = html.unescape(text)
    return '\n'.join([line.strip() for line in text.split('\n') if line.strip()])

def format_date_base(iso_date_str):
    if not iso_date_str: return ""
    try:
        # Assuming iso_date_str is YYYY-MM-DD...
        date_obj = datetime.datetime.strptime(iso_date_str[:10], "%Y-%m-%d")
        return date_obj.strftime("%d/%m/%Y")
    except: return iso_date_str

def get_sf_data(sf, case_id):
    print(f"--- [SF] Lấy dữ liệu Case {case_id} ---")
    query = f"SELECT Id, Subject, Customer_Complain_Content__c, So_LSX__c, Date_Export__c, Number_Container__c, CreatedDate, Account.Account_Code__c FROM Case WHERE Id = '{case_id}'"
    res = sf.query(query)
    if not res['records']: return None
    rec = res['records'][0]
    return {
        KEYS['MA_KH']: rec.get('Account', {}).get('Account_Code__c', ''),
        KEYS['NGAY_PHAN_ANH']: format_date_base(rec.get('CreatedDate')),
        KEYS['NOI_DUNG']: convert_html_to_richtext(rec.get('Customer_Complain_Content__c')),
        KEYS['NGAY_XUAT']: format_date_base(rec.get('Date_Export__c')),
        KEYS['SO_CONT']: rec.get('Number_Container__c', ''),
        KEYS['LSX']: rec.get('So_LSX__c', ''),
        "subject": rec.get('Subject', 'No Subject')
    }

def download_sf_files(sf, case_id):
    files_payload = []
    q = f"SELECT ContentDocument.Title, ContentDocument.FileExtension, ContentDocument.LatestPublishedVersionId FROM ContentDocumentLink WHERE LinkedEntityId = '{case_id}'"
    res = sf.query(q)
    for rec in res['records']:
        ver_id = rec['ContentDocument']['LatestPublishedVersionId']
        fname = f"{rec['ContentDocument']['Title']}.{rec['ContentDocument']['FileExtension']}"
        d_url = f"https://{sf.sf_instance}/services/data/v52.0/sobjects/ContentVersion/{ver_id}/VersionData"
        r = requests.get(d_url, headers={"Authorization": f"Bearer {sf.session_id}"}, stream=True)
        if r.status_code == 200:
            files_payload.append(('root_file[]', (fname, io.BytesIO(r.content), 'application/octet-stream')))
    return files_payload

def find_ticket_id(subject):
    resp = requests.post(URL_GET_ALL, data={"access_token_v2": os.getenv("SERVICE_ACCESS_TOKEN"), "service_id": BASE_SERVICE_ID})
    try:
        data = resp.json()
        for t in data.get('tickets', []):
            if t.get('name', '').strip() == subject.strip(): return t.get('id')
    except Exception as e:
        print(f"Error finding ticket: {e}")
    return None

def create_ticket(subject, sf_data):
    print("--- [BASE] Tạo phiếu mới ---")
    payload = {
        "access_token_v2": os.getenv("SERVICE_ACCESS_TOKEN"),
        "service_id": BASE_SERVICE_ID,
        "block_id": BASE_BLOCK_ID_CREATE,
        "username": BASE_USERNAME,
        "name": subject
    }
    # Tối ưu: Update thẳng data vào payload Create, không cần custom_field_ids
    payload.update({k: v for k, v in sf_data.items() if k != 'subject'})
    resp = requests.post(URL_CREATE, data=payload)
    return resp.json().get('data', {}).get('id')

def update_smart(ticket_id, sf_data, files):
    print(f"--- [BASE] Kiểm tra đồng bộ Ticket {ticket_id} ---")
    detail = requests.post(URL_GET_DETAIL, data={"access_token_v2": os.getenv("SERVICE_ACCESS_TOKEN"), "id": ticket_id}).json()
    ticket = detail.get('tickets', [{}])[0]
    
    # 1. So sánh Field
    current_fields = {f['key']: str(f.get('value', '')).strip() for f in ticket.get('custom_object', [])}
    fields_to_up = {}
    for k, v in sf_data.items():
        if k == 'subject': continue
        target = str(v or '').strip()
        if current_fields.get(k) != target:
            fields_to_up[k] = target

    # 2. So sánh File
    existing_files = {f.get('name') for f in ticket.get('files', [])}
    if 'root_export' in ticket:
        existing_files.update({f.get('name') for f in ticket['root_export'].get('files', [])})
    
    files_to_up = [f for f in files if f[1][0] not in existing_files]

    if not fields_to_up and not files_to_up:
        print("   -> Đã đồng bộ. Bỏ qua.")
        return

    # 3. Gửi Update: Bắt buộc kèm custom_field_ids
    payload = {
        "access_token_v2": os.getenv("SERVICE_ACCESS_TOKEN"),
        "service_id": BASE_SERVICE_ID,
        "ticket_id": ticket_id,
        "username": BASE_USERNAME,
        "custom_field_ids": ",".join(fields_to_up.keys())
    }
    payload.update(fields_to_up)
    resp = requests.post(URL_EDIT_CUSTOM, data=payload, files=files_to_up if files_to_up else None)
    print(f"   -> Kết quả: {resp.status_code}")

@app.get("/sync-base-service")
async def sync_base_service(case_id: str = SF_CASE_ID_DEFAULT):
    """
    Sync logic from Salesforce Case to Base.vn Ticket
    """
    try:
        sf = get_salesforce_connection()
        data = get_sf_data(sf, case_id)
        if not data:
            return {"status": "error", "message": "Case không tồn tại."}
        
        files = download_sf_files(sf, case_id)
        t_id = find_ticket_id(data['subject'])
        
        action = "none"
        if not t_id:
            t_id = create_ticket(data['subject'], data)
            action = "created"
        
        if t_id:
            update_smart(t_id, data, files)
            if action == "none": action = "checked/updated"

        # Close files
        for _, f in files: f[1].close()

        return {
            "status": "success",
            "message": f"Synced successfully. Ticket ID: {t_id}",
            "action": action,
            "case_id": case_id
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error syncing to Base.vn: {str(e)}")


def expand_invoice_items_table(ws, template_row: int, n: int) -> None:
    max_col = ws.max_column
    row_style = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=template_row, column=col)
        row_style.append(style_copy(cell._style) if cell.has_style else None)
    row_height = ws.row_dimensions[template_row].height
    add_rows = max(0, n - 1)

    merges_to_shift = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > template_row:
            merges_to_shift.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))

    for mr in merges_to_shift:
        rng = f"{get_column_letter(mr[2])}{mr[0]}:{get_column_letter(mr[3])}{mr[1]}"
        ws.unmerge_cells(rng)

    if add_rows > 0:
        ws.insert_rows(template_row + 1, amount=add_rows)
        for offset in range(1, add_rows + 1):
            r = template_row + offset
            for col in range(1, max_col + 1):
                dst = ws.cell(row=r, column=col)
                dst.value = None
                st = row_style[col - 1]
                if st is not None:
                    dst._style = style_copy(st)
            if row_height is not None:
                ws.row_dimensions[r].height = row_height

    for mr in merges_to_shift:
        new_min_row = mr[0] + add_rows
        new_max_row = mr[1] + add_rows
        rng = f"{get_column_letter(mr[2])}{new_min_row}:{get_column_letter(mr[3])}{new_max_row}"
        ws.merge_cells(rng)

@app.get("/generate_invoice/{shipment_id}")
def generate_invoice(shipment_id: str):
    sf = get_salesforce_connection()

    # Base and discount templates
    base_template_path = "./templates/invoice_template.xlsx"
    discount_template_path = "./templates/invoice_template_w_discount.xlsx"
    template_path = base_template_path

    # Get all picklist values dynamically
    freight_options = get_picklist_values(sf, 'Shipment__c', 'Freight__c')
    terms_of_sales_options = get_picklist_values(sf, 'Shipment__c', 'Terms_of_Sales__c')
    terms_of_payment_options = get_picklist_values(sf, 'Shipment__c', 'Terms_of_Payment__c')

    shipment_query = f"""
    SELECT Name, Consignee__c, Invoice_Packing_list_no__c, Issued_date__c,
           Port_of_Origin__c, Final_Destination__c, Stockyard__c,
           Ocean_Vessel__c, B_L_No__c, Freight__c,
           Departure_Date_ETD__c, Arrival_Schedule_ETA__c,
           Remark_number_on_documents__c,
           Terms_of_Sales__c, Terms_of_Payment__c,
           Subtotal_USD__c, Fumigation__c, In_words__c,
           Total_Price_USD__c, Surcharge_amount_USD__c,
           Discount_Percentage__c, Discount_Amount__c
    FROM Shipment__c
    WHERE Id = '{shipment_id}'
    """
    shipment_result = sf.query(shipment_query)
    if not shipment_result["records"]:
        raise ValueError(f"No Shipment found with ID: {shipment_id}")
    shipment = shipment_result["records"][0]

    # Determine if discount exists on the shipment
    discount_percentage = shipment.get("Discount_Percentage__c")
    discount_amount = shipment.get("Discount_Amount__c")

    discount_exists = any(
        v not in (None, 0, "", "0", 0.0)
        for v in (discount_percentage, discount_amount)
    )

    # Choose template based on discount
    if discount_exists:
        template_path = discount_template_path

    # Account / Consignee
    if shipment.get("Consignee__c"):
        account_query = f"""
        SELECT Name, BillingStreet, BillingCity, BillingPostalCode, BillingCountry,
               Phone, Fax__c, VAT__c
        FROM Account
        WHERE Id = '{shipment['Consignee__c']}'
        """
        account_result = sf.query(account_query)
        account = account_result["records"][0] if account_result["records"] else {}
    else:
        account = {}

    # Container items
    items_query = f"""
    SELECT Line_item_no_for_print__c, Product_Description__c,
           Length__c, Width__c, Height__c,
           Quantity_For_print__c, Unit_for_print__c,
           Sales_Price_USD__c, Charge_Unit__c,
           Total_Price_USD__c, Order_No__c,
           Container__r.STT_Cont__c
    FROM Container_Item__c
    WHERE Shipment__c = '{shipment_id}'
    ORDER BY Line_item_no_for_print__c
    """
    items_result = sf.query_all(items_query)
    items = items_result["records"]

    # 💡 FIXED: use existing fields only
    deposit_query = f"""
    SELECT Contract_PI__r.Name, Reconciled_Amount__c
    FROM Receipt_Reconciliation__c
    WHERE Invoice__c = '{shipment_id}'
    """
    deposits = sf.query_all(deposit_query)["records"]

    refunds_query = f"""
    SELECT Reason, Refund_Amount__c
    FROM Case
    WHERE Refund_in_Shipment__c = '{shipment_id}'
    """
    refunds = sf.query_all(refunds_query)["records"]

    # Build debug data for response
    debug_data = {
        "shipment": {k: v for k, v in shipment.items() if k != "attributes"},
        "account": {k: v for k, v in account.items() if k != "attributes"} if account else {},
        "container_items": [
            {k: v for k, v in item.items() if k != "attributes"}
            for item in items
        ],
        "deposits": [
            {k: v for k, v in dep.items() if k != "attributes"}
            for dep in deposits
        ],
        "refunds": [
            {k: v for k, v in ref.items() if k != "attributes"}
            for ref in refunds
        ],
        "picklist_options": {
            "Freight__c": freight_options,
            "Terms_of_Sales__c": terms_of_sales_options,
            "Terms_of_Payment__c": terms_of_payment_options,
        },
        "discount_exists": discount_exists,
        "template_used": template_path,
    }

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Invoice"] if "Invoice" in wb.sheetnames else wb.active

    # Format Port of Origin in uppercase
    port_of_origin = (shipment.get("Port_of_Origin__c") or "").upper()

    replacements = {
        "{{Shipment__c.Consignee__r.Name}}": account.get("Name") or "",
        "{{Shipment__c.Consignee__r.BillingStreet}}": account.get("BillingStreet") or "",
        "{{Shipment__c.Consignee__r.BillingCity}}": account.get("BillingCity") or "",
        "{{Shipment__c.Consignee__r.BillingPostalCode}}": account.get("BillingPostalCode") or "",
        "{{Shipment__c.Consignee__r.BillingCountry}}": account.get("BillingCountry") or "",
        "{{Shipment__c.Consignee__r.Phone}}": account.get("Phone") or "",
        "{{Shipment__c.Consignee__r.Fax__c}}": account.get("Fax__c") or "",
        "{{Shipment__c.Consignee__r.VAT__c}}": account.get("VAT__c") or "",
        "{{Shipment__c.Invoice_Packing_list_no__c}}": shipment.get("Invoice_Packing_list_no__c") or "",
        "{{Shipment__c.Issued_date__c}}": shipment.get("Issued_date__c") or "",
        "{{Shipment__c.Port_of_Origin__c}}": port_of_origin,
        "{{Shipment__c.Final_Destination__c}}": shipment.get("Final_Destination__c") or "",
        "{{Shipment__c.Stockyard__c}}": shipment.get("Stockyard__c") or "",
        "{{Shipment__c.Ocean_Vessel__c}}": shipment.get("Ocean_Vessel__c") or "",
        "{{Shipment__c.B_L_No__c}}": shipment.get("B_L_No__c") or "",
        "{{Shipment__c.Departure_Date_ETD__c}}": shipment.get("Departure_Date_ETD__c") or "",
        "{{Shipment__c.Arrival_Schedule_ETA__c}}": shipment.get("Arrival_Schedule_ETA__c") or "",
        "{{Shipment__c.Remark_number_on_documents__c}}": shipment.get("Remark_number_on_documents__c") or "",
        "{{Shipment__c.Subtotal_USD__c\\# #,##0.##}}": shipment.get("Subtotal_USD__c") or 0,
        "{{Shipment__c.Fumigation__c}}": shipment.get("Fumigation__c") or "",
        "{{Shipment__c.Total_Price_USD__c\\# #,##0.##}}": shipment.get("Total_Price_USD__c") or 0,
        "{{Shipment__c.In_words__c}}": shipment.get("In_words__c") or "",

        # 🔹 NEW: discount placeholders used by invoice_template_w_discount.xlsx
        "{{Shipment__c.Discount_Percentage__c}}": shipment.get("Discount_Percentage__c") or "",
        "{{Shipment__c.Discount_Amount__c\\# #,##0.##}}": shipment.get("Discount_Amount__c") or 0,
    }

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for placeholder, value in replacements.items():
                    cell.value = cell.value.replace(placeholder, str(value))

    # Clean "None"
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "None" in cell.value:
                cell.value = cell.value.replace("None", "")

    # 💡 FIXED: correct parameter order for checkboxes + uppercase
    freight_checkbox_text = format_picklist_checkboxes(
        freight_options, shipment.get("Freight__c"), uppercase=True
    )
    terms_of_sales_checkbox_text = format_picklist_checkboxes(
        terms_of_sales_options, shipment.get("Terms_of_Sales__c"), uppercase=True
    )
    terms_of_payment_checkbox_text = format_picklist_checkboxes(
        terms_of_payment_options, shipment.get("Terms_of_Payment__c"), uppercase=True
    )

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "{{Shipment__c.Freight__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Freight__c}}", freight_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                if "{{Shipment__c.Terms_of_Sales__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Terms_of_Sales__c}}", terms_of_sales_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                if "{{Shipment__c.Terms_of_Payment__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Terms_of_Payment__c}}", terms_of_payment_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)

    # --- ContainerItems table expansion ---
    table_start_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and "{{TableStart:ContainerItems}}" in cell.value:
                table_start_row = cell.row
                break
        if table_start_row:
            break

    if not table_start_row:
        raise ValueError("No ContainerItems table start marker found in template")

    expand_invoice_items_table(ws, table_start_row, len(items) if items else 1)

    for idx, item in enumerate(items):
        row_idx = table_start_row + idx
        container_r = item.get("Container__r") or {}
        line_item_no = item.get("Line_item_no_for_print__c") or str(idx + 1)
        ws.cell(row_idx, 1).value = line_item_no
        ws.cell(row_idx, 2).value = item.get("Product_Description__c")
        ws.cell(row_idx, 3).value = item.get("Length__c")
        ws.cell(row_idx, 4).value = item.get("Width__c")
        ws.cell(row_idx, 5).value = item.get("Height__c")
        ws.cell(row_idx, 6).value = item.get("Quantity_For_print__c")
        ws.cell(row_idx, 7).value = item.get("Unit_for_print__c")
        ws.cell(row_idx, 8).value = container_r.get("STT_Cont__c") or container_r.get("Name")
        ws.cell(row_idx, 9).value = f"{item.get('Sales_Price_USD__c') or ''} {item.get('Charge_Unit__c') or ''}".strip()
        ws.cell(row_idx, 10).value = item.get("Total_Price_USD__c")
        ws.cell(row_idx, 11).value = item.get("Order_No__c")

    # --- Deposits / refunds / surcharge sections (back to working behaviour) ---
    deposit_text_cell = None
    deposit_amount_cell = None
    refund_cell = None
    surcharge_text_cell = None
    surcharge_amount_cell = None

    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            val = cell.value
            if "{{TableStart:InvoiceDeposit}}" in val:
                deposit_text_cell = cell
            if "Reconciled_Amount__c" in val:
                deposit_amount_cell = cell
            if "{{TableStart:Shipment__c.r.Cases__r}}" in val:
                refund_cell = cell
            if "{{TableStart:Surcharges}}" in val:
                surcharge_text_cell = cell
            if "Surcharge_amount_USD__c" in val:
                surcharge_amount_cell = cell

    # Deposits: multi-line "Deduct: Deposit of PI X"
    if deposit_text_cell and deposit_amount_cell:
        if deposits:
            labels = []
            amounts = []
            for rec in deposits:
                pi_name = (rec.get("Contract_PI__r") or {}).get("Name") or ""
                labels.append(f"Deduct: Deposit of PI {pi_name}".strip())
                amt = rec.get("Reconciled_Amount__c")
                amounts.append("" if amt is None else f"{amt:,.2f}")
            deposit_text_cell.value = "\n".join(labels)
            deposit_amount_cell.value = "\n".join(amounts)
        else:
            deposit_text_cell.value = None
            deposit_amount_cell.value = None

    # Refunds section
    if refund_cell:
        if refunds:
            lines = []
            for rec in refunds:
                reason = rec.get("Reason") or ""
                amt = rec.get("Refund_Amount__c")
                part = reason
                if amt is not None:
                    part = f"{reason} {amt:,.2f}".strip()
                lines.append(part)
            refund_cell.value = "\n".join(lines)
        else:
            refund_cell.value = None

    # Clean remaining refund placeholders if no refunds
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "{{Refund_Amount__c" in cell.value or "{{TableEnd:Shipment__c.r.Cases__r}}" in cell.value:
                    if not refunds:
                        cell.value = cell.value.replace("{{Refund_Amount__c\\# #,##0.##}}", "")
                        cell.value = cell.value.replace("{{TableEnd:Shipment__c.r.Cases__r}}", "")
                        if cell.value and not cell.value.strip():
                            cell.value = None

    # Surcharge
    surcharge_amount = shipment.get("Surcharge_amount_USD__c")
    if surcharge_text_cell or surcharge_amount_cell:
        if surcharge_amount:
            if surcharge_text_cell:
                surcharge_text_cell.value = "Surcharge:"
            if surcharge_amount_cell:
                surcharge_amount_cell.value = f"{surcharge_amount:,.2f}"
        else:
            if surcharge_text_cell:
                surcharge_text_cell.value = None
            if surcharge_amount_cell:
                surcharge_amount_cell.value = None


    # Save file
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"Invoice_{shipment.get('Invoice_Packing_list_no__c', shipment['Name'])}_{timestamp}.xlsx"

    # Use appropriate output directory based on environment
    output_dir = get_output_directory()
    file_path = output_dir / file_name

    wb.save(str(file_path))

    # Upload to Salesforce as ContentVersion
    with open(file_path, "rb") as f:
        data = f.read()
    encoded = base64.b64encode(data).decode("utf-8")

    content_version = sf.ContentVersion.create(
        {
            "Title": file_name.rsplit(".", 1)[0],
            "PathOnClient": file_name,
            "VersionData": encoded,
            "FirstPublishLocationId": shipment_id,
        }
    )

    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"],
        "freight_options_used": freight_options,
        "deposit_count": len(deposits),
        "refund_count": len(refunds),
        "discount_exists": discount_exists,
        "template_used": template_path,
        "debug_data": debug_data,
    }

@app.get("/generate-combined-export/{shipment_id}")
def generate_combined_export(shipment_id: str):
    """
    Generate combined packing list and invoice in one Excel file with two sheets.
    First sheet: "Packing List"
    Second sheet: "Invoice"
    
    Parameters:
    - shipment_id: Salesforce Shipment ID
    """
    sf = get_salesforce_connection()
    
    # Templates
    packing_list_template_path = os.getenv('TEMPLATE_PATH', 'templates/packing_list_template.xlsx')
    base_invoice_template_path = "./templates/invoice_template.xlsx"
    discount_invoice_template_path = "./templates/invoice_template_w_discount.xlsx"
    
    # Verify templates exist
    if not os.path.exists(packing_list_template_path):
        raise HTTPException(
            status_code=404,
            detail=f"Packing list template not found at: {packing_list_template_path}"
        )
    
    # Get picklist values dynamically
    freight_options = get_picklist_values(sf, 'Shipment__c', 'Freight__c')
    terms_of_sales_options = get_picklist_values(sf, 'Shipment__c', 'Terms_of_Sales__c')
    terms_of_payment_options = get_picklist_values(sf, 'Shipment__c', 'Terms_of_Payment__c')
    
    # Query shipment data (combining fields from both packing list and invoice)
    shipment_query = f"""
    SELECT Name, Consignee__c, Invoice_Packing_list_no__c, Issued_date__c,
           Port_of_Origin__c, Final_Destination__c, Stockyard__c,
           Ocean_Vessel__c, B_L_No__c, Freight__c,
           Departure_Date_ETD__c, Arrival_Schedule_ETA__c,
           Remark_number_on_documents__c,
           Terms_of_Sales__c, Terms_of_Payment__c,
           Subtotal_USD__c, Fumigation__c, In_words__c,
           Total_Price_USD__c, Surcharge_amount_USD__c,
           Discount_Percentage__c, Discount_Amount__c
    FROM Shipment__c
    WHERE Id = '{shipment_id}'
    """
    shipment_result = sf.query(shipment_query)
    if not shipment_result["records"]:
        raise HTTPException(status_code=404, detail=f"No Shipment found with ID: {shipment_id}")
    shipment = shipment_result["records"][0]
    
    # Determine if discount exists
    discount_percentage = shipment.get("Discount_Percentage__c")
    discount_amount = shipment.get("Discount_Amount__c")
    discount_exists = any(
        v not in (None, 0, "", "0", 0.0)
        for v in (discount_percentage, discount_amount)
    )
    invoice_template_path = discount_invoice_template_path if discount_exists else base_invoice_template_path
    
    # Query account/consignee data
    if shipment.get("Consignee__c"):
        account_query = f"""
        SELECT Name, BillingStreet, BillingCity, BillingPostalCode, BillingCountry,
               Phone, Fax__c, VAT__c
        FROM Account
        WHERE Id = '{shipment['Consignee__c']}'
        """
        account_result = sf.query(account_query)
        account = account_result["records"][0] if account_result["records"] else {}
    else:
        account = {}
    
    # Query bookings (for packing list)
    bookings_query = f"""
    SELECT Id, Cont_Quantity__c
    FROM Booking__c
    WHERE Shipment__c = '{shipment_id}'
    """
    bookings_result = sf.query_all(bookings_query)
    bookings = bookings_result['records']
    total_containers_from_bookings = sum(booking.get('Cont_Quantity__c') or 0 for booking in bookings)
    
    # Query container items (both packing list and invoice use this)
    items_query = f"""
    SELECT Line_item_no_for_print__c, Product_Description__c,
           Length__c, Width__c, Height__c,
           Quantity_For_print__c, Unit_for_print__c,
           Crates__c, Packing__c, Order_No__c,
           Sales_Price_USD__c, Charge_Unit__c, Total_Price_USD__c,
           Container__r.Name, Container__r.Container_Weight_Regulation__c,
           Container__r.STT_Cont__c
    FROM Container_Item__c
    WHERE Shipment__c = '{shipment_id}'
    ORDER BY Line_item_no_for_print__c
    """
    items_result = sf.query_all(items_query)
    items = items_result["records"]
    
    # Query deposits (for invoice)
    deposit_query = f"""
    SELECT Contract_PI__r.Name, Reconciled_Amount__c
    FROM Receipt_Reconciliation__c
    WHERE Invoice__c = '{shipment_id}'
    """
    deposits = sf.query_all(deposit_query)["records"]
    
    # Query refunds (for invoice)
    refunds_query = f"""
    SELECT Reason, Refund_Amount__c
    FROM Case
    WHERE Refund_in_Shipment__c = '{shipment_id}'
    """
    refunds = sf.query_all(refunds_query)["records"]
    
    # ===== GENERATE PACKING LIST SHEET =====
    wb_packing = openpyxl.load_workbook(packing_list_template_path)
    ws_packing = wb_packing['PackingList']
    
    # Packing list replacements
    packing_replacements = {
        '{{Shipment__c.Consignee__r.Name}}': account.get('Name') or '',
        '{{Shipment__c.Consignee__r.BillingStreet}}': account.get('BillingStreet') or '',
        '{{Shipment__c.Consignee__r.BillingCity}}': account.get('BillingCity') or '',
        '{{Shipment__c.Consignee__r.BillingPostalCode}}': account.get('BillingPostalCode') or '',
        '{{Shipment__c.Consignee__r.BillingCountry}}': account.get('BillingCountry') or '',
        '{{Shipment__c.Consignee__r.Phone}}': account.get('Phone') or '',
        '{{Shipment__c.Consignee__r.Fax__c}}': account.get('Fax__c') or '',
        '{{Shipment__c.Consignee__r.VAT__c}}': account.get('VAT__c') or '',
        '{{Shipment__c.Invoice_Packing_list_no__c}}': shipment.get('Invoice_Packing_list_no__c') or '',
        '{{Shipment__c.Issued_date__c}}': shipment.get('Issued_date__c') or '',
        '{{Shipment__c.Port_of_Origin__c}}': shipment.get('Port_of_Origin__c') or '',
        '{{Shipment__c.Final_Destination__c}}': shipment.get('Final_Destination__c') or '',
        '{{Shipment__c.Stockyard__c}}': shipment.get('Stockyard__c') or '',
        '{{Shipment__c.Ocean_Vessel__c}}': shipment.get('Ocean_Vessel__c') or '',
        '{{Shipment__c.B_L_No__c}}': shipment.get('B_L_No__c') or '',
        '{{Shipment__c.Departure_Date_ETD__c}}': shipment.get('Departure_Date_ETD__c') or '',
        '{{Shipment__c.Arrival_Schedule_ETA__c}}': shipment.get('Arrival_Schedule_ETA__c') or '',
        '{{Shipment__c.Remark_number_on_documents__c}}': shipment.get('Remark_number_on_documents__c') or '',
        '{{Shipment__c.Terms_of_Sales__c}}': shipment.get('Terms_of_Sales__c') or '',
        '{{Shipment__c.Terms_of_Payment__c}}': shipment.get('Terms_of_Payment__c') or '',
    }
    
    for row in ws_packing.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for placeholder, value in packing_replacements.items():
                    cell.value = cell.value.replace(placeholder, str(value))
                if '{{TableStart:Shipment__c.r.Bookings__r}}' in cell.value:
                    cell.value = str(total_containers_from_bookings)
    
    # Remove "None" values
    for row in ws_packing.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and 'None' in cell.value:
                cell.value = cell.value.replace('None', '')
    
    # Handle freight checkboxes for packing list
    freight_value = (shipment.get('Freight__c') or '').strip()
    freight_upper = freight_value.upper()
    checked_box = '☑'
    unchecked_box = '☐'
    
    lines = []
    for opt in freight_options:
        mark = checked_box if opt.upper() == freight_upper else unchecked_box
        lines.append(f"{mark} {opt}")
    checkbox_text = "\n".join(lines)
    
    for row in ws_packing.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '{{Shipment__c.Freight__c}}' in cell.value:
                cell.value = cell.value.replace('{{Shipment__c.Freight__c}}', checkbox_text)
                if cell.alignment:
                    new_alignment = style_copy(cell.alignment)
                else:
                    from openpyxl.styles import Alignment
                    new_alignment = Alignment()
                new_alignment.wrap_text = True
                cell.alignment = new_alignment
    
    # Find table start row for packing list
    table_start_row = None
    for row in ws_packing.iter_rows(min_row=1, max_row=ws_packing.max_row, min_col=1, max_col=13):
        for cell in row:
            if cell.value and '{{TableStart:ContainerItems}}' in str(cell.value):
                table_start_row = cell.row
                break
        if table_start_row:
            break
    
    if not table_start_row:
        raise ValueError("No table start marker found in packing list template")
    
    # Expand table for packing list
    expand_items_table(ws_packing, table_start_row, len(items) if items else 1)
    
    # Fill in item data for packing list
    for idx, item in enumerate(items):
        row = table_start_row + idx
        container_r = item.get('Container__r', {})
        line_item_no = item.get('Line_item_no_for_print__c') or str(idx + 1)
        ws_packing.cell(row, 1).value = line_item_no
        ws_packing.cell(row, 2).value = item.get('Product_Description__c')
        ws_packing.cell(row, 3).value = item.get('Length__c')
        ws_packing.cell(row, 4).value = item.get('Width__c')
        ws_packing.cell(row, 5).value = item.get('Height__c')
        ws_packing.cell(row, 6).value = item.get('Quantity_For_print__c') or ''
        ws_packing.cell(row, 7).value = item.get('Unit_for_print__c') or ''
        ws_packing.cell(row, 8).value = item.get('Crates__c')
        ws_packing.cell(row, 9).value = f"{item.get('Packing__c') or ''} pcs/crate"
        ws_packing.cell(row, 10).value = container_r.get('Container_Weight_Regulation__c')
        ws_packing.cell(row, 11).value = container_r.get('Name')
        ws_packing.cell(row, 13).value = item.get('Order_No__c')
    
    # ===== GENERATE INVOICE SHEET =====
    wb_invoice = openpyxl.load_workbook(invoice_template_path)
    ws_invoice = wb_invoice["Invoice"] if "Invoice" in wb_invoice.sheetnames else wb_invoice.active
    
    # Format Port of Origin in uppercase
    port_of_origin = (shipment.get("Port_of_Origin__c") or "").upper()
    
    # Invoice replacements
    invoice_replacements = {
        "{{Shipment__c.Consignee__r.Name}}": account.get("Name") or "",
        "{{Shipment__c.Consignee__r.BillingStreet}}": account.get("BillingStreet") or "",
        "{{Shipment__c.Consignee__r.BillingCity}}": account.get("BillingCity") or "",
        "{{Shipment__c.Consignee__r.BillingPostalCode}}": account.get("BillingPostalCode") or "",
        "{{Shipment__c.Consignee__r.BillingCountry}}": account.get("BillingCountry") or "",
        "{{Shipment__c.Consignee__r.Phone}}": account.get("Phone") or "",
        "{{Shipment__c.Consignee__r.Fax__c}}": account.get("Fax__c") or "",
        "{{Shipment__c.Consignee__r.VAT__c}}": account.get("VAT__c") or "",
        "{{Shipment__c.Invoice_Packing_list_no__c}}": shipment.get("Invoice_Packing_list_no__c") or "",
        "{{Shipment__c.Issued_date__c}}": shipment.get("Issued_date__c") or "",
        "{{Shipment__c.Port_of_Origin__c}}": port_of_origin,
        "{{Shipment__c.Final_Destination__c}}": shipment.get("Final_Destination__c") or "",
        "{{Shipment__c.Stockyard__c}}": shipment.get("Stockyard__c") or "",
        "{{Shipment__c.Ocean_Vessel__c}}": shipment.get("Ocean_Vessel__c") or "",
        "{{Shipment__c.B_L_No__c}}": shipment.get("B_L_No__c") or "",
        "{{Shipment__c.Departure_Date_ETD__c}}": shipment.get("Departure_Date_ETD__c") or "",
        "{{Shipment__c.Arrival_Schedule_ETA__c}}": shipment.get("Arrival_Schedule_ETA__c") or "",
        "{{Shipment__c.Remark_number_on_documents__c}}": shipment.get("Remark_number_on_documents__c") or "",
        "{{Shipment__c.Subtotal_USD__c\\# #,##0.##}}": shipment.get("Subtotal_USD__c") or 0,
        "{{Shipment__c.Fumigation__c}}": shipment.get("Fumigation__c") or "",
        "{{Shipment__c.Total_Price_USD__c\\# #,##0.##}}": shipment.get("Total_Price_USD__c") or 0,
        "{{Shipment__c.In_words__c}}": shipment.get("In_words__c") or "",
        "{{Shipment__c.Discount_Percentage__c}}": shipment.get("Discount_Percentage__c") or "",
        "{{Shipment__c.Discount_Amount__c\\# #,##0.##}}": shipment.get("Discount_Amount__c") or 0,
    }
    
    for row in ws_invoice.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for placeholder, value in invoice_replacements.items():
                    cell.value = cell.value.replace(placeholder, str(value))
    
    # Clean "None"
    for row in ws_invoice.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "None" in cell.value:
                cell.value = cell.value.replace("None", "")
    
    # Format picklist fields with checkboxes (uppercase)
    freight_checkbox_text = format_picklist_checkboxes(
        freight_options, shipment.get("Freight__c"), uppercase=True
    )
    terms_of_sales_checkbox_text = format_picklist_checkboxes(
        terms_of_sales_options, shipment.get("Terms_of_Sales__c"), uppercase=True
    )
    terms_of_payment_checkbox_text = format_picklist_checkboxes(
        terms_of_payment_options, shipment.get("Terms_of_Payment__c"), uppercase=True
    )
    
    for row in ws_invoice.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "{{Shipment__c.Freight__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Freight__c}}", freight_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                if "{{Shipment__c.Terms_of_Sales__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Terms_of_Sales__c}}", terms_of_sales_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                if "{{Shipment__c.Terms_of_Payment__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Terms_of_Payment__c}}", terms_of_payment_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
    
    # Find ContainerItems table for invoice
    invoice_table_start_row = None
    for row in ws_invoice.iter_rows(min_row=1, max_row=ws_invoice.max_row, min_col=1, max_col=ws_invoice.max_column):
        for cell in row:
            if isinstance(cell.value, str) and "{{TableStart:ContainerItems}}" in cell.value:
                invoice_table_start_row = cell.row
                break
        if invoice_table_start_row:
            break
    
    if not invoice_table_start_row:
        raise ValueError("No ContainerItems table start marker found in invoice template")
    
    expand_invoice_items_table(ws_invoice, invoice_table_start_row, len(items) if items else 1)
    
    for idx, item in enumerate(items):
        row_idx = invoice_table_start_row + idx
        container_r = item.get("Container__r") or {}
        line_item_no = item.get("Line_item_no_for_print__c") or str(idx + 1)
        ws_invoice.cell(row_idx, 1).value = line_item_no
        ws_invoice.cell(row_idx, 2).value = item.get("Product_Description__c")
        ws_invoice.cell(row_idx, 3).value = item.get("Length__c")
        ws_invoice.cell(row_idx, 4).value = item.get("Width__c")
        ws_invoice.cell(row_idx, 5).value = item.get("Height__c")
        ws_invoice.cell(row_idx, 6).value = item.get("Quantity_For_print__c")
        ws_invoice.cell(row_idx, 7).value = item.get("Unit_for_print__c")
        ws_invoice.cell(row_idx, 8).value = container_r.get("STT_Cont__c") or container_r.get("Name")
        ws_invoice.cell(row_idx, 9).value = f"{item.get('Sales_Price_USD__c') or ''} {item.get('Charge_Unit__c') or ''}".strip()
        ws_invoice.cell(row_idx, 10).value = item.get("Total_Price_USD__c")
        ws_invoice.cell(row_idx, 11).value = item.get("Order_No__c")
    
    # Handle deposits / refunds / surcharge sections
    deposit_text_cell = None
    deposit_amount_cell = None
    refund_cell = None
    surcharge_text_cell = None
    surcharge_amount_cell = None
    
    for row in ws_invoice.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            val = cell.value
            if "{{TableStart:InvoiceDeposit}}" in val:
                deposit_text_cell = cell
            if "Reconciled_Amount__c" in val:
                deposit_amount_cell = cell
            if "{{TableStart:Shipment__c.r.Cases__r}}" in val:
                refund_cell = cell
            if "{{TableStart:Surcharges}}" in val:
                surcharge_text_cell = cell
            if "Surcharge_amount_USD__c" in val:
                surcharge_amount_cell = cell
    
    if deposit_text_cell and deposit_amount_cell:
        if deposits:
            labels = []
            amounts = []
            for rec in deposits:
                pi_name = (rec.get("Contract_PI__r") or {}).get("Name") or ""
                labels.append(f"Deduct: Deposit of PI {pi_name}".strip())
                amt = rec.get("Reconciled_Amount__c")
                amounts.append("" if amt is None else f"{amt:,.2f}")
            deposit_text_cell.value = "\n".join(labels)
            deposit_amount_cell.value = "\n".join(amounts)
        else:
            deposit_text_cell.value = None
            deposit_amount_cell.value = None
    
    if refund_cell:
        if refunds:
            lines = []
            for rec in refunds:
                reason = rec.get("Reason") or ""
                amt = rec.get("Refund_Amount__c")
                part = reason
                if amt is not None:
                    part = f"{reason} {amt:,.2f}".strip()
                lines.append(part)
            refund_cell.value = "\n".join(lines)
        else:
            refund_cell.value = None
    
    for row in ws_invoice.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "{{Refund_Amount__c" in cell.value or "{{TableEnd:Shipment__c.r.Cases__r}}" in cell.value:
                    if not refunds:
                        cell.value = cell.value.replace("{{Refund_Amount__c\\# #,##0.##}}", "")
                        cell.value = cell.value.replace("{{TableEnd:Shipment__c.r.Cases__r}}", "")
                        if cell.value and not cell.value.strip():
                            cell.value = None
    
    surcharge_amount = shipment.get("Surcharge_amount_USD__c")
    if surcharge_text_cell or surcharge_amount_cell:
        if surcharge_amount:
            if surcharge_text_cell:
                surcharge_text_cell.value = "Surcharge:"
            if surcharge_amount_cell:
                surcharge_amount_cell.value = f"{surcharge_amount:,.2f}"
        else:
            if surcharge_text_cell:
                surcharge_text_cell.value = None
            if surcharge_amount_cell:
                surcharge_amount_cell.value = None
    
    # ===== COMBINE INTO ONE WORKBOOK =====
    # Create a new workbook and copy sheets
    combined_wb = openpyxl.Workbook()
    combined_wb.remove(combined_wb.active)  # Remove default sheet
    
    # Copy packing list sheet
    ws_packing_copy = combined_wb.create_sheet("Packing List")
    for row in ws_packing.iter_rows():
        for cell in row:
            new_cell = ws_packing_copy.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = style_copy(cell.font)
                new_cell.border = style_copy(cell.border)
                new_cell.fill = style_copy(cell.fill)
                new_cell.number_format = style_copy(cell.number_format)
                new_cell.protection = style_copy(cell.protection)
                new_cell.alignment = style_copy(cell.alignment)
    
    # Copy column dimensions
    for col in ws_packing.column_dimensions:
        if col in ws_packing.column_dimensions:
            combined_wb["Packing List"].column_dimensions[col].width = ws_packing.column_dimensions[col].width
    
    # Copy row dimensions
    for row in ws_packing.row_dimensions:
        if row in ws_packing.row_dimensions:
            combined_wb["Packing List"].row_dimensions[row].height = ws_packing.row_dimensions[row].height
    
    # Copy merged cells
    for merged_cell in ws_packing.merged_cells.ranges:
        combined_wb["Packing List"].merge_cells(str(merged_cell))
    
    # Copy invoice sheet
    ws_invoice_copy = combined_wb.create_sheet("Invoice")
    for row in ws_invoice.iter_rows():
        for cell in row:
            new_cell = ws_invoice_copy.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = style_copy(cell.font)
                new_cell.border = style_copy(cell.border)
                new_cell.fill = style_copy(cell.fill)
                new_cell.number_format = style_copy(cell.number_format)
                new_cell.protection = style_copy(cell.protection)
                new_cell.alignment = style_copy(cell.alignment)
    
    # Copy column dimensions
    for col in ws_invoice.column_dimensions:
        if col in ws_invoice.column_dimensions:
            combined_wb["Invoice"].column_dimensions[col].width = ws_invoice.column_dimensions[col].width
    
    # Copy row dimensions
    for row in ws_invoice.row_dimensions:
        if row in ws_invoice.row_dimensions:
            combined_wb["Invoice"].row_dimensions[row].height = ws_invoice.row_dimensions[row].height
    
    # Copy merged cells
    for merged_cell in ws_invoice.merged_cells.ranges:
        combined_wb["Invoice"].merge_cells(str(merged_cell))
    
    
    # Save combined file
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"Combined_Export_{shipment.get('Invoice_Packing_list_no__c', shipment['Name'])}_{timestamp}.xlsx"
    
    # Use appropriate output directory based on environment
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    
    combined_wb.save(str(file_path))
    
    # Upload to Salesforce as ContentVersion
    with open(file_path, "rb") as f:
        data = f.read()
    encoded = base64.b64encode(data).decode("utf-8")
    
    content_version = sf.ContentVersion.create(
        {
            "Title": file_name.rsplit(".", 1)[0],
            "PathOnClient": file_name,
            "VersionData": encoded,
            "FirstPublishLocationId": shipment_id,
        }
    )
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"],
        "sheets": ["Packing List", "Invoice"],
        "item_count": len(items),
        "deposit_count": len(deposits),
        "refund_count": len(refunds),
        "discount_exists": discount_exists,
        "template_used": {
            "packing_list": packing_list_template_path,
            "invoice": invoice_template_path
        }
    }

@app.get("/download/{file_name}")
async def download_file(file_name: str):
    """
    Download a generated packing list file
    
    Parameters:
    - file_name: Name of the file to download
    
    Note: In serverless environments (Vercel), files in /tmp are ephemeral.
    The download endpoint may not work reliably. Files are always uploaded to Salesforce.
    """
    # Try to find the file in the appropriate output directory
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    
    # Also check the legacy output directory for backwards compatibility
    if not file_path.exists():
        legacy_path = Path("output") / file_name
        if legacy_path.exists():
            file_path = legacy_path
    
    if not file_path.exists():
        raise HTTPException(
            status_code=404, 
            detail=f"File not found. In serverless environments, use the Salesforce attachment instead."
        )
    
    return FileResponse(
        path=str(file_path),
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --- New Helper Functions for PI, PO, Quote ---

def sanitize_filename(name):
    """
    Sanitize filename by removing or replacing invalid characters.
    """
    if not name:
        return "Unknown"
    # Replace invalid characters with underscore
    return re.sub(r'[<>:"/\\|?*]', '_', str(name))

def expand_table_by_tag(ws, start_tag, end_tag, data):
    """
    Expand a single row table based on start and end tags.
    Matches logic from test_fill_pi_no_discount.py
    """
    # Find the row containing the tags
    table_row_idx = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if start_tag in cell.value:
                    table_row_idx = cell.row
                    break
        if table_row_idx:
            break
            
    if not table_row_idx:
        print(f"Warning: Table tags {start_tag} not found.")
        return None

    if not data:
        # Clear tags and placeholders, keep static text
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=table_row_idx, column=col)
            if cell.value and isinstance(cell.value, str):
                # Remove tags
                val = cell.value.replace(start_tag, "").replace(end_tag, "")
                # Remove any remaining placeholders {{...}}
                val = re.sub(r"\{\{.*?\}\}", "", val)
                cell.value = val
        return table_row_idx

    num_rows = len(data)
    add_rows = max(0, num_rows - 1)
    
    # Capture styles from the template row
    max_col = ws.max_column
    row_style = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=table_row_idx, column=col)
        row_style.append(style_copy(cell._style) if cell.has_style else None)
    
    row_height = ws.row_dimensions[table_row_idx].height

    # Handle merged cells (shift them down)
    merges_to_shift = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > table_row_idx:
            merges_to_shift.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))
    
    for mr in merges_to_shift:
        rng = f"{get_column_letter(mr[2])}{mr[0]}:{get_column_letter(mr[3])}{mr[1]}"
        ws.unmerge_cells(rng)

    # Insert rows if needed
    if add_rows > 0:
        ws.insert_rows(table_row_idx + 1, amount=add_rows)
        
        for offset in range(1, add_rows + 1):
            r = table_row_idx + offset
            # Copy row height
            if row_height is not None:
                ws.row_dimensions[r].height = row_height
                
            for col in range(1, max_col + 1):
                dst = ws.cell(row=r, column=col)
                # Copy value from template row (to preserve placeholders)
                src_val = ws.cell(row=table_row_idx, column=col).value
                dst.value = src_val
                
                # Copy style
                st = row_style[col - 1]
                if st is not None:
                    dst._style = style_copy(st)
                    
    # Re-merge shifted cells
    for mr in merges_to_shift:
        new_min_row = mr[0] + add_rows
        new_max_row = mr[1] + add_rows
        rng = f"{get_column_letter(mr[2])}{new_min_row}:{get_column_letter(mr[3])}{new_max_row}"
        ws.merge_cells(rng)
                    
    # Fill data
    for i, record in enumerate(data):
        current_row_idx = table_row_idx + i
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=current_row_idx, column=col)
            if cell.value and isinstance(cell.value, str):
                cell_val = cell.value.replace(start_tag, "").replace(end_tag, "")
                
                for key, value in record.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in cell_val:
                        cell_val = cell_val.replace(placeholder, str(value) if value is not None else "")
                        
                    pattern = f"\\{{{{{key}\\\\#(.*?)\\}}}}"
                    matches = re.findall(pattern, cell_val)
                    for fmt in matches:
                        try:
                            if value is not None:
                                if isinstance(value, (int, float)):
                                    if "#,##0.##" in fmt:
                                         formatted_val = "{:,.2f}".format(value)
                                    else:
                                         formatted_val = str(value)
                                    cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", formatted_val)
                                else:
                                    cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value))
                            else:
                                cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", "")
                        except:
                             cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value))

                cell.value = cell_val
                
                # Attempt to convert to number if it looks like one
                if isinstance(cell.value, str):
                    try:
                        clean_val = cell.value.replace(',', '')
                        f_val = float(clean_val)
                        is_leading_zero = (len(clean_val) > 1 and clean_val.startswith('0') and not clean_val.startswith('0.'))
                        
                        if not is_leading_zero:
                            if f_val.is_integer():
                                cell.value = int(f_val)
                            else:
                                cell.value = f_val
                    except ValueError:
                        pass
    
    return table_row_idx

def apply_bold_formatting(ws, start_row, data, col_idx, key_name='Product__r'):
    """
    Apply bold formatting to product description based on product name.
    """
    if not start_row or not data:
        return

    for i, item in enumerate(data):
        row_idx = start_row + i
        cell = ws.cell(row=row_idx, column=col_idx)
        
        product_name = None
        if key_name == 'Product__r':
            product_name = item.get('Product__r', {}).get('Name')
        else:
            product_name = item.get(key_name) # Direct access for Quote/Order items

        current_desc = str(cell.value) if cell.value else ""
        
        if product_name and current_desc:
            match = re.match(r"^([^\d\(]+)", product_name)
            if match:
                bold_target = match.group(1).strip()
                if bold_target and bold_target in current_desc:
                    start_idx = current_desc.find(bold_target)
                    if start_idx != -1:
                        parts = []
                        if start_idx > 0: parts.append(current_desc[:start_idx])
                        parts.append(TextBlock(InlineFont(b=True), bold_target))
                        end_idx = start_idx + len(bold_target)
                        if end_idx < len(current_desc): parts.append(current_desc[end_idx:])
                        cell.value = CellRichText(parts)

def merge_identical_cells(ws, start_row, count, col_idx):
    """
    Merge identical cells in a column and adjust row heights.
    """
    if not start_row or count <= 0:
        return

    start_merge_row = start_row
    current_val = str(ws.cell(row=start_merge_row, column=col_idx).value)
    
    for i in range(1, count):
        row_idx = start_row + i
        cell_val = str(ws.cell(row=row_idx, column=col_idx).value)
        
        if cell_val != current_val:
            if row_idx - 1 > start_merge_row:
                ws.merge_cells(start_row=start_merge_row, start_column=col_idx, end_row=row_idx-1, end_column=col_idx)
                ws.cell(row=start_merge_row, column=col_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                adjust_row_height_for_merged_cell(ws, start_merge_row, row_idx-1, col_idx, current_val)
            start_merge_row = row_idx
            current_val = cell_val
            
    # Last block
    last_row = start_row + count - 1
    if last_row > start_merge_row:
        ws.merge_cells(start_row=start_merge_row, start_column=col_idx, end_row=last_row, end_column=col_idx)
        ws.cell(row=start_merge_row, column=col_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        adjust_row_height_for_merged_cell(ws, start_merge_row, last_row, col_idx, current_val)

def adjust_row_height_for_merged_cell(ws, start_row, end_row, col_idx, text, line_height_base=25):
    """
    Helper to adjust row height for merged cells.
    """
    val_str = str(text) if text else ""
    text_len = len(val_str)
    
    col_letter = get_column_letter(col_idx)
    col_width = ws.column_dimensions[col_letter].width
    if not col_width: col_width = 30
    
    chars_per_line = int(col_width * 1.2)
    if chars_per_line < 10: chars_per_line = 30
    
    explicit_lines = val_str.count('\n') + 1
    wrap_lines = (text_len // chars_per_line) + 1
    estimated_lines = max(explicit_lines, wrap_lines)
    
    if estimated_lines > 1:
        required_height = estimated_lines * line_height_base
    else:
        required_height = 30
    required_height += 10
    
    current_total_height = 0
    for r in range(start_row, end_row + 1):
        h = ws.row_dimensions[r].height
        if h is None: h = 15
        current_total_height += h
    
    if required_height > current_total_height:
        extra_per_row = (required_height - current_total_height) / (end_row - start_row + 1)
        for r in range(start_row, end_row + 1):
            h = ws.row_dimensions[r].height
            if h is None: h = 15
            ws.row_dimensions[r].height = h + extra_per_row
# --- Helper: Number to Words (English USD) ---
def number_to_text(n):
    if n < 0:
        return "Minus " + number_to_text(-n)
    if n == 0:
        return "Zero"

    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
    teens = ["", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "Ten", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    thousands = ["", "Thousand", "Million", "Billion"]

    def convert_chunk(n):
        if n == 0: return ""
        if n < 10: return units[n]
        if n < 20: return teens[n - 10] if n > 10 else tens[n // 10]
        if n < 100:
            return tens[n // 10] + (" " + units[n % 10] if n % 10 > 0 else "")
        return units[n // 100] + " Hundred" + (" " + convert_chunk(n % 100) if n % 100 > 0 else "")

    parts = []
    chunk_count = 0
    while n > 0:
        chunk = n % 1000
        if chunk > 0:
            part = convert_chunk(chunk)
            if chunk_count > 0:
                part += " " + thousands[chunk_count]
            parts.insert(0, part)
        n //= 1000
        chunk_count += 1

    return " ".join(parts)

def amount_to_words_usd(amount):
    try:
        amount = float(amount)
    except:
        return ""
        
    dollars = int(amount)
    cents = int(round((amount - dollars) * 100))
    
    text = number_to_text(dollars) + " US Dollars"
    if cents > 0:
        text += " And " + number_to_text(cents) + " Cents"
    
    return text.strip() + " Only"

# --- PI No Discount Generation ---

def generate_pi_no_discount_file(contract_id: str, template_path: str):
    sf = get_salesforce_connection()
    
    # Get picklist values for PI
    incoterms_options = get_picklist_values(sf, 'Contract__c', 'Incoterms__c')
    terms_of_sale_options = get_picklist_values(sf, 'Contract__c', 'Terms_of_Sale__c')
    terms_of_payment_options = get_picklist_values(sf, 'Contract__c', 'Terms_of_Payment__c')
    
    # Query Contract (Full Query from reference script)
    contract_query = f"""
    SELECT Id, IsDeleted, Name, CreatedDate, LastModifiedDate, SystemModstamp, LastActivityDate, LastViewedDate, LastReferencedDate, Cont__c, Container_Weight_Regulations__c, Crates__c, Height__c, Length__c, Line_Number__c, Packing__c, Sales_Price__c, Tons__c, Width__c, List_Price__c, Discount__c, Charge_Unit__c, Quantity__c, m2__c, m3__c, ml__c, Total_Price_USD__c, L_PI__c, W_PI__c, H_PI__c, PCS_PI__c, Crates_PI__c, Created_Date__c, Packing_PI__c, Product_Discription__c, Charge_Unit_PI__c, Actual_Cont__c, Pending_Cont__c, Clear__c, Actual_Crates__c, Actual_m2__c, Actual_m3__c, Actual_ml__c, Actual_Quantity__c, Actual_Tons__c, Actual_Total_Price_USD__c, Pending_Crates__c, Pending_m2__c, Pending_m3__c, Pending_ml__c, Pending_Quantity__c, Pending_Tons__c, Pending_Amount_USD__c, Delivery_Date__c, Delivery_Quantity__c, Is_Delivery_Quantity_Valid__c, Delivery_Quantity_number__c, Unscheduled_Quantity__c, Line_number_For_print__c, Product__r.Id, Product__r.Name, Product__r.ProductCode, Product__r.Description, Product__r.QuantityScheduleType, Product__r.QuantityInstallmentPeriod, Product__r.NumberOfQuantityInstallments, Product__r.RevenueScheduleType, Product__r.RevenueInstallmentPeriod, Product__r.NumberOfRevenueInstallments, Product__r.IsActive, Product__r.CreatedDate, Product__r.CreatedById, Product__r.LastModifiedDate, Product__r.LastModifiedById, Product__r.SystemModstamp, Product__r.Family, Product__r.ExternalDataSourceId, Product__r.ExternalId, Product__r.DisplayUrl, Product__r.QuantityUnitOfMeasure, Product__r.IsDeleted, Product__r.IsArchived, Product__r.LastViewedDate, Product__r.LastReferencedDate, Product__r.StockKeepingUnit, Product__r.Product_description_in_Vietnamese__c, Product__r.specific_gravity__c, Product__r.Bottom_cladding_coefficient__c, Product__r.STONE_Color_Type__c, Product__r.Packing__c, Product__r.Long__c, Product__r.High__c, Product__r.Width__c, Product__r.Long_special__c, Product__r.High_special__c, Product__r.Image__c, Product__r.Charge_Unit__c, Product__r.Width_special__c, Product__r.STONE_Class__c, Product__r.Description__c, Product__r.List_Price__c, Product__r.Weight_per_unit__c, Product__r.Edge_Finish__c, Product__r.Suppliers__c, Product__r.m_per_unit__c, Product__r.Application__c, Product__r.Surface_Finish__c, Product__r.m3_per_unit__c, Product__r.Pricing_Method__c, Contract__r.Id, Contract__r.OwnerId, Contract__r.IsDeleted, Contract__r.Name, Contract__r.CreatedDate, Contract__r.CreatedById, Contract__r.LastModifiedDate, Contract__r.LastModifiedById, Contract__r.SystemModstamp, Contract__r.LastActivityDate, Contract__r.LastViewedDate, Contract__r.LastReferencedDate, Contract__r.Account__c, Contract__r.Quote__c, Contract__r.Bill_To__c, Contract__r.Bill_To_Name__c, Contract__r.Contact_Name__c, Contract__r.Expiration_Date__c, Contract__r.Export_Route_Carrier__c, Contract__r.Fax__c, Contract__r.Phone__c, Contract__r.Fumigation__c, Contract__r.Incoterms__c, Contract__r.In_words__c, Contract__r.Packing__c, Contract__r.Port_of_Discharge__c, Contract__r.REMARK_NUMBER_ON_DOCUMENTS__c, Contract__r.Shipping_Schedule__c, Contract__r.Total_Conts__c, Contract__r.Total_Crates__c, Contract__r.Total_m3__c, Contract__r.Sub_Total_USD__c, Contract__r.Total_Tons__c, Contract__r.Deposit_Percentage__c, Contract__r.Discount__c, Contract__r.Total_Price_USD__c, Contract__r.Deposit__c, Contract__r.Stage__c, Contract__r.Total_Payment_Received__c, Contract__r.Expected_ETD__c, Contract__r.Port_of_Origin__c, Contract__r.Price_Book__c, Contract__r.Stockyard__c, Contract__r.Created_Date__c, Contract__r.Total_Contract_Product__c, Contract__r.Pending_Products__c, Contract__r.Total_Payment_Received_USD__c, Contract__r.Production_Order_Number__c, Contract__r.Total_m2__c, Contract__r.Total_Pcs__c, Contract__r.Total_Pcs_PO__c, Contract__r.Planned_Shipments__c, Contract__r.Is_approved__c, Contract__r.Deposited_amount_USD__c, Contract__r.Design_confirmed__c, Contract__r.Contract_type__c, Contract__r.Fully_deposited__c, Contract__r.Discount_Amount__c, Contract__r.Terms_of_Payment__c,    Contract__r.Terms_of_Sale__c, Contract__r.Total_surcharge__c, Contract__r.Customer_PO_number__c FROM Contract_Product__c where Contract__r.Id = '{contract_id}' ORDER BY Line_Number__c ASC
    """
    
    try:
        result = sf.query_all(contract_query)
    except Exception as e:
        print(f"Error querying contract: {e}")
        raise ValueError(f"Error querying contract: {e}")

    if not result['records']:
        raise ValueError(f"No contract items found for ID: {contract_id}")

    contract_items = result['records']
    first_item = contract_items[0]
    if 'Contract__r' in first_item and first_item['Contract__r']:
        contract = first_item['Contract__r']
    else:
        raise ValueError("Contract data missing in line items.")
    
    # Flatten Data
    full_data = {}
    for k, v in contract.items():
        full_data[f"Contract__c.{k}"] = v
        
    # Fetch Account
    account_id = contract.get('Account__c')
    if account_id:
        acc_fields = ["Name", "BillingStreet", "BillingCity", "BillingPostalCode", "BillingCountry", "Phone", "Fax__c", "VAT__c"]
        try:
            acc = sf.Account.get(account_id)
            for k in acc_fields:
                full_data[f"Contract__c.Account__r.{k}"] = acc.get(k)
        except: pass

    # Inject Sequential Number
    for idx, item in enumerate(contract_items):
        item['Line_number_For_print__c'] = idx + 1

    # Query Surcharges
    surcharge_query = f"SELECT Id, Name, Surcharge_amount_USD__c FROM Expense__c WHERE Contract_PI__r.Id = '{contract_id}' AND Surcharge_amount_USD__c != 0"
    try:
        sur_result = sf.query_all(surcharge_query)
        surcharge_records = sur_result['records']
    except Exception as e:
        surcharge_records = []
        
    surcharge_items = []
    for item in surcharge_records:
        surcharge_items.append({
            "Name": item.get('Name'),
            "Surcharge_amount_USD__c": item.get('Surcharge_amount_USD__c')
        })

    # Query Deposits (Receipt_Reconciliation__c)
    deposit_query = f"SELECT Id, Name, Reconciled_Amount__c, Contract_PI__r.Name FROM Receipt_Reconciliation__c WHERE Contract_PI__r.Id = '{contract_id}'"
    try:
        dep_result = sf.query_all(deposit_query)
        deposit_records = dep_result['records']
    except Exception as e:
        deposit_records = []
    
    deposit_items = []
    for item in deposit_records:
        deposit_items.append({
            "Name": item.get('Name'),
            "Reconciled_Amount__c": item.get('Reconciled_Amount__c'),
            "Contract_PI__r.Name": item.get('Contract_PI__r', {}).get('Name')
        })

    # Query Discounts (Discount_Item__c - Placeholder)
    discount_items = []
    try:
        discount_query = f"SELECT Id, Name, Discount_Amount__c FROM Discount_Item__c WHERE Contract_PI__r.Id = '{contract_id}'"
        disc_result = sf.query_all(discount_query)
        discount_records = disc_result['records']
        for item in discount_records:
            val = item.get('Discount_Amount__c')
            if val is not None:
                try: val = float(val)
                except: pass
            discount_items.append({
                "Name": item.get('Name'),
                "Discount_Amount__c": val
            })
    except Exception:
        pass # Ignore if object doesn't exist

    # Determine Template based on Discount
    discount_val = contract.get('Discount__c')
    discount_amt = contract.get('Discount_Amount__c')
    
    has_discount = False
    for v in (discount_val, discount_amt):
        if v not in (None, 0, 0.0, "", "0", "0.0"):
            has_discount = True
            break

    # Calculate In_words if missing
    if not full_data.get('Contract__c.In_words__c'):
        total_price = contract.get('Total_Price_USD__c') or 0
        try:
            in_words = amount_to_words_usd(total_price)
            full_data['Contract__c.In_words__c'] = in_words
        except:
            pass
            
    if has_discount:
        template_path = "templates/proforma_invoice_template_new.xlsx"
    else:
        template_path = "templates/proforma_invoice_template_no_discount.xlsx"

    # Verify template exists
    if not os.path.exists(template_path):
        # Fallback to check root directory
        base_name = os.path.basename(template_path)
        if os.path.exists(base_name):
            template_path = base_name
        else:
            print(f"Warning: Template {template_path} not found, falling back to original argument or risking error.")

    # Load Template
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Fill Main Data
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value

                # ===== Handle Incoterms with checkbox formatting =====
                if "{{Contract__c.Incoterms__c}}" in val:
                    incoterms_value = full_data.get('Contract__c.Incoterms__c', '')
                    incoterms_checkbox_text = format_picklist_checkboxes(
                        incoterms_options, incoterms_value, uppercase=True
                    )
                    val = val.replace("{{Contract__c.Incoterms__c}}", incoterms_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                # ===== Handle Terms of Sale with checkbox formatting =====
                if "{{Contract__c.Terms_of_Sale__c}}" in val:
                    terms_of_sale_value = full_data.get('Contract__c.Terms_of_Sale__c', '')
                    terms_of_sale_checkbox_text = format_picklist_checkboxes(
                        terms_of_sale_options, terms_of_sale_value, uppercase=True
                    )
                    val = val.replace("{{Contract__c.Terms_of_Sale__c}}", terms_of_sale_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                # ===== Handle Terms of Payment with checkbox formatting =====
                if "{{Contract__c.Terms_of_Payment__c}}" in val:
                    terms_of_payment_value = full_data.get('Contract__c.Terms_of_Payment__c', '')
                    terms_of_payment_checkbox_text = format_picklist_checkboxes(
                        terms_of_payment_options, terms_of_payment_value, uppercase=True
                    )
                    val = val.replace("{{Contract__c.Terms_of_Payment__c}}", terms_of_payment_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment
                
                # Conditional Logic
                if_pattern = r"\{\{#if\s+([\w\.]+)\s+'=='\s+'([^']+)'\}\}(.*?)\{\{else\}\}(.*?)\{\{/if\}\}"
                if_matches = re.findall(if_pattern, val)
                for match in if_matches:
                    key, target_val, true_text, false_text = match
                    full_match_str = f"{{{{#if {key} '==' '{target_val}'}}}}{true_text}{{{{else}}}}{false_text}{{{{/if}}}}"
                    actual_val = str(full_data.get(key, ""))
                    if actual_val.lower() == target_val.lower():
                        val = val.replace(full_match_str, true_text)
                    else:
                        val = val.replace(full_match_str, false_text)

                # Float Fields
                float_fields = [
                    "{{Contract__c.Total_Crates__c}}", "{{Contract__c.Total_m3__c}}",
                    "{{Contract__c.Total_Tons__c}}", "{{Contract__c.Total_Conts__c}}",
                    "{{Contract__c.Total_m2__c}}",
                    "{{Contract__c.Sub_Total_USD__c\\# #,##0.##}}",
                    "{{Contract__c.Total_Price_USD__c\\# #,##0.##}}",
                    "{{Contract__c.Deposit__c\\# #,##0.##}}",
                    "{{Contract__c.Discount_Amount__c\\# #,##0.##}}",
                    "{{Contract__c.Discount_Amount__c}}"
                ]
                is_float_field = False
                for field in float_fields:
                    if field in val:
                        key_part = field.replace("{{", "").replace("}}", "").split("\\#")[0]
                        value = full_data.get(key_part)
                        if value is not None:
                            try:
                                f_val = float(value)
                                cell.value = f_val
                                # Smart Formatting: Integer if whole number, else float with 2 decimal places
                                if f_val.is_integer():
                                    cell.number_format = '#,##0'
                                else:
                                    cell.number_format = '#,##0.00'
                                is_float_field = True
                            except ValueError:
                                pass
                        break
                if is_float_field:
                    continue

                # Int Fields
                int_fields = [
                   "{{Contract__c.Total_Pcs__c}}",
                   "{{Contract__c.Total_Pcs_PO__c}}",
                   "{{Contract__c.Customer_PO_number__c}}"
                ]
                is_int_field = False
                for field in int_fields:
                    # STRICT check: Only convert to number if the cell contains JUST the placeholder
                    if val and field == val.strip():
                        key_part = field.replace("{{", "").replace("}}", "").split("\#")[0]
                        value = full_data.get(key_part)
                        if value is not None:
                            try:
                                cell.value = int(float(value))
                                cell.number_format = '#,##0'
                                is_int_field = True
                            except ValueError:
                                pass
                        break
                
                if is_int_field:
                    continue

                # General Replacement
                for key, value in full_data.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in val:
                        val = val.replace(placeholder, str(value) if value is not None else "")
                    
                    pattern = f"\\{{{{{key}\\\\#(.*?)\\}}}}"
                    matches = re.findall(pattern, val)
                    for fmt in matches:
                         if value is not None and isinstance(value, (int, float)):
                             if "#,##0.##" in fmt:
                                 formatted_val = "{:,.2f}".format(value)
                             else:
                                 formatted_val = str(value)
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", formatted_val)
                         else:
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value) if value is not None else "")
                cell.value = val

    # Fill Product Table
    table_start_row = expand_table_by_tag(ws, "{{TableStart:ContractProduct2}}", "{{TableEnd:ContractProduct2}}", contract_items)
    
    if table_start_row and contract_items:
        # Merge duplicate "TÊN HÀNG" (Column B / 2) - MỚI
        if contract_items:
            start_row = table_start_row 
            end_row = table_start_row + len(contract_items) - 1 
            col_b_idx = 2
            
            merge_start_row = start_row
            current_val = ws.cell(row=start_row, column=col_b_idx).value
            
            from openpyxl.styles import Alignment
            for r in range(start_row + 1, end_row + 2): 
                val = ws.cell(row=r, column=col_b_idx).value if r <= end_row else "SENTINEL"
                
                should_break = (val != current_val)
                
                if should_break:
                    if r - 1 > merge_start_row:
                        ws.merge_cells(start_row=merge_start_row, start_column=col_b_idx, end_row=r-1, end_column=col_b_idx)
                        ws.cell(row=merge_start_row, column=col_b_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        # Merge K, L, M (11, 12, 13)
                        for col_idx in [11, 12, 13]:
                            ws.merge_cells(start_row=merge_start_row, start_column=col_idx, end_row=r-1, end_column=col_idx)
                            # Center alignment for merged price/amount cells
                            ws.cell(row=merge_start_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    merge_start_row = r
                    current_val = val

        # Format Price Columns (L=12, M=13) and Packing (G=7)
        for i in range(len(contract_items)):
            row_idx = table_start_row + i
            
            # Column 14 (Packing): Custom Format "pcs/crates"
            cell_packing = ws.cell(row=row_idx, column=14)
            if cell_packing.value is not None:
                try:
                    # Ensure it is a number
                    if isinstance(cell_packing.value, str):
                        val_curr = float(str(cell_packing.value).replace(',', ''))
                        cell_packing.value = int(val_curr) if val_curr.is_integer() else val_curr
                    
                    # Apply custom number format
                    cell_packing.number_format = '#,##0 "pcs/crates"'
                except ValueError:
                    pass

            # Unit Price
            cell = ws.cell(row=row_idx, column=12)
            item = contract_items[i]
            
            # Use raw numeric value from record to ensure it's a number in Excel
            val_raw = item.get('Sales_Price__c')
            if val_raw is not None:
                try:
                    cell.value = float(val_raw)

                    # Determine Unit Suffix
                    unit_raw = item.get('Charge_Unit_PI__c')
                    
                    if unit_raw:
                        unit_clean = unit_raw.strip().upper()
                        if "USD" in unit_clean:
                           suffix = unit_clean
                        else:
                           suffix = f"USD/{unit_clean}"
                    else:
                        suffix = "USD"

                    cell.number_format = f'#,##0.00 "{suffix}"'
                except: pass
            
            # Total Price
            cell = ws.cell(row=row_idx, column=13)
            # Use raw numeric value for total price too
            total_raw = item.get('Total_Price_USD__c')
            if total_raw is not None:
                try:
                    cell.value = float(total_raw)
                    cell.number_format = '#,##0.00'
                except: pass

    # Fill Surcharges
    sur_start = expand_table_by_tag(ws, "{{TableStart:PISurcharge}}", "{{TableEnd:PISurcharge}}", surcharge_items)
    if sur_start and surcharge_items:
        for i in range(len(surcharge_items)):
            r = sur_start + i
            ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)

    # Fill Deposits
    dep_start = expand_table_by_tag(ws, "{{TableStart:PIDeposit}}", "{{TableEnd:PIDeposit}}", deposit_items)
    if dep_start and deposit_items:
        for i in range(len(deposit_items)):
            r = dep_start + i
            ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)

    # Fill Discounts
    disc_start = expand_table_by_tag(ws, "{{TableStart:PIDiscount}}", "{{TableEnd:PIDiscount}}", discount_items)
    if disc_start and discount_items:
        for i in range(len(discount_items)):
            r = disc_start + i
            ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)

    # -------------------------------------------------------------------------
    # SỬA ĐỔI: LOGIC MERGE CỘT A-J (1-10) TỰ ĐỘNG THEO SUBTOTAL VÀ TOTAL
    # -------------------------------------------------------------------------
    from openpyxl.styles import Alignment
    
    start_merge_row = None
    end_merge_row = None

    # Quét qua các dòng để tìm vị trí thực tế của Subtotal và Total
    # Vì các bảng ở trên (Surcharge, Discount...) có thể giãn ra, số dòng sẽ thay đổi.
    for r in range(1, ws.max_row + 1):
        # Lấy nội dung text của cả dòng để kiểm tra từ khóa
        row_text_u = ""
        for c in range(1, 15): # Quét 15 cột đầu
            val = ws.cell(row=r, column=c).value
            if val:
                row_text_u += str(val).upper()
        
        # 1. Tìm dòng bắt đầu (SUBTOTAL)
        # Kiểm tra từ khóa: SUBTOTAL, SUB TOTAL, SUB-TOTAL
        if "SUBTOTAL" in row_text_u or "SUB TOTAL" in row_text_u or "SUB-TOTAL" in row_text_u:
            # Chỉ lấy dòng Subtotal đầu tiên tìm thấy sau khi các bảng đã giãn
            if start_merge_row is None:
                start_merge_row = r
        
        # 2. Tìm dòng kết thúc (TOTAL)
        # Kiểm tra từ khóa: TOTAL (tránh Subtotal), TỔNG CỘNG, GRAND TOTAL
        # Lưu ý: "TOTAL" phải nằm dưới "SUBTOTAL"
        if ("TOTAL" in row_text_u and "SUB" not in row_text_u) or "TỔNG CỘNG" in row_text_u or "GRAND TOTAL" in row_text_u:
            if start_merge_row is not None and r > start_merge_row:
                end_merge_row = r
                # Nếu tìm thấy Total hợp lệ thì có thể dừng vòng lặp hoặc tiếp tục để tìm Total cuối cùng (nếu có nhiều block)
                # Ở đây ta giả định PI chỉ có 1 phần tổng ở cuối, nên ta cập nhật end_merge_row liên tục nếu có nhiều dòng Total
    
    # Thực hiện Merge nếu tìm thấy cả 2 mốc
    if start_merge_row and end_merge_row and end_merge_row > start_merge_row:
        # Phạm vi cột cần merge: A (1) đến J (10)
        min_col, max_col = 1, 10
        
        # Bước 1: Unmerge (Gỡ bỏ) các ô đã merge sẵn nằm trong vùng này để tránh lỗi
        ranges_to_unmerge = []
        for mr in ws.merged_cells.ranges:
            # Kiểm tra xem vùng merge có giao nhau với vùng ta chuẩn bị merge không
            if (mr.min_row <= end_merge_row and mr.max_row >= start_merge_row and
                mr.min_col <= max_col and mr.max_col >= min_col):
                ranges_to_unmerge.append(mr)
        
        for mr in ranges_to_unmerge:
            try:
                ws.unmerge_cells(str(mr))
            except Exception as e:
                pass # Bỏ qua lỗi nếu ô đó không thực sự merge

        # Bước 2: Thực hiện Merge từ dòng Subtotal đến dòng Total
        try:
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=end_merge_row-1, end_column=10)
            
            # Merge cho dòng Total riêng biệt
            ws.merge_cells(start_row=end_merge_row, start_column=1, end_row=end_merge_row, end_column=10)
            
            # Bước 3: Căn chỉnh lại text (Căn trái, lên trên)
            cell = ws.cell(row=start_merge_row, column=1)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            cell_total = ws.cell(row=end_merge_row, column=1)
            cell_total.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            print(f"Merged A-J from row {start_merge_row} (Subtotal) to {end_merge_row-1} (Total-1) and row {end_merge_row} (Total)")
        except Exception as e:
            print(f"Merge error A-J: {e}")
    else:
        print("Warning: Could not identify Subtotal and Total rows to merge A-J columns.")

    # -------------------------------------------------------------------------
    # END MERGE LOGIC
    # -------------------------------------------------------------------------

    output_dir = get_output_directory()
    safe_name = sanitize_filename(contract.get('Name'))
    prefix = "PI_Discount_" if has_discount else "PI_NoDiscount_"
    file_name = f"{prefix}{safe_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = output_dir / file_name
    wb.save(str(file_path))
    
    # Upload to Salesforce
    with open(file_path, "rb") as f:
        file_data = f.read()
    encoded = base64.b64encode(file_data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": contract_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"]
    }

@app.get("/generate-pi-no-discount/{contract_id}")
async def generate_pi_no_discount_endpoint(contract_id: str):
    try:
        template_path = os.getenv('PI_NO_DISCOUNT_TEMPLATE_PATH', 'templates/proforma_invoice_template_no_discount.xlsx')
        if not os.path.exists(template_path):
             # Fallback to root if not in templates
             if os.path.exists('proforma_invoice_template_no_discount.xlsx'):
                 template_path = 'proforma_invoice_template_no_discount.xlsx'
             else:
                 raise HTTPException(status_code=404, detail=f"Template not found: {template_path}")
        
        result = generate_pi_no_discount_file(contract_id, template_path)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Production Order Generation ---

def generate_production_order_file(contract_id: str, template_path: str):
    sf = get_salesforce_connection()
    
    # Query Contract
    contract_query = f"""
        SELECT Id, Production_Order_Number__c, Name, CreatedDate, Port_of_Origin__c, 
               Port_of_Discharge__c, Stockyard__c, Total_Pcs_PO__c, Total_Crates__c, 
               Total_m2__c, Total_m3__c, Total_Tons__c, Total_Conts__c, Terms_of_Sale__c
        FROM Contract__c 
        WHERE Id = '{contract_id}'
    """
    try:
        contract_res = sf.query(contract_query)
        contract_data = contract_res['records'][0] if contract_res['totalSize'] > 0 else {}
    except Exception as e:
        print(f"Error querying Contract: {e}")
        raise ValueError(f"Error querying Contract: {e}")

    # Query Order Products
    products_query = f"""
        SELECT Id, IsDeleted, Name, CreatedDate, LastModifiedDate, SystemModstamp, LastActivityDate, LastViewedDate, LastReferencedDate, Charge_Unit__c, Cont__c, Container_Weight_Regulations__c, Crates__c, Height__c, Length__c, List_Price__c, Quantity__c, Width__c, m2__c, m3__c, ml__c, Packing__c, Sales_Price__c, Tons__c, Total_Price_USD__c, Actual_Cont__c, Actual_Crates__c, Actual_Quantity__c, Actual_Tons__c, Actual_m2__c, Actual_m3__c, Actual_ml__c, Product_Description__c, Actual_Total_Price_USD__c, Pending_Cont__c, Pending_Crates__c, Pending_m2__c, Pending_m3__c, Pending_ml__c, Pending_Quantity__c, Pending_Amount_USD__c, Pending_Tons__c, Delivery_Date__c, Planned_Quantity__c, Total_Child_Order_Actual_Quantity__c, Pending_Quantity_for_child_2__c, Delivered_date__c, Line_number__c, Line_item_no_for_print__c, SKU__c, Vietnamese_Description__c, Order__r.Name, Order__r.Delivery_Date__c, Contract_PI__r.Id 
        FROM Order_Product__c 
        WHERE Contract_PI__r.Id = '{contract_id}' 
        ORDER BY Line_number__c ASC
    """
    try:
        products_res = sf.query(products_query)
        products_data = products_res['records']
        
        # FALLBACK: If no Order Products found, query Contract Product items instead
        if not products_data:
            print(f"No Order Products found for {contract_id}, falling back to Contract Products...")
            cp_query = f"""
                SELECT Id, Name, Quantity__c, Crates__c, m2__c, m3__c, Tons__c, Cont__c, 
                       Length__c, Width__c, Height__c, Packing__c, Delivery_Date__c,
                       Product__r.Name, Product__r.ProductCode, Product__r.Product_description_in_Vietnamese__c,
                       Contract__r.Name
                FROM Contract_Product__c 
                WHERE Contract__r.Id = '{contract_id}'
                ORDER BY Line_Number__c ASC
            """
            cp_res = sf.query(cp_query)
            if cp_res['records']:
                for item in cp_res['records']:
                    # Normalize fields to match template expectations
                    norm_item = {
                        "Order__r": {"Name": item.get("Contract__r", {}).get("Name", "")},
                        "SKU__c": item.get("Product__r", {}).get("ProductCode"),
                        "Vietnamese_Description__c": item.get("Product__r", {}).get("Product_description_in_Vietnamese__c"),
                        "Length__c": item.get("Length__c"),
                        "Width__c": item.get("Width__c"),
                        "Height__c": item.get("Height__c"),
                        "Quantity__c": item.get("Quantity__c"),
                        "Crates__c": item.get("Crates__c"),
                        "m2__c": item.get("m2__c"),
                        "m3__c": item.get("m3__c"),
                        "Tons__c": item.get("Tons__c"),
                        "Cont__c": item.get("Cont__c"),
                        "Packing__c": item.get("Packing__c"),
                        "Delivery_Date__c": item.get("Delivery_Date__c")
                    }
                    products_data.append(norm_item)
                    
    except Exception as e:
        print(f"Error querying items: {e}")
        products_data = []

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Flatten data
    flat_data = {}
    if contract_data:
        for k, v in contract_data.items():
            flat_data[f"Contract__c.{k}"] = v
            if "Date" in k and v:
                try:
                    dt = datetime.datetime.strptime(v[:10], "%Y-%m-%d")
                    flat_data[f"Contract__c.{k}\\@dd/MM/yyyy"] = dt.strftime("%d/%m/%Y")
                except: pass

    # Fill placeholders
    total_fields = [
        "Contract__c.Total_Pcs_PO__c",
        "Contract__c.Total_Crates__c",
        "Contract__c.Total_m2__c",
        "Contract__c.Total_m3__c",
        "Contract__c.Total_Tons__c",
        "Contract__c.Total_Conts__c"
    ]

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                
                # Check for smart formatting fields first (exact match of {{Placeholder}})
                is_numeric_total = False
                for field in total_fields:
                    placeholder = f"{{{{{field}}}}}"
                    placeholder_with_fmt = f"{{{{{field}\\#0}}}}" # Handle existing template format if any
                    
                    if val.strip() == placeholder or val.strip() == placeholder_with_fmt:
                        raw_val = flat_data.get(field)
                        if raw_val is not None:
                            try:
                                num_val = float(raw_val)
                                cell.value = num_val
                                if num_val.is_integer():
                                    cell.number_format = '#,##0'
                                else:
                                    cell.number_format = '#,##0.00'
                                is_numeric_total = True
                                break
                            except: pass
                
                if is_numeric_total:
                    continue

                matches = re.findall(r"\{\{([^\}]+)\}\}", val)
                for match in matches:
                    key_part = match.split('\\')[0].strip()
                    format_part = match.split('\\@')[1].strip() if '\\@' in match else None
                    
                    if key_part in flat_data:
                        replace_val = flat_data[key_part]
                        if replace_val is None: replace_val = ""
                        
                        if format_part and replace_val:
                            try:
                                 val_str = str(replace_val).split('T')[0]
                                 if 'T' in str(replace_val):
                                      dt = datetime.datetime.strptime(str(replace_val).split('+')[0].split('.')[0], "%Y-%m-%dT%H:%M:%S")
                                 else:
                                      dt = datetime.datetime.strptime(val_str, "%Y-%m-%d")
                                 py_format = format_part.replace('dd', '%d').replace('MM', '%m').replace('yyyy', '%Y')
                                 replace_val = dt.strftime(py_format)
                            except: pass
                        
                        val = val.replace(f"{{{{{match}}}}}", str(replace_val))
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal=cell.alignment.horizontal if cell.alignment else 'left')
                cell.value = val

    # Fill Table
    table_start_row = None
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val and "{{TableStart:ProPlanProduct}}" in str(cell_val):
            table_start_row = r
            break
            
    if table_start_row:
        if products_data:
            num_items = len(products_data)
            rows_to_insert = num_items - 1
            
            if rows_to_insert > 0:
                # FOOTER PRESERVATION: Capture footer content BEFORE inserting rows
                # Footer starts after the TỔNG CỘNG row (table_start_row + 2 = LƯU Ý row in template)
                # TỔNG CỘNG is at table_start_row + 1, we capture from row after it
                footer_start_row_original = table_start_row + 2  # Skip TỔNG CỘNG, start at LƯU Ý
                footer_data = []  # List of {row_offset, col, value, font, alignment, border, number_format}
                footer_merges = []  # List of merge info relative to footer_start
                
                # Capture all cell values and styles in footer (from LƯU Ý to end of sheet)
                for r in range(footer_start_row_original, ws.max_row + 1):
                    for c in range(1, 16):
                        cell = ws.cell(row=r, column=c)
                        if cell.value is not None:  # Only capture cells with values
                            footer_data.append({
                                'row_offset': r - footer_start_row_original,
                                'col': c,
                                'value': cell.value,
                                'font': style_copy(cell.font) if cell.font else None,
                                'alignment': style_copy(cell.alignment) if cell.alignment else None,
                                'border': style_copy(cell.border) if cell.border else None,
                                'number_format': cell.number_format
                            })
                
                # Capture merged ranges in footer zone
                for merged_range in list(ws.merged_cells.ranges):
                    if merged_range.min_row >= footer_start_row_original:
                        footer_merges.append({
                            'min_row_offset': merged_range.min_row - footer_start_row_original,
                            'max_row_offset': merged_range.max_row - footer_start_row_original,
                            'min_col': merged_range.min_col,
                            'max_col': merged_range.max_col
                        })
                
                # Now insert rows
                ws.insert_rows(table_start_row + 1, amount=rows_to_insert)
                
                # FOOTER RESTORATION: Restore captured footer content to new positions
                # New footer start row = product rows + TỔNG CỘNG = table_start_row + num_items + 1
                footer_start_row_new = table_start_row + num_items + 1
                
                # Restore cell values
                for item in footer_data:
                    new_row = footer_start_row_new + item['row_offset']
                    cell = ws.cell(row=new_row, column=item['col'])
                    cell.value = item['value']
                    if item['font']: cell.font = item['font']
                    if item['alignment']: cell.alignment = item['alignment']
                    if item['border']: cell.border = item['border']
                    if item['number_format']: cell.number_format = item['number_format']
                
                # Restore merged ranges (first unmerge any existing, then re-merge)
                for merge_info in footer_merges:
                    new_min_row = footer_start_row_new + merge_info['min_row_offset']
                    new_max_row = footer_start_row_new + merge_info['max_row_offset']
                    # Unmerge any existing ranges in this area first
                    for col in range(merge_info['min_col'], merge_info['max_col'] + 1):
                        cell = ws.cell(row=new_min_row, column=col)
                        for existing_merge in list(ws.merged_cells.ranges):
                            if cell.coordinate in existing_merge:
                                try: ws.unmerge_cells(str(existing_merge))
                                except: pass
                    # Re-apply the merge
                    try:
                        ws.merge_cells(start_row=new_min_row, start_column=merge_info['min_col'],
                                       end_row=new_max_row, end_column=merge_info['max_col'])
                    except: pass

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Copy styles
            if num_items > 1:
                for i in range(1, num_items):
                    for col in range(1, 16):
                        source_cell = ws.cell(row=table_start_row, column=col)
                        target_cell = ws.cell(row=table_start_row + i, column=col)
                        if source_cell.border: target_cell.border = style_copy(source_cell.border)
                        if source_cell.font: target_cell.font = style_copy(source_cell.font)
                        if source_cell.alignment: target_cell.alignment = style_copy(source_cell.alignment)
                        if source_cell.number_format: target_cell.number_format = style_copy(source_cell.number_format)

            ws.cell(row=table_start_row, column=1).value = ""
            
            for i, item in enumerate(products_data):
                row_idx = table_start_row + i
                
                # Unmerge
                for col in range(1, 16):
                    cell = ws.cell(row=row_idx, column=col)
                    for merged_range in list(ws.merged_cells.ranges):
                        if cell.coordinate in merged_range:
                            try: ws.unmerge_cells(str(merged_range))
                            except: pass
                    ws.cell(row=row_idx, column=col).border = thin_border

                # Map Data
                ws.cell(row=row_idx, column=1).value = i + 1
                ws.cell(row=row_idx, column=1).alignment = align_center
                ws.cell(row=row_idx, column=2).value = item.get("Order__r", {}).get("Name") if item.get("Order__r") else ""
                ws.cell(row=row_idx, column=2).alignment = align_center
                ws.cell(row=row_idx, column=3).value = item.get("SKU__c")
                ws.cell(row=row_idx, column=3).alignment = align_left
                
                # Rich Text Description
                desc_val = item.get("Vietnamese_Description__c") or ""
                if desc_val and '-' in str(desc_val):
                    parts = str(desc_val).split('-', 1)
                    rich_text = CellRichText(
                        TextBlock(InlineFont(b=True, rFont='Times New Roman', sz=11), parts[0]),
                        TextBlock(InlineFont(b=False, rFont='Times New Roman', sz=11), '-' + parts[1])
                    )
                    ws.cell(row=row_idx, column=4).value = rich_text
                else:
                    ws.cell(row=row_idx, column=4).value = desc_val
                ws.cell(row=row_idx, column=4).alignment = align_left
                
                # Dimensions & Quantity
                ws.cell(row=row_idx, column=5).value = item.get("Length__c")
                ws.cell(row=row_idx, column=6).value = item.get("Width__c")
                ws.cell(row=row_idx, column=7).value = item.get("Height__c")
                ws.cell(row=row_idx, column=8).value = item.get("Quantity__c")
                ws.cell(row=row_idx, column=9).value = item.get("Crates__c")
                
                if item.get("m2__c"): 
                    ws.cell(row=row_idx, column=10).value = float(item.get("m2__c"))
                    ws.cell(row=row_idx, column=10).number_format = '0.00'
                if item.get("m3__c"):
                    ws.cell(row=row_idx, column=11).value = float(item.get("m3__c"))
                    ws.cell(row=row_idx, column=11).number_format = '0.00'
                    
                ws.cell(row=row_idx, column=12).value = item.get("Tons__c")
                ws.cell(row=row_idx, column=13).value = item.get("Cont__c")
                
                for col in range(5, 14): ws.cell(row=row_idx, column=col).alignment = align_center
                
                # Packing
                packing_val = item.get("Packing__c")
                if packing_val:
                    try:
                        # Chuyển sang int
                        ws.cell(row=row_idx, column=14).value = int(float(packing_val))
                        ws.cell(row=row_idx, column=14).number_format = '0 "viên/kiện"'
                    except:
                        ws.cell(row=row_idx, column=14).value = f"{packing_val}\nviên/kiện"
                ws.cell(row=row_idx, column=14).alignment = align_center
                
                # Delivery Date
                del_date = item.get("Delivery_Date__c")
                # Fallback to Order's Delivery Date if None
                if not del_date:
                    del_date = item.get("Order__r", {}).get("Delivery_Date__c")

                if del_date:
                    try:
                        dt = datetime.datetime.strptime(del_date[:10], "%Y-%m-%d")
                        ws.cell(row=row_idx, column=15).value = dt.strftime("%d/%m/%Y")
                    except:
                        ws.cell(row=row_idx, column=15).value = del_date
                ws.cell(row=row_idx, column=15).alignment = align_center

            # Merge duplicate "TÊN HÀNG" (Column D / 4) - Sync with Delivery Date Logic
            start_merge_row = table_start_row
            current_val = ws.cell(row=start_merge_row, column=4).value
            for i in range(1, len(products_data)):
                row_idx = table_start_row + i
                val = ws.cell(row=row_idx, column=4).value
                if val != current_val:
                    if row_idx - 1 > start_merge_row:
                        ws.merge_cells(start_row=start_merge_row, start_column=4, end_row=row_idx-1, end_column=4)
                        ws.cell(row=start_merge_row, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    start_merge_row = row_idx
                    current_val = val
            last_row = table_start_row + len(products_data) - 1
            if last_row > start_merge_row:
                ws.merge_cells(start_row=start_merge_row, start_column=4, end_row=last_row, end_column=4)
                ws.cell(row=start_merge_row, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Merge duplicate "THỜI GIAN GIAO HÀNG" (Column O / 15)
            start_merge_row = table_start_row
            current_val = ws.cell(row=start_merge_row, column=15).value
            for i in range(1, len(products_data)):
                row_idx = table_start_row + i
                val = ws.cell(row=row_idx, column=15).value
                if val != current_val:
                    if row_idx - 1 > start_merge_row:
                        ws.merge_cells(start_row=start_merge_row, start_column=15, end_row=row_idx-1, end_column=15)
                        ws.cell(row=start_merge_row, column=15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    start_merge_row = row_idx
                    current_val = val
            last_row = table_start_row + len(products_data) - 1
            if last_row > start_merge_row:
                ws.merge_cells(start_row=start_merge_row, start_column=15, end_row=last_row, end_column=15)
                ws.cell(row=start_merge_row, column=15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            # CLEANUP: If no products_data, clear placeholders and tags from the template row
            for col in range(1, 16):
                cell = ws.cell(row=table_start_row, column=col)
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.replace("{{TableStart:ProPlanProduct}}", "").replace("{{TableEnd:ProPlanProduct}}", "")
                    val = re.sub(r"\{\{.*?\}\}", "", val)
                    cell.value = val.strip() if val.strip() else None







    # ----------------------------------------------------
    # MERGE I, J, K FOR ROWS WITH "Người soạn lệnh" OR "Ngọc Bích"
    # ----------------------------------------------------
    for r in range(1, ws.max_row + 1):
        found_keyword = False
        row_values_ijk = []
        target_val = None
        
        # Check entire row to find the keyword "Người soạn lệnh" or "Ngọc Bích"
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=col)
            if cell.value and isinstance(cell.value, str):
                val_upper = str(cell.value).strip().upper()
                if "NGƯỜI SOẠN LỆNH" in val_upper or "NGƯỜI SOAN LỆNH" in val_upper or "NGỌC BÍCH" in val_upper:
                    found_keyword = True
                    # If we found it, we break. But we need to know WHICH value to keep in IJK merge.
                    # We will resume scanning IJK specifically below.
                    break
        
        if found_keyword:
            # Check content in I(9), J(10), K(11) to preserve it
            # We want to keep the value if it exists in one of these cells
            val_9 = ws.cell(row=r, column=9).value
            val_10 = ws.cell(row=r, column=10).value
            val_11 = ws.cell(row=r, column=11).value
            
            # Prioritize the first non-empty value among them
            final_val = val_9 if val_9 is not None else (val_10 if val_10 is not None else val_11)
            
            try:
                # Unmerge if already merged (sanity check)
                # Then set value to I(9) and clear J, K
                ws.cell(row=r, column=9).value = final_val
                ws.cell(row=r, column=10).value = None
                ws.cell(row=r, column=11).value = None
                
                # Apply Styling
                val_str = str(final_val).upper() if final_val else ""
                if "NGƯỜI SOẠN LỆNH" in val_str or "NGƯỜI SOAN LỆNH" in val_str:
                    ws.cell(row=r, column=9).font = Font(bold=True, underline='single', name='Times New Roman', size=11)
                elif "NGỌC BÍCH" in val_str:
                    ws.cell(row=r, column=9).font = Font(bold=True, name='Times New Roman', size=11)
                
                ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
                ws.cell(row=r, column=9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except Exception as e:
                print(f"Error merging IJK at row {r}: {e}")

    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    safe_name = sanitize_filename(contract_data.get('Production_Order_Number__c', 'Draft'))
    file_name = f"Production_Order_{safe_name}_{timestamp}.xlsx"
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    wb.save(str(file_path))
    
    # Upload to Salesforce
    with open(file_path, "rb") as f:
        file_data = f.read()
    encoded = base64.b64encode(file_data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": contract_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"]
    }



# --- Quote No Discount Generation ---

def generate_quote_no_discount_file(quote_id: str, template_path: str):
    sf = get_salesforce_connection()

    # Get picklist values for Quote
    incoterms_options = get_picklist_values(sf, 'Quote', 'Incoterms__c')
    terms_of_sale_options = get_picklist_values(sf, 'Quote', 'Terms_of_Sale__c')
    terms_of_payment_options = get_picklist_values(sf, 'Quote', 'Terms_of_Payment__c')
    
    # Query Quote Items (Full Query)
    query = f"""
    SELECT Id, IsDeleted, LineNumber, CreatedDate, LastModifiedDate, SystemModstamp, LastViewedDate, LastReferencedDate, Quantity, UnitPrice, Discount, HasRevenueSchedule, HasQuantitySchedule, Description, ServiceDate, SortOrder, HasSchedule, ListPrice, Subtotal, TotalPrice, Product_Description__c, Length__c, Width__c, Height__c, Line_Number__c, Packing__c, Total_Price_to_sumup__c, Cont__c, Crates__c, Tons__c, Container_Weight_Regulations__c, Discount__c, Unit_Price__c, L_x_W_x_H__c, ml_x_m2_x_m3__c, Crates_and_Packing__c, Unit_Price_USD__c, ChargeUnit__c, Product_Name__c, m2__c, m3__c, ml__c, Total_Price_USD__c, L_Quote__c, W_Quote__c, H_Quote__c, PCS_Quote__c, Crates_Quote__c, Charge_Unit_Quote__c, Packing_Quote__c, Quote_Line_Item_Number_Quote__c, Opportunity_Id__c, Quote_display_name__c, Quote.Id, Quote.OwnerId, Quote.IsDeleted, Quote.Name, Quote.RecordTypeId, Quote.CreatedDate, Quote.CreatedById, Quote.LastModifiedDate, Quote.LastModifiedById, Quote.SystemModstamp, Quote.LastViewedDate, Quote.LastReferencedDate, Quote.OpportunityId, Quote.Pricebook2Id, Quote.ContactId, Quote.QuoteNumber, Quote.IsSyncing, Quote.ShippingHandling, Quote.Tax, Quote.Status, Quote.ExpirationDate, Quote.Description, Quote.Subtotal, Quote.TotalPrice, Quote.LineItemCount, Quote.BillingStreet, Quote.BillingCity, Quote.BillingState, Quote.BillingPostalCode, Quote.BillingCountry, Quote.BillingLatitude, Quote.BillingLongitude, Quote.BillingGeocodeAccuracy, Quote.BillingAddress, Quote.ShippingStreet, Quote.ShippingCity, Quote.ShippingState, Quote.ShippingPostalCode, Quote.ShippingCountry, Quote.ShippingLatitude, Quote.ShippingLongitude, Quote.ShippingGeocodeAccuracy, Quote.ShippingAddress, Quote.QuoteToStreet, Quote.QuoteToCity, Quote.QuoteToState, Quote.QuoteToPostalCode, Quote.QuoteToCountry, Quote.QuoteToLatitude, Quote.QuoteToLongitude, Quote.QuoteToGeocodeAccuracy, Quote.QuoteToAddress, Quote.AdditionalStreet, Quote.AdditionalCity, Quote.AdditionalState, Quote.AdditionalPostalCode, Quote.AdditionalCountry, Quote.AdditionalLatitude, Quote.AdditionalLongitude, Quote.AdditionalGeocodeAccuracy, Quote.AdditionalAddress, Quote.BillingName, Quote.ShippingName, Quote.QuoteToName, Quote.AdditionalName, Quote.Email, Quote.Phone, Quote.Fax, Quote.ContractId, Quote.AccountId, Quote.Discount, Quote.GrandTotal, Quote.CanCreateQuoteLineItems, Quote.Sub_Total_USD__c, Quote.Fumigation__c, Quote.Total_Crates__c, Quote.Total_m3__c, Quote.Total_Tons__c, Quote.Total_Conts__c, Quote.REMARK_NUMBER_ON_DOCUMENTS__c, Quote.Packing__c, Quote.Shipping_Schedule__c, Quote.Port_of_Discharge__c, Quote.Export_Route_Carrier__c, Quote.In_words__c, Quote.Discount__c, Quote.Total_Price_USD__c, Quote.Total_Quote_Line_Items__c, Quote.Port_of_Origin__c, Quote.Stockyard__c, Quote.Created_Date__c, Quote.Discount_Amount__c, Quote.Is_new_quote__c, Quote.First_approved_by__c, Quote.Final_approved_by__c, Quote.Account_approved_pricebook__c, Quote.Is_approved__c, Quote.Terms_of_Sale__c, Quote.Terms_of_Payment__c, Quote.Incoterms__c 
    FROM QuoteLineItem 
    WHERE QuoteId = '{quote_id}' 
    ORDER BY Quote_Line_Item_Number_Quote__c ASC
    """
    
    try:
        result = sf.query_all(query)
    except Exception as e:
        print(f"Error querying quote items: {e}")
        raise ValueError(f"Error querying quote items: {e}")

    if not result['records']:
        # Try fetching just the Quote if no items
        try:
            q_res = sf.query(f"SELECT Id, Name FROM Quote WHERE Id = '{quote_id}'")
            if q_res['records']:
                quote_data = q_res['records'][0]
                quote_items = []
            else:
                raise ValueError(f"Quote not found: {quote_id}")
        except:
            raise ValueError(f"Quote not found: {quote_id}")
    else:
        quote_items = result['records']
        first_item = quote_items[0]
        if 'Quote' in first_item and first_item['Quote']:
            quote_data = first_item['Quote']
        else:
            raise ValueError("Quote data missing in line items.")

    # Flatten Data
    full_data = {}
    for k, v in quote_data.items():
        full_data[f"Quote.{k}"] = v
        
    # Fetch Account
    account_id = quote_data.get('AccountId')
    if account_id:
        acc_fields = ["Name", "BillingStreet", "BillingCity", "BillingPostalCode", "BillingCountry", "Phone", "Fax__c", "VAT__c"]
        try:
            acc = sf.Account.get(account_id)
            for k in acc_fields:
                full_data[f"Quote.Account.{k}"] = acc.get(k)
        except: pass

    # Inject Sequential Number
    for idx, item in enumerate(quote_items):
        item['Quote_Line_Item_Number_Quote__c'] = idx + 1

    # Query Discounts (Discount_Item__c - Placeholder)
    discount_items = []
    try:
        discount_query = f"SELECT Id, Name, Discount_Amount__c FROM Discount_Item__c WHERE Quote__c = '{quote_id}'"
        disc_result = sf.query_all(discount_query)
        discount_records = disc_result['records']
        for item in discount_records:
            val = item.get('Discount_Amount__c')
            if val is not None:
                try: val = float(val)
                except: pass
            discount_items.append({
                "Name": item.get('Name'),
                "Discount_Amount__c": val
            })
    except Exception:
        pass

    # Determine Template based on Discount
    discount_val = quote_data.get('Discount')
    discount_amt = quote_data.get('Discount_Amount__c') # Check field name from query
    
    has_discount = False
    for v in (discount_val, discount_amt):
        if v not in (None, 0, 0.0, "", "0", "0.0"):
            has_discount = True
            break
            
    if has_discount:
        template_path = "templates/quotation_template_new.xlsx"
    else:
        template_path = "templates/quotation_template_no_discount.xlsx"

    # Verify template exists
    if not os.path.exists(template_path):
        # Fallback to check root directory
        base_name = os.path.basename(template_path)
        if os.path.exists(base_name):
            template_path = base_name
        else:
             print(f"Warning: Template {template_path} not found, falling back to original argument or risking error.")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Fill Main Data
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value

                # ===== Handle Incoterms with checkbox formatting =====
                if "{{Quote.Incoterms__c}}" in val:
                    incoterms_value = full_data.get('Quote.Incoterms__c', '')
                    incoterms_checkbox_text = format_picklist_checkboxes(
                        incoterms_options, incoterms_value, uppercase=True
                    )
                    val = val.replace("{{Quote.Incoterms__c}}", incoterms_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                # ===== Handle Terms of Sale with checkbox formatting =====
                if "{{Quote.Terms_of_Sale__c}}" in val:
                    terms_of_sale_value = full_data.get('Quote.Terms_of_Sale__c', '')
                    terms_of_sale_checkbox_text = format_picklist_checkboxes(
                        terms_of_sale_options, terms_of_sale_value, uppercase=True
                    )
                    val = val.replace("{{Quote.Terms_of_Sale__c}}", terms_of_sale_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                # ===== Handle Terms of Payment with checkbox formatting =====
                if "{{Quote.Terms_of_Payment__c}}" in val:
                    terms_of_payment_value = full_data.get('Quote.Terms_of_Payment__c', '')
                    terms_of_payment_checkbox_text = format_picklist_checkboxes(
                        terms_of_payment_options, terms_of_payment_value, uppercase=True
                    )
                    val = val.replace("{{Quote.Terms_of_Payment__c}}", terms_of_payment_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment
                
                # Conditional Logic
                if_pattern = r"\{\{#if\s+([\w\.]+)\s+'=='\s+'([^']+)'\}\}(.*?)\{\{else\}\}(.*?)\{\{/if\}\}"
                if_matches = re.findall(if_pattern, val)
                for match in if_matches:
                    key, target_val, true_text, false_text = match
                    full_match_str = f"{{{{#if {key} '==' '{target_val}'}}}}{true_text}{{{{else}}}}{false_text}{{{{/if}}}}"
                    actual_val = str(full_data.get(key, ""))
                    if actual_val.lower() == target_val.lower():
                        val = val.replace(full_match_str, true_text)
                    else:
                        val = val.replace(full_match_str, false_text)

                # Float Fields
                float_fields = [
                    "{{Quote.Total_Crates__c}}", "{{Quote.Total_m3__c}}",
                    "{{Quote.Total_Tons__c}}", "{{Quote.Total_Conts__c}}",
                    "{{Quote.Sub_Total_USD__c\\# #,##0.##}}",
                    "{{Quote.Total_Price_USD__c\\# #,##0.##}}",
                    "{{Quote.Discount_Amount__c\\# #,##0.##}}",
                    "{{Quote.Discount_Amount__c}}"
                ]
                is_float_field = False
                for field in float_fields:
                    if field in val:
                        key_part = field.replace("{{", "").replace("}}", "").split("\\#")[0]
                        value = full_data.get(key_part)
                        if value is not None:
                            try:
                                cell.value = float(value)
                                cell.number_format = '#,##0.00'
                                is_float_field = True
                            except ValueError:
                                pass
                        break
                if is_float_field:
                    continue

                # General Replacement
                for key, value in full_data.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in val:
                        val = val.replace(placeholder, str(value) if value is not None else "")
                    
                    pattern = f"\\{{{{{key}\\\\#(.*?)\\}}}}"
                    matches = re.findall(pattern, val)
                    for fmt in matches:
                         if value is not None and isinstance(value, (int, float)):
                             if "#,##0.##" in fmt:
                                 formatted_val = "{:,.2f}".format(value)
                             else:
                                 formatted_val = str(value)
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", formatted_val)
                         else:
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value) if value is not None else "")
                cell.value = val

    # Fill Product Table
    table_start_row = expand_table_by_tag(ws, "{{TableStart:GetQuoteLine}}", "{{TableEnd:GetQuoteLine}}", quote_items)

    # Fill Discounts
    disc_start = expand_table_by_tag(ws, "{{TableStart:QuoteDiscount}}", "{{TableEnd:QuoteDiscount}}", discount_items)
    if disc_start and discount_items:
        for i in range(len(discount_items)):
            r = disc_start + i
            ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)

    # Merge Cols A-J (1-10) from "All prices quoted herein" down to "TOTAL"
    from openpyxl.styles import Alignment
    start_merge_row = None
    end_merge_row = None

    # Find start row ("All prices quoted herein" or similar)
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val and isinstance(cell_val, str) and ("All prices quoted herein" in cell_val or "Prices quoted herein" in cell_val):
            start_merge_row = r
            break
            
    # Find end row ("TOTAL")
    if start_merge_row:
        for r in range(start_merge_row, ws.max_row + 1):
            found_total = False
            for c in range(11, 16): # Check deeper K-O
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str) and "TOTAL" in val.upper():
                    end_merge_row = r
                    found_total = True
                    break
            if found_total:
                break
                
    if start_merge_row and end_merge_row and end_merge_row > start_merge_row:
        # Define range
        min_row, max_row = start_merge_row, end_merge_row
        min_col, max_col = 1, 10
        
        # Unmerge overlapping ranges
        ranges_to_unmerge = []
        for mr in ws.merged_cells.ranges:
            if (mr.min_row <= max_row and mr.max_row >= min_row and
                mr.min_col <= max_col and mr.max_col >= min_col):
                ranges_to_unmerge.append(mr)
        
        for mr in ranges_to_unmerge:
            try: ws.unmerge_cells(str(mr))
            except: pass

        try:
             ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=end_merge_row, end_column=10)
             cell = ws.cell(row=start_merge_row, column=1)
             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        except Exception as e:
            print(f"Merge error: {e}")

    if table_start_row and quote_items:
        # Merge duplicate "TÊN HÀNG" (Column B / 2) - MỚI
        if quote_items:
            start_row = table_start_row
            end_row = table_start_row + len(quote_items) - 1
            col_b_idx = 2
            
            merge_start_row = start_row
            current_val = ws.cell(row=start_row, column=col_b_idx).value
            
            from openpyxl.styles import Alignment
            for r in range(start_row + 1, end_row + 2):
                val = ws.cell(row=r, column=col_b_idx).value if r <= end_row else "SENTINEL"
                
                should_break = (val != current_val)
                
                if should_break:
                    if r - 1 > merge_start_row:
                        ws.merge_cells(start_row=merge_start_row, start_column=col_b_idx, end_row=r-1, end_column=col_b_idx)
                        ws.cell(row=merge_start_row, column=col_b_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    merge_start_row = r
                    current_val = val

        # Format Price Columns (L=12, M=13)
        for i in range(len(quote_items)):
            row_idx = table_start_row + i
            # Unit Price
            cell = ws.cell(row=row_idx, column=12)
            if cell.value:
                try:
                    val_str = str(cell.value).replace(',', '')
                    cell.value = float(val_str)
                    cell.number_format = '#,##0.00 "USD"'
                except: pass
            # Total Price
            cell = ws.cell(row=row_idx, column=13)
            if cell.value:
                try:
                    val_str = str(cell.value).replace(',', '')
                    cell.value = float(val_str)
                    cell.number_format = '#,##0.00'
                except: pass

    # Footer Row Height
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "All prices quoted herein" in cell.value:
                ws.row_dimensions[cell.row].height = 50
                ws.row_dimensions[cell.row + 1].height = 50
                break

    # Save
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    safe_name = sanitize_filename(quote_data.get('Name'))
    prefix = "Quote_Discount_" if has_discount else "Quote_NoDiscount_"
    file_name = f"{prefix}{safe_name}_{timestamp}.xlsx"
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    wb.save(str(file_path))
    
    # Upload to Salesforce
    with open(file_path, "rb") as f:
        file_data = f.read()
    encoded = base64.b64encode(file_data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": quote_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"]
    }



# ==========================================
# NEW ENDPOINTS: PI, Quote, Production Order
# ==========================================

# --- PI No Discount Logic ---
def expand_table_pi(ws, start_tag, end_tag, data):
    """
    Expand a single row table based on start and end tags (PI version).
    """
    # Find the row containing the tags
    table_row_idx = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if start_tag in cell.value:
                    table_row_idx = cell.row
                    break
        if table_row_idx:
            break
            
    if not table_row_idx:
        print(f"Warning: Table tags {start_tag} not found.")
        return None

    if not data:
        # Clear tags
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=table_row_idx, column=col)
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace(start_tag, "").replace(end_tag, "")
        return table_row_idx

    num_rows = len(data)
    add_rows = max(0, num_rows - 1)
    
    # Capture styles from the template row
    max_col = ws.max_column
    row_style = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=table_row_idx, column=col)
        row_style.append(style_copy(cell._style) if cell.has_style else None)
    
    row_height = ws.row_dimensions[table_row_idx].height

    # Handle merged cells (shift them down)
    merges_to_shift = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > table_row_idx:
            merges_to_shift.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))
    
    for mr in merges_to_shift:
        rng = f"{get_column_letter(mr[2])}{mr[0]}:{get_column_letter(mr[3])}{mr[1]}"
        ws.unmerge_cells(rng)

    # Insert rows if needed
    if add_rows > 0:
        ws.insert_rows(table_row_idx + 1, amount=add_rows)
        
        for offset in range(1, add_rows + 1):
            r = table_row_idx + offset
            # Copy row height
            if row_height is not None:
                ws.row_dimensions[r].height = row_height
                
            for col in range(1, max_col + 1):
                dst = ws.cell(row=r, column=col)
                # Copy value from template row
                src_val = ws.cell(row=table_row_idx, column=col).value
                dst.value = src_val
                
                # Copy style
                st = row_style[col - 1]
                if st is not None:
                    dst._style = style_copy(st)
                    
    # Re-merge shifted cells
    for mr in merges_to_shift:
        new_min_row = mr[0] + add_rows
        new_max_row = mr[1] + add_rows
        rng = f"{get_column_letter(mr[2])}{new_min_row}:{get_column_letter(mr[3])}{new_max_row}"
        ws.merge_cells(rng)
                    
    # Fill data
    for i, record in enumerate(data):
        current_row_idx = table_row_idx + i
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=current_row_idx, column=col)
            if cell.value and isinstance(cell.value, str):
                cell_val = cell.value.replace(start_tag, "").replace(end_tag, "")
                
                for key, value in record.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in cell_val:
                        cell_val = cell_val.replace(placeholder, str(value) if value is not None else "")
                        
                    pattern = f"\\{{{{{key}\\\\#(.*?)\\}}}}"
                    matches = re.findall(pattern, cell_val)
                    for fmt in matches:
                        try:
                            if value is not None:
                                if isinstance(value, (int, float)):
                                    if "#,##0.##" in fmt:
                                         formatted_val = "{:,.2f}".format(value)
                                    else:
                                         formatted_val = str(value)
                                    cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", formatted_val)
                                else:
                                    cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value))
                            else:
                                cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", "")
                        except:
                             cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value))
                
                cell.value = cell_val
                
                # Attempt to convert to number if it looks like one
                if isinstance(cell.value, str):
                    try:
                        clean_val = cell.value.replace(',', '')
                        f_val = float(clean_val)
                        is_leading_zero = (len(clean_val) > 1 and clean_val.startswith('0') and not clean_val.startswith('0.'))
                        
                        if not is_leading_zero:
                            if f_val.is_integer():
                                cell.value = int(f_val)
                            else:
                                cell.value = f_val
                    except ValueError:
                        pass
    
    return table_row_idx

def get_picklist_values(sf, object_name, field_name):
    """
    Fetch picklist values for a given object and field from Salesforce.
    """
    try:
        desc = getattr(sf, object_name).describe()
        for field in desc['fields']:
            if field['name'] == field_name:
                return [entry['value'] for entry in field['picklistValues']]
    except Exception as e:
        print(f"Error fetching picklist values for {object_name}.{field_name}: {e}")
    return []

def format_picklist_checkboxes(options, selected_value, uppercase=False):
    """
    Format picklist options as a checkbox list.
    Mark the selected value with [x], others with [ ].
    """
    formatted_lines = []
    if selected_value is None:
        selected_value = ""
    
    # Normalize selected value for comparison
    selected_value_norm = str(selected_value).strip().lower()
    
    for opt in options:
        opt_label = str(opt)
        if uppercase:
            opt_label = opt_label.upper()
            
        # Check if this option is selected
        is_selected = opt.strip().lower() == selected_value_norm
        
        checkbox = "☑" if is_selected else "☐"
        formatted_lines.append(f"{checkbox} {opt_label}")
        
    return "\n".join(formatted_lines)

def safe_float(val):
    try:
        if val is None: return 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def generate_pi_no_discount_logic(contract_id, template_path):
    sf = get_salesforce_connection()
    
    # Get picklist values for checkbox formatting
    incoterms_options = get_picklist_values(sf, 'Contract__c', 'Incoterms__c')
    terms_of_sale_options = get_picklist_values(sf, 'Contract__c', 'Terms_of_Sale__c')
    terms_of_payment_options = get_picklist_values(sf, 'Contract__c', 'Terms_of_Payment__c')
    
    # Query Contract
    query = f"""
    SELECT Id, IsDeleted, Name, CreatedDate, LastModifiedDate, SystemModstamp, LastActivityDate, LastViewedDate, LastReferencedDate, Cont__c, Container_Weight_Regulations__c, Crates__c, Height__c, Length__c, Line_Number__c, Packing__c, Sales_Price__c, Tons__c, Width__c, List_Price__c, Discount__c, Charge_Unit__c, Quantity__c, m2__c, m3__c, ml__c, Total_Price_USD__c, L_PI__c, W_PI__c, H_PI__c, PCS_PI__c, Crates_PI__c, Created_Date__c, Packing_PI__c, Product_Discription__c, Charge_Unit_PI__c, Actual_Cont__c, Pending_Cont__c, Clear__c, Actual_Crates__c, Actual_m2__c, Actual_m3__c, Actual_ml__c, Actual_Quantity__c, Actual_Tons__c, Actual_Total_Price_USD__c, Pending_Crates__c, Pending_m2__c, Pending_m3__c, Pending_ml__c, Pending_Quantity__c, Pending_Tons__c, Pending_Amount_USD__c, Delivery_Date__c, Delivery_Quantity__c, Is_Delivery_Quantity_Valid__c, Delivery_Quantity_number__c, Unscheduled_Quantity__c, Line_number_For_print__c, Product__r.Id, Product__r.Name, Product__r.ProductCode, Product__r.Description, Product__r.QuantityScheduleType, Product__r.QuantityInstallmentPeriod, Product__r.NumberOfQuantityInstallments, Product__r.RevenueScheduleType, Product__r.RevenueInstallmentPeriod, Product__r.NumberOfRevenueInstallments, Product__r.IsActive, Product__r.CreatedDate, Product__r.CreatedById, Product__r.LastModifiedDate, Product__r.LastModifiedById, Product__r.SystemModstamp, Product__r.Family, Product__r.ExternalDataSourceId, Product__r.ExternalId, Product__r.DisplayUrl, Product__r.QuantityUnitOfMeasure, Product__r.IsDeleted, Product__r.IsArchived, Product__r.LastViewedDate, Product__r.LastReferencedDate, Product__r.StockKeepingUnit, Product__r.Product_description_in_Vietnamese__c, Product__r.specific_gravity__c, Product__r.Bottom_cladding_coefficient__c, Product__r.STONE_Color_Type__c, Product__r.Packing__c, Product__r.Long__c, Product__r.High__c, Product__r.Width__c, Product__r.Long_special__c, Product__r.High_special__c, Product__r.Image__c, Product__r.Charge_Unit__c, Product__r.Width_special__c, Product__r.STONE_Class__c, Product__r.Description__c, Product__r.List_Price__c, Product__r.Weight_per_unit__c, Product__r.Edge_Finish__c, Product__r.Suppliers__c, Product__r.m_per_unit__c, Product__r.Application__c, Product__r.Surface_Finish__c, Product__r.m3_per_unit__c, Product__r.Pricing_Method__c, Contract__r.Id, Contract__r.OwnerId, Contract__r.IsDeleted, Contract__r.Name, Contract__r.CreatedDate, Contract__r.CreatedById, Contract__r.LastModifiedDate, Contract__r.LastModifiedById, Contract__r.SystemModstamp, Contract__r.LastActivityDate, Contract__r.LastViewedDate, Contract__r.LastReferencedDate, Contract__r.Account__c, Contract__r.Quote__c, Contract__r.Bill_To__c, Contract__r.Bill_To_Name__c, Contract__r.Contact_Name__c, Contract__r.Expiration_Date__c, Contract__r.Export_Route_Carrier__c, Contract__r.Fax__c, Contract__r.Phone__c, Contract__r.Fumigation__c, Contract__r.Incoterms__c, Contract__r.In_words__c, Contract__r.Packing__c, Contract__r.Port_of_Discharge__c, Contract__r.REMARK_NUMBER_ON_DOCUMENTS__c, Contract__r.Shipping_Schedule__c, Contract__r.Total_Conts__c, Contract__r.Total_Crates__c, Contract__r.Total_m3__c, Contract__r.Sub_Total_USD__c, Contract__r.Total_Tons__c, Contract__r.Deposit_Percentage__c, Contract__r.Discount__c, Contract__r.Total_Price_USD__c, Contract__r.Deposit__c, Contract__r.Stage__c, Contract__r.Total_Payment_Received__c, Contract__r.Expected_ETD__c, Contract__r.Port_of_Origin__c, Contract__r.Price_Book__c, Contract__r.Stockyard__c, Contract__r.Created_Date__c, Contract__r.Total_Contract_Product__c, Contract__r.Pending_Products__c, Contract__r.Total_Payment_Received_USD__c, Contract__r.Production_Order_Number__c, Contract__r.Total_m2__c, Contract__r.Total_Pcs__c, Contract__r.Total_Pcs_PO__c, Contract__r.Planned_Shipments__c, Contract__r.Is_approved__c, Contract__r.Deposited_amount_USD__c, Contract__r.Design_confirmed__c, Contract__r.Contract_type__c, Contract__r.Fully_deposited__c, Contract__r.Discount_Amount__c, Contract__r.Terms_of_Payment__c, Contract__r.Terms_of_Sale__c, Contract__r.Total_surcharge__c FROM Contract_Product__c where Contract__r.Id = '{contract_id}' ORDER BY Line_Number__c ASC
    """
    
    result = sf.query_all(query)
    if not result['records']:
        raise ValueError(f"No contract items found for ID: {contract_id}")
    
    contract_items = result['records']
    first_item = contract_items[0]
    if 'Contract__r' in first_item and first_item['Contract__r']:
        contract_data = first_item['Contract__r']
    else:
        raise ValueError("Contract data missing in line items.")

    full_data = {}
    for k, v in contract_data.items():
        full_data[f"Contract__c.{k}"] = v
        
    # --- Calculate Totals Locally ---
    total_crates = 0.0
    total_m2 = 0.0
    total_m3 = 0.0
    total_tons = 0.0
    total_conts = 0.0
    
    for item in contract_items:
        total_crates += safe_float(item.get('Crates__c'))
        total_m2 += safe_float(item.get('m2__c'))
        total_m3 += safe_float(item.get('m3__c'))
        total_tons += safe_float(item.get('Tons__c'))
        total_conts += safe_float(item.get('Cont__c'))
        
    full_data['Contract__c.Total_Crates__c'] = total_crates
    full_data['Contract__c.Total_m2__c'] = total_m2
    full_data['Contract__c.Total_m3__c'] = total_m3
    full_data['Contract__c.Total_Tons__c'] = total_tons
    full_data['Contract__c.Total_Conts__c'] = total_conts
    # --------------------------------

    account_id = contract_data.get('Account__c')
    if account_id:
        acc_fields = ["Name", "BillingStreet", "BillingCity", "BillingPostalCode", "BillingCountry", "Phone", "Fax__c", "VAT__c"]
        try:
            acc = sf.Account.get(account_id)
            for k in acc_fields:
                full_data[f"Contract__c.Account__r.{k}"] = acc.get(k)
        except Exception as e:
            print(f"Error fetching account: {e}")
            
    for idx, item in enumerate(contract_items):
        item['Line_number_For_print__c'] = idx + 1
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                
                # ===== Handle Incoterms with checkbox formatting =====
                if "{{Contract__c.Incoterms__c}}" in val:
                    incoterms_value = full_data.get('Contract__c.Incoterms__c', '')
                    incoterms_checkbox_text = format_picklist_checkboxes(
                        incoterms_options, incoterms_value, uppercase=True
                    )
                    val = val.replace("{{Contract__c.Incoterms__c}}", incoterms_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                # ===== Handle Terms of Sale with checkbox formatting =====
                if "{{Contract__c.Terms_of_Sale__c}}" in val:
                    terms_of_sale_value = full_data.get('Contract__c.Terms_of_Sale__c', '')
                    terms_of_sale_checkbox_text = format_picklist_checkboxes(
                        terms_of_sale_options, terms_of_sale_value, uppercase=True
                    )
                    val = val.replace("{{Contract__c.Terms_of_Sale__c}}", terms_of_sale_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                # ===== Handle Terms of Payment with checkbox formatting =====
                if "{{Contract__c.Terms_of_Payment__c}}" in val:
                    terms_of_payment_value = full_data.get('Contract__c.Terms_of_Payment__c', '')
                    terms_of_payment_checkbox_text = format_picklist_checkboxes(
                        terms_of_payment_options, terms_of_payment_value, uppercase=True
                    )
                    val = val.replace("{{Contract__c.Terms_of_Payment__c}}", terms_of_payment_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                if_pattern = r"\{\{#if\s+([\w\.]+)\s+'(==|contains)'\s+'([^']+)'\}\}(.*?)\{\{else\}\}(.*?)\{\{/if\}\}"
                if_matches = re.findall(if_pattern, val)
                for match in if_matches:
                    key, operator, target_val, true_text, false_text = match
                    full_match_str = f"{{{{#if {key} '{operator}' '{target_val}'}}}}{true_text}{{{{else}}}}{false_text}{{{{/if}}}}"
                    
                    actual_val = full_data.get(key)
                    if actual_val is None:
                        actual_val = ""
                    else:
                        actual_val = str(actual_val)
                        
                    condition_met = False
                    if operator == '==':
                        condition_met = actual_val.lower() == target_val.lower()
                    elif operator == 'contains':
                        condition_met = target_val.lower() in actual_val.lower()
                        
                    if condition_met:
                        val = val.replace(full_match_str, true_text)
                    else:
                        val = val.replace(full_match_str, false_text)

                # --- NEW: Handle Discount Logic (Clear if 0) ---
                if "{{Contract__c.Discount__c}}" in val or "{{Contract__c.Discount_Amount__c" in val:
                    discount_val = full_data.get('Contract__c.Discount__c')
                    discount_amt = full_data.get('Contract__c.Discount_Amount__c')
                    
                    # Check if essentially 0
                    is_zero = True
                    try:
                        if discount_val and float(discount_val) != 0:
                            is_zero = False
                        if discount_amt and float(discount_amt) != 0:
                            is_zero = False
                    except:
                        pass
                        
                    if is_zero:
                        cell.value = ""
                        continue
                # -----------------------------------------------

                float_fields = [
                    "{{Contract__c.Total_Crates__c}}",
                    "{{Contract__c.Total_m2__c}}",
                    "{{Contract__c.Total_m3__c}}",
                    "{{Contract__c.Total_Tons__c}}",
                    "{{Contract__c.Total_Conts__c}}",
                    "{{Contract__c.Sub_Total_USD__c\\# #,##0.##}}",
                    "{{Contract__c.Total_Price_USD__c\\# #,##0.##}}",
                    "{{Contract__c.Deposit__c\\# #,##0.##}}"
                ]
                
                is_float_field = False
                for field in float_fields:
                    if field in val:
                        key_part = field.replace("{{", "").replace("}}", "").split("\\#")[0]
                        value = full_data.get(key_part)
                        if value is not None:
                            try:
                                cell.value = float(value)
                                cell.number_format = '#,##0.00'
                                is_float_field = True
                            except ValueError:
                                pass
                        break
                
                if is_float_field:
                    continue

                for key, value in full_data.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in val:
                        val = val.replace(placeholder, str(value) if value is not None else "")
                    
                    pattern = f"\\{{{{{key}\\\\#(.*?)\\}}}}"
                    matches = re.findall(pattern, val)
                    for fmt in matches:
                         if value is not None and isinstance(value, (int, float)):
                             if "#,##0.##" in fmt:
                                 formatted_val = "{:,.2f}".format(value)
                             else:
                                 formatted_val = str(value)
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", formatted_val)
                         else:
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value) if value is not None else "")
                
                cell.value = val

    table_start_row = expand_table_pi(ws, "{{TableStart:ContractProduct2}}", "{{TableEnd:ContractProduct2}}", contract_items)
    
    if table_start_row and contract_items:
        col_b_idx = 2
        for i, item in enumerate(contract_items):
            row_idx = table_start_row + i
            cell = ws.cell(row=row_idx, column=col_b_idx)
            product_name = None
            if 'Product__r' in item and item['Product__r']:
                product_name = item['Product__r'].get('Name')
            current_desc = str(cell.value) if cell.value else ""

    if table_start_row and contract_items:
        col_b_idx = 2
        start_merge_row = table_start_row
        current_val = str(ws.cell(row=start_merge_row, column=col_b_idx).value)
        for i in range(1, len(contract_items)):
            row_idx = table_start_row + i
            cell_val = str(ws.cell(row=row_idx, column=col_b_idx).value)
            if cell_val == current_val:
                continue
            else:
                if row_idx - 1 > start_merge_row:
                    ws.merge_cells(start_row=start_merge_row, start_column=col_b_idx, end_row=row_idx-1, end_column=col_b_idx)
                    cell = ws.cell(row=start_merge_row, column=col_b_idx)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                start_merge_row = row_idx
                current_val = cell_val
        last_row = table_start_row + len(contract_items) - 1
        if last_row > start_merge_row:
             ws.merge_cells(start_row=start_merge_row, start_column=col_b_idx, end_row=last_row, end_column=col_b_idx)
             ws.cell(row=start_merge_row, column=col_b_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    if table_start_row and contract_items:
        for i in range(len(contract_items)):
            row_idx = table_start_row + i
            cell = ws.cell(row=row_idx, column=12)
            if cell.value:
                try:
                    val_str = str(cell.value).replace(',', '')
                    cell.value = float(val_str)
                    cell.number_format = '#,##0.00 "USD"'
                except ValueError:
                    pass 
            cell = ws.cell(row=row_idx, column=13)
            if cell.value:
                try:
                    val_str = str(cell.value).replace(',', '')
                    cell.value = float(val_str)
                    cell.number_format = '#,##0.00'
                except ValueError:
                    pass

    # Fill Surcharge Table
    # The template uses {{TableStart:PISurcharge}}...{{TableEnd:PISurcharge}}
    surcharge_query = f"""
    SELECT Id, Name, Surcharge_amount_USD__c 
    FROM Expense__c 
    WHERE Contract_PI__r.Id = '{contract_id}' AND Surcharge_amount_USD__c != 0
    """
    
    try:
        surcharge_result = sf.query_all(surcharge_query)
        surcharge_records = surcharge_result['records']
    except Exception as e:
        print(f"Error querying surcharge expenses: {e}")
        surcharge_records = []

    surcharge_items = []
    if surcharge_records:
        for item in surcharge_records:
            surcharge_items.append({
                "Name": item.get('Name'),
                "Surcharge_amount_USD__c": item.get('Surcharge_amount_USD__c')
            })

    expand_table_pi(ws, "{{TableStart:PISurcharge}}", "{{TableEnd:PISurcharge}}", surcharge_items)

    # Fill Deposit Table (Single row from Contract)
    deposit_items = []
    if contract_data:
        total_amount = safe_float(contract_data.get('Total_Price_USD__c'))
        deposit_amount = safe_float(contract_data.get('Deposit__c'))
        balance = total_amount - deposit_amount
        
        deposit_items.append({
            "Deposit__c": contract_data.get('Deposit__c'),
            "Deposit_Percentage__c": contract_data.get('Deposit_Percentage__c'),
            "Total_Price_USD__c": contract_data.get('Total_Price_USD__c'),
            "Balance__c": balance 
        })
        
    expand_table_pi(ws, "{{TableStart:PIDeposit}}", "{{TableEnd:PIDeposit}}", deposit_items)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "All prices quoted herein" in cell.value:
                ws.row_dimensions[cell.row].height = 50
                ws.row_dimensions[cell.row + 1].height = 50
                break

    output_dir = get_output_directory()
    safe_name = sanitize_filename(contract_data.get('Name'))
    file_name = f"PI_{safe_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = output_dir / file_name
    wb.save(str(file_path))

    # Upload to Salesforce
    with open(file_path, "rb") as f:
        file_data = f.read()
    encoded = base64.b64encode(file_data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": contract_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"]
    }

@app.get("/generate-pi-no-discount/{contract_id}")
async def generate_pi_no_discount_endpoint(contract_id: str):
    try:
        # Check if contract has discount first
        sf = get_salesforce_connection()
        query = f"SELECT Discount__c, Discount_Amount__c FROM Contract__c WHERE Id = '{contract_id}'"
        res = sf.query(query)
        has_discount = False
        if res['totalSize'] > 0:
            rec = res['records'][0]
            d_percent = rec.get('Discount__c')
            d_amount = rec.get('Discount_Amount__c')
            if (d_percent and float(d_percent) != 0) or (d_amount and float(d_amount) != 0):
                has_discount = True
        
        if has_discount:
            template_path = os.getenv('PI_TEMPLATE_PATH', 'templates/proforma_invoice_template_new.xlsx')
        else:
             template_path = 'templates/proforma_invoice_template_no_discount.xlsx'

        if not os.path.exists(template_path):
             # Fallback
             if has_discount:
                 template_path = 'templates/proforma_invoice_template_new.xlsx'
             else:
                 template_path = 'templates/proforma_invoice_template_no_discount.xlsx'
             
             if not os.path.exists(template_path):
                 # Ultimate fallback
                 fallback = 'templates/proforma_invoice_template_new.xlsx'
                 if os.path.exists(fallback):
                     template_path = fallback
                 else:
                     raise HTTPException(status_code=404, detail=f"PI Template not found: {template_path}")

        result = generate_pi_no_discount_logic(contract_id, template_path)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Quote No Discount Logic ---
def expand_table_quote(ws, start_tag, end_tag, data):
    """
    Expand a single row table based on start and end tags (Quote version with strict types).
    """
    table_row_idx = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if start_tag in cell.value:
                    table_row_idx = cell.row
                    break
        if table_row_idx:
            break
            
    if not table_row_idx:
        print(f"Warning: Table tags {start_tag} not found.")
        return None

    if not data:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=table_row_idx, column=col)
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace(start_tag, "").replace(end_tag, "")
        return table_row_idx

    num_rows = len(data)
    add_rows = max(0, num_rows - 1)
    
    max_col = ws.max_column
    row_style = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=table_row_idx, column=col)
        row_style.append(style_copy(cell._style) if cell.has_style else None)
    
    row_height = ws.row_dimensions[table_row_idx].height

    merges_to_shift = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > table_row_idx:
            merges_to_shift.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))
    
    for mr in merges_to_shift:
        rng = f"{get_column_letter(mr[2])}{mr[0]}:{get_column_letter(mr[3])}{mr[1]}"
        ws.unmerge_cells(rng)

    if add_rows > 0:
        ws.insert_rows(table_row_idx + 1, amount=add_rows)
        for offset in range(1, add_rows + 1):
            r = table_row_idx + offset
            if row_height is not None:
                ws.row_dimensions[r].height = row_height
            for col in range(1, max_col + 1):
                dst = ws.cell(row=r, column=col)
                src_val = ws.cell(row=table_row_idx, column=col).value
                dst.value = src_val
                st = row_style[col - 1]
                if st is not None:
                    dst._style = style_copy(st)
                    
    for mr in merges_to_shift:
        new_min_row = mr[0] + add_rows
        new_max_row = mr[1] + add_rows
        rng = f"{get_column_letter(mr[2])}{new_min_row}:{get_column_letter(mr[3])}{new_max_row}"
        ws.merge_cells(rng)
                    
    for i, record in enumerate(data):
        current_row_idx = table_row_idx + i
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=current_row_idx, column=col)
            if cell.value and isinstance(cell.value, str):
                cell_val = cell.value.replace(start_tag, "").replace(end_tag, "")
                
                # --- Enforce int/float types for specific fields ---
                int_fields = [
                    "L_Quote__c", "W_Quote__c", "H_Quote__c", 
                    "PCS_Quote__c", "Crates_Quote__c", "Packing_Quote__c", 
                    "Quote_Line_Item_Number_Quote__c"
                ]
                float_fields = ["m2__c", "m3__c", "Tons__c", "Cont__c"]
                
                is_numeric_cell = False
                for key, value in record.items():
                    placeholder = f"{{{{{key}}}}}"
                    if cell_val.strip() == placeholder:
                        if key in int_fields and value is not None:
                            try:
                                cell.value = int(float(value))
                                is_numeric_cell = True
                            except (ValueError, TypeError):
                                pass
                        elif key in float_fields and value is not None:
                            try:
                                cell.value = float(value)
                                cell.number_format = '#,##0.00'
                                is_numeric_cell = True
                            except (ValueError, TypeError):
                                pass
                        if is_numeric_cell:
                            break
                
                if is_numeric_cell:
                    continue
                # --------------------------------------------------------

                for key, value in record.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in cell_val:
                        cell_val = cell_val.replace(placeholder, str(value) if value is not None else "")
                        
                    pattern = f"\\{{{{{key}\\\\#(.*?)\\}}}}"
                    matches = re.findall(pattern, cell_val)
                    for fmt in matches:
                        try:
                            if value is not None:
                                if isinstance(value, (int, float)):
                                    if "#,##0.##" in fmt:
                                         formatted_val = "{:,.2f}".format(value)
                                    else:
                                         formatted_val = str(value)
                                    cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", formatted_val)
                                else:
                                    cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value))
                            else:
                                cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", "")
                        except:
                             cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value))
                
                cell.value = cell_val
    
    return table_row_idx

def generate_quote_no_discount_logic(quote_id, template_path):
    sf = get_salesforce_connection()
    
    # Get picklist values for checkbox formatting
    incoterms_options = get_picklist_values(sf, 'Quote', 'Incoterms__c')
    terms_of_sale_options = get_picklist_values(sf, 'Quote', 'Terms_of_Sale__c')
    terms_of_payment_options = get_picklist_values(sf, 'Quote', 'Terms_of_Payment__c')
    
    query = f"""
    SELECT Id, IsDeleted, LineNumber, CreatedDate, LastModifiedDate, SystemModstamp, LastViewedDate, LastReferencedDate, Quantity, UnitPrice, Discount, HasRevenueSchedule, HasQuantitySchedule, Description, ServiceDate, SortOrder, HasSchedule, ListPrice, Subtotal, TotalPrice, Product_Description__c, Length__c, Width__c, Height__c, Line_Number__c, Packing__c, Total_Price_to_sumup__c, Cont__c, Crates__c, Tons__c, Container_Weight_Regulations__c, Discount__c, Unit_Price__c, L_x_W_x_H__c, ml_x_m2_x_m3__c, Crates_and_Packing__c, Unit_Price_USD__c, ChargeUnit__c, Product_Name__c, m2__c, m3__c, ml__c, Total_Price_USD__c, L_Quote__c, W_Quote__c, H_Quote__c, PCS_Quote__c, Crates_Quote__c, Charge_Unit_Quote__c, Packing_Quote__c, Quote_Line_Item_Number_Quote__c, Opportunity_Id__c, Quote_display_name__c, Quote.Id, Quote.OwnerId, Quote.IsDeleted, Quote.Name, Quote.RecordTypeId, Quote.CreatedDate, Quote.CreatedById, Quote.LastModifiedDate, Quote.LastModifiedById, Quote.SystemModstamp, Quote.LastViewedDate, Quote.LastReferencedDate, Quote.OpportunityId, Quote.Pricebook2Id, Quote.ContactId, Quote.QuoteNumber, Quote.IsSyncing, Quote.ShippingHandling, Quote.Tax, Quote.Status, Quote.ExpirationDate, Quote.Description, Quote.Subtotal, Quote.TotalPrice, Quote.LineItemCount, Quote.BillingStreet, Quote.BillingCity, Quote.BillingState, Quote.BillingPostalCode, Quote.BillingCountry, Quote.BillingLatitude, Quote.BillingLongitude, Quote.BillingGeocodeAccuracy, Quote.BillingAddress, Quote.ShippingStreet, Quote.ShippingCity, Quote.ShippingState, Quote.ShippingPostalCode, Quote.ShippingCountry, Quote.ShippingLatitude, Quote.ShippingLongitude, Quote.ShippingGeocodeAccuracy, Quote.ShippingAddress, Quote.QuoteToStreet, Quote.QuoteToCity, Quote.QuoteToState, Quote.QuoteToPostalCode, Quote.QuoteToCountry, Quote.QuoteToLatitude, Quote.QuoteToLongitude, Quote.QuoteToGeocodeAccuracy, Quote.QuoteToAddress, Quote.AdditionalStreet, Quote.AdditionalCity, Quote.AdditionalState, Quote.AdditionalPostalCode, Quote.AdditionalCountry, Quote.AdditionalLatitude, Quote.AdditionalLongitude, Quote.AdditionalGeocodeAccuracy, Quote.AdditionalAddress, Quote.BillingName, Quote.ShippingName, Quote.QuoteToName, Quote.AdditionalName, Quote.Email, Quote.Phone, Quote.Fax, Quote.ContractId, Quote.AccountId, Quote.Discount, Quote.GrandTotal, Quote.CanCreateQuoteLineItems, Quote.Sub_Total_USD__c, Quote.Fumigation__c, Quote.Total_Crates__c, Quote.Total_m3__c, Quote.Total_Tons__c, Quote.Total_Conts__c, Quote.REMARK_NUMBER_ON_DOCUMENTS__c, Quote.Packing__c, Quote.Shipping_Schedule__c, Quote.Port_of_Discharge__c, Quote.Export_Route_Carrier__c, Quote.In_words__c, Quote.Discount__c, Quote.Total_Price_USD__c, Quote.Total_Quote_Line_Items__c, Quote.Port_of_Origin__c, Quote.Stockyard__c, Quote.Created_Date__c, Quote.Discount_Amount__c, Quote.Is_new_quote__c, Quote.First_approved_by__c, Quote.Final_approved_by__c, Quote.Account_approved_pricebook__c, Quote.Is_approved__c, Quote.Terms_of_Sale__c, Quote.Terms_of_Payment__c, Quote.Incoterms__c 
    FROM QuoteLineItem 
    WHERE QuoteId = '{quote_id}' 
    ORDER BY Quote_Line_Item_Number_Quote__c ASC
    """
    
    result = sf.query_all(query)
    if not result['records']:
        # Try fetching just the Quote
        q_res = sf.query(f"SELECT Id, Name FROM Quote WHERE Id = '{quote_id}'")
        if q_res['records']:
            quote_data = q_res['records'][0]
            quote_items = []
        else:
            raise ValueError(f"Quote not found: {quote_id}")
    else:
        quote_items = result['records']
        first_item = quote_items[0]
        if 'Quote' in first_item and first_item['Quote']:
            quote_data = first_item['Quote']
        else:
            raise ValueError("Quote data missing in line items.")

    full_data = {}
    for k, v in quote_data.items():
        full_data[f"Quote.{k}"] = v
        
    # --- Calculate Totals Locally ---
    total_crates = 0.0
    total_m3 = 0.0
    total_tons = 0.0
    total_conts = 0.0
    
    for item in quote_items:
        total_crates += safe_float(item.get('Crates_Quote__c'))
        total_m3 += safe_float(item.get('m3__c'))
        total_tons += safe_float(item.get('Tons__c'))
        total_conts += safe_float(item.get('Cont__c'))
        
    full_data['Quote.Total_Crates__c'] = total_crates
    full_data['Quote.Total_m3__c'] = total_m3
    full_data['Quote.Total_Tons__c'] = total_tons
    full_data['Quote.Total_Conts__c'] = total_conts
    # --------------------------------
    account_id = quote_data.get('AccountId')
    if account_id:
        acc_fields = ["Name", "BillingStreet", "BillingCity", "BillingPostalCode", "BillingCountry", "Phone", "Fax__c", "VAT__c"]
        try:
            acc = sf.Account.get(account_id)
            for k in acc_fields:
                full_data[f"Quote.Account.{k}"] = acc.get(k)
        except Exception as e:
            print(f"Error fetching account: {e}")
    
    for idx, item in enumerate(quote_items):
        item['Quote_Line_Item_Number_Quote__c'] = idx + 1
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value

                # ===== Handle Incoterms with checkbox formatting =====
                if "{{Quote.Incoterms__c}}" in val:
                    incoterms_value = full_data.get('Quote.Incoterms__c', '')
                    incoterms_checkbox_text = format_picklist_checkboxes(
                        incoterms_options, incoterms_value, uppercase=True
                    )
                    val = val.replace("{{Quote.Incoterms__c}}", incoterms_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                # ===== Handle Terms of Sale with checkbox formatting =====
                if "{{Quote.Terms_of_Sale__c}}" in val:
                    terms_of_sale_value = full_data.get('Quote.Terms_of_Sale__c', '')
                    terms_of_sale_checkbox_text = format_picklist_checkboxes(
                        terms_of_sale_options, terms_of_sale_value, uppercase=True
                    )
                    val = val.replace("{{Quote.Terms_of_Sale__c}}", terms_of_sale_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                # ===== Handle Terms of Payment with checkbox formatting =====
                if "{{Quote.Terms_of_Payment__c}}" in val:
                    terms_of_payment_value = full_data.get('Quote.Terms_of_Payment__c', '')
                    terms_of_payment_checkbox_text = format_picklist_checkboxes(
                        terms_of_payment_options, terms_of_payment_value, uppercase=True
                    )
                    val = val.replace("{{Quote.Terms_of_Payment__c}}", terms_of_payment_checkbox_text)
                    if cell.alignment:
                        new_alignment = style_copy(cell.alignment)
                    else:
                        new_alignment = Alignment()
                    new_alignment.wrap_text = True
                    cell.alignment = new_alignment

                if_pattern = r"\{\{#if\s+([\w\.]+)\s+'(==|contains)'\s+'([^']+)'\}\}(.*?)\{\{else\}\}(.*?)\{\{/if\}\}"
                if_matches = re.findall(if_pattern, val)
                for match in if_matches:
                    key, operator, target_val, true_text, false_text = match
                    full_match_str = f"{{{{#if {key} '{operator}' '{target_val}'}}}}{true_text}{{{{else}}}}{false_text}{{{{/if}}}}"
                    
                    actual_val = full_data.get(key)
                    if actual_val is None:
                        actual_val = ""
                    else:
                        actual_val = str(actual_val)
                        
                    condition_met = False
                    if operator == '==':
                        condition_met = actual_val.lower() == target_val.lower()
                    elif operator == 'contains':
                        condition_met = target_val.lower() in actual_val.lower()
                        
                    if condition_met:
                        val = val.replace(full_match_str, true_text)
                    else:
                        val = val.replace(full_match_str, false_text)

                float_fields = [
                    "{{Quote.Total_Crates__c}}",
                    "{{Quote.Total_m3__c}}",
                    "{{Quote.Total_Tons__c}}",
                    "{{Quote.Total_Conts__c}}",
                    "{{Quote.Sub_Total_USD__c\\# #,##0.##}}",
                    "{{Quote.Total_Price_USD__c\\# #,##0.##}}"
                ]
                
                is_float_field = False
                for field in float_fields:
                    if field in val:
                        key_part = field.replace("{{", "").replace("}}", "").split("\\#")[0]
                        value = full_data.get(key_part)
                        if value is not None:
                            try:
                                cell.value = float(value)
                                cell.number_format = '#,##0.00'
                                is_float_field = True
                            except ValueError:
                                pass
                        break
                
                if is_float_field:
                    continue

                for key, value in full_data.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in val:
                        val = val.replace(placeholder, str(value) if value is not None else "")
                    
                    pattern = f"\\{{{{{key}\\\\#(.*?)\\}}}}"
                    matches = re.findall(pattern, val)
                    for fmt in matches:
                         if value is not None and isinstance(value, (int, float)):
                             if "#,##0.##" in fmt:
                                 formatted_val = "{:,.2f}".format(value)
                             else:
                                 formatted_val = str(value)
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", formatted_val)
                         else:
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value) if value is not None else "")
                
                cell.value = val

    table_start_row = expand_table_quote(ws, "{{TableStart:GetQuoteLine}}", "{{TableEnd:GetQuoteLine}}", quote_items)
    if table_start_row and quote_items:
        col_b_idx = 2
        for i, item in enumerate(quote_items):
            row_idx = table_start_row + i
            cell = ws.cell(row=row_idx, column=col_b_idx)
            product_name = item.get('Product_Name__c')
            current_desc = str(cell.value) if cell.value else ""

    if table_start_row and quote_items:
        col_b_idx = 2
        start_merge_row = table_start_row
        current_val = str(ws.cell(row=start_merge_row, column=col_b_idx).value)
        for i in range(1, len(quote_items)):
            row_idx = table_start_row + i
            cell_val = str(ws.cell(row=row_idx, column=col_b_idx).value)
            if cell_val == current_val:
                continue
            else:
                if row_idx - 1 > start_merge_row:
                    ws.merge_cells(start_row=start_merge_row, start_column=col_b_idx, end_row=row_idx-1, end_column=col_b_idx)
                    ws.cell(row=start_merge_row, column=col_b_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                start_merge_row = row_idx
                current_val = cell_val
        last_row = table_start_row + len(quote_items) - 1
        if last_row > start_merge_row:
             ws.merge_cells(start_row=start_merge_row, start_column=col_b_idx, end_row=last_row, end_column=col_b_idx)
             ws.cell(row=start_merge_row, column=col_b_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
 
    if table_start_row and quote_items:
        for i in range(len(quote_items)):
            row_idx = table_start_row + i
            cell = ws.cell(row=row_idx, column=12)
            if cell.value:
                try:
                    val_str = str(cell.value).replace(',', '')
                    cell.value = float(val_str)
                    cell.number_format = '#,##0.00 "USD"'
                except ValueError:
                    pass 
            cell = ws.cell(row=row_idx, column=13)
            if cell.value:
                try:
                    val_str = str(cell.value).replace(',', '')
                    cell.value = float(val_str)
                    cell.number_format = '#,##0.00'
                except ValueError:
                    pass

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "All prices quoted herein" in cell.value:
                ws.row_dimensions[cell.row].height = 50
                ws.row_dimensions[cell.row + 1].height = 50
                break

    output_dir = get_output_directory()
    safe_name = sanitize_filename(quote_data.get('Name'))
    file_name = f"Quote_{safe_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = output_dir / file_name
    wb.save(str(file_path))

    # Upload to Salesforce
    with open(file_path, "rb") as f:
        file_data = f.read()
    encoded = base64.b64encode(file_data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": quote_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"]
    }

@app.get("/generate-quote-no-discount/{quote_id}")
async def generate_quote_no_discount_endpoint(quote_id: str):
    try:
        template_path = os.getenv('QUOTE_TEMPLATE_PATH', 'templates/quotation_template_no_discount.xlsx')
        if not os.path.exists(template_path):
             raise HTTPException(status_code=404, detail=f"Quote Template not found")

        result = generate_quote_no_discount_logic(quote_id, template_path)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Production Order Logic ---
def get_production_order_data(sf, contract_id):
    if not sf:
        return None, None

    # Query Contract
    contract_query = f"""
        SELECT Id, Production_Order_Number__c, Name, CreatedDate, Port_of_Origin__c, 
               Port_of_Discharge__c, Stockyard__c, Total_Pcs_PO__c, Total_Crates__c, 
               Total_m2__c, Total_m3__c, Total_Tons__c, Total_Conts__c, Terms_of_Sale__c
        FROM Contract__c 
        WHERE Id = '{contract_id}'
    """
    try:
        contract_res = sf.query(contract_query)
        contract_data = contract_res['records'][0] if contract_res['totalSize'] > 0 else {}
    except Exception as e:
        print(f"Error querying Contract: {e}")
        contract_data = {}

    # Query Order Products
    products_query = f"""
        SELECT Id, IsDeleted, Name, CreatedDate, LastModifiedDate, SystemModstamp, LastActivityDate, LastViewedDate, LastReferencedDate, Charge_Unit__c, Cont__c, Container_Weight_Regulations__c, Crates__c, Height__c, Length__c, List_Price__c, Quantity__c, Width__c, m2__c, m3__c, ml__c, Packing__c, Sales_Price__c, Tons__c, Total_Price_USD__c, Actual_Cont__c, Actual_Crates__c, Actual_Quantity__c, Actual_Tons__c, Actual_m2__c, Actual_m3__c, Actual_ml__c, Product_Description__c, Actual_Total_Price_USD__c, Pending_Cont__c, Pending_Crates__c, Pending_m2__c, Pending_m3__c, Pending_ml__c, Pending_Quantity__c, Pending_Amount_USD__c, Pending_Tons__c, Delivery_Date__c, Planned_Quantity__c, Total_Child_Order_Actual_Quantity__c, Pending_Quantity_for_child_2__c, Delivered_date__c, Line_number__c, Line_item_no_for_print__c, SKU__c, Vietnamese_Description__c, Order__r.Name, Contract_PI__r.Id 
        FROM Order_Product__c 
        WHERE Contract_PI__r.Id = '{contract_id}' 
        ORDER BY Line_number__c ASC
    """
    try:
        products_res = sf.query(products_query)
        products_data = products_res['records']
    except Exception as e:
        print(f"Error querying Order Products: {e}")
        products_data = []

    return contract_data, products_data

def fill_production_order_template(template_path, output_path, contract_data, products_data):
    print(f"Filling template: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Flatten contract data for easier replacement
    flat_data = {}
    if contract_data:
        for k, v in contract_data.items():
            flat_data[f"Contract__c.{k}"] = v
            # Handle date formatting
            if "Date" in k and v:
                try:
                    dt = datetime.datetime.strptime(v[:10], "%Y-%m-%d")
                    flat_data[f"Contract__c.{k}\\@dd/MM/yyyy"] = dt.strftime("%d/%m/%Y")
                except:
                    pass

    # Fill simple placeholders
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                matches = re.findall(r"\{\{([^\}]+)\}\}", val)
                for match in matches:
                    key_part = match.split('\\')[0].strip()
                    format_part = None
                    if '\\@' in match:
                        format_part = match.split('\\@')[1].strip()
                    
                    if key_part in flat_data:
                        replace_val = flat_data[key_part]
                        if replace_val is None:
                            replace_val = ""
                        
                        if format_part and replace_val:
                            try:
                                val_str = str(replace_val)
                                if 'T' in val_str:
                                    dt = datetime.datetime.strptime(val_str.split('+')[0].split('.')[0], "%Y-%m-%dT%H:%M:%S")
                                else:
                                    dt = datetime.datetime.strptime(val_str, "%Y-%m-%d")
                                
                                py_format = format_part.replace('dd', '%d').replace('MM', '%m').replace('yyyy', '%Y')
                                replace_val = dt.strftime(py_format)
                            except Exception as e:
                                # print(f"Error formatting date {replace_val} with {format_part}: {e}")
                                replace_val = str(replace_val).split('T')[0]

                        total_fields = [
                            "Contract__c.Total_Pcs_PO__c", "Contract__c.Total_Crates__c", "Contract__c.Total_m2__c",
                            "Contract__c.Total_m3__c", "Contract__c.Total_Tons__c", "Contract__c.Total_Conts__c"
                        ]
                        if key_part in total_fields and replace_val is not None:
                            try:
                                float_val = float(replace_val)
                                replace_val = int(float_val) if float_val.is_integer() else float_val
                            except (ValueError, TypeError):
                                pass

                        val = val.replace(f"{{{{{match}}}}}", str(replace_val))
                        
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal=cell.alignment.horizontal if cell.alignment else 'left')
                        
                        val_str = str(replace_val)
                        explicit_lines = val_str.count('\n') + 1
                        wrap_lines = (len(val_str) // 20) + 1 
                        est_lines = max(explicit_lines, wrap_lines)
                        
                        if est_lines > 1:
                            current_height = ws.row_dimensions[cell.row].height or 15
                            ws.row_dimensions[cell.row].height = max(current_height, est_lines * 20)
                    else:
                        pass
                
                try:
                    clean_val = str(val).replace(',', '')
                    float_val = float(clean_val)
                    cell.value = int(float_val) if float_val.is_integer() else float_val
                except ValueError:
                    cell.value = val

    # Fill Table
    table_start_row = None
    total_row_template_idx = None # Thêm biến để lưu chỉ mục dòng Tổng Cộng template
    
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val and "{{TableStart:ProPlanProduct}}" in str(cell_val):
            table_start_row = r
        if ws.cell(row=r, column=4).value and "TỔNG CỘNG" in str(ws.cell(row=r, column=4).value).upper():
            total_row_template_idx = r
            
    if not table_start_row:
        print("Error: Table start marker {{TableStart:ProPlanProduct}} not found.")
        return

    num_items = len(products_data)
    
    if products_data:
        print(f"Found table start at row {table_start_row}. Expanding for {num_items} items.")
        
        # 1. Expand table (Nếu có nhiều hơn 1 sản phẩm)
        if num_items > 1:
            ws.insert_rows(table_start_row + 1, amount=num_items - 1)
        
        # Tính lại vị trí dòng Tổng Cộng mới
        if total_row_template_idx:
            total_row = total_row_template_idx + (num_items - 1) if num_items > 0 else total_row_template_idx
        else:
            total_row = table_start_row + num_items # Vị trí nếu không tìm thấy dòng tổng cộng

        # Define styles
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 2. Copy styles (Giữ nguyên logic copy style)
        if num_items > 1:
            for i in range(1, num_items):
                target_row = table_start_row + i
                source_row = table_start_row
                for col in range(1, 16):
                    source_cell = ws.cell(row=source_row, column=col)
                    target_cell = ws.cell(row=target_row, column=col)
                    
                    if source_cell.border:
                        target_cell.border = style_copy(source_cell.border)
                    if source_cell.font:
                        target_cell.font = style_copy(source_cell.font)
                    if source_cell.alignment:
                        target_cell.alignment = style_copy(source_cell.alignment)
                    if source_cell.fill:
                        target_cell.fill = style_copy(source_cell.fill)
                    if source_cell.number_format:
                        target_cell.number_format = style_copy(source_cell.number_format)

        # 3. Clear the first row template marker
        ws.cell(row=table_start_row, column=1).value = ""

        # 4. Fill data
        for i, item in enumerate(products_data):
            row_idx = table_start_row + i
            
            # CRITICAL: Unmerge cells before writing (Giữ nguyên phần này)
            for col in range(1, 16):
                cell = ws.cell(row=row_idx, column=col)
                is_merged = False
                target_range = None
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        target_range = merged_range
                        break
                
                if is_merged and target_range:
                    try:
                        ws.unmerge_cells(str(target_range))
                    except KeyError:
                        pass
                
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border 

            # Map item fields
            item_map = {
                "Order__r.Name": item.get("Order__r", {}).get("Name") if item.get("Order__r") else "",
                "SKU__c": item.get("SKU__c"),
                "Vietnamese_Description__c": item.get("Vietnamese_Description__c"),
                "Length": item.get("Length__c"), "Width": item.get("Width__c"), "Height": item.get("Height__c"),
                "Quantity": item.get("Quantity__c"), "Crates__c": item.get("Crates__c"),
                "m2__c": item.get("m2__c"), "m3__c": item.get("m3__c"), "Tons__c": item.get("Tons__c"),
                "Cont__c": item.get("Cont__c"), "Packing__c": item.get("Packing__c"),
                "Delivery_Date__c": item.get("Delivery_Date__c")
            }

            # Write data - columns A through O
            ws.cell(row=row_idx, column=1).value = i + 1 
            ws.cell(row=row_idx, column=1).alignment = align_center
            
            ws.cell(row=row_idx, column=2).value = item_map["Order__r.Name"]
            ws.cell(row=row_idx, column=2).alignment = align_center
            
            ws.cell(row=row_idx, column=3).value = item_map["SKU__c"]
            ws.cell(row=row_idx, column=3).alignment = align_left
            
            desc_val = item_map["Vietnamese_Description__c"] or ""
            
            # Handle Bold Text before Hyphen (Nếu là RichText, việc merge sẽ cần xử lý khác)
            if desc_val and '-' in str(desc_val):
                parts = str(desc_val).split('-', 1)
                bold_part = parts[0]
                normal_part = '-' + parts[1]
                
                rich_text = CellRichText(
                    TextBlock(InlineFont(b=True, rFont='Times New Roman', sz=11), bold_part),
                    TextBlock(InlineFont(b=False, rFont='Times New Roman', sz=11), normal_part)
                )
                ws.cell(row=row_idx, column=4).value = rich_text
            else:
                ws.cell(row=row_idx, column=4).value = desc_val
            
            ws.cell(row=row_idx, column=4).alignment = align_left
            
            # Auto-adjust row height (Giữ nguyên phần này)
            desc_str = str(desc_val)
            explicit_lines = desc_str.count('\n') + 1
            wrap_lines = (len(desc_str) // 25) + 1 
            order_str = str(item_map["Order__r.Name"])
            order_lines = (len(order_str) // 10) + 1
            max_lines = max(explicit_lines, wrap_lines, order_lines)
            
            ws.row_dimensions[row_idx].height = max_lines * 20 if max_lines > 1 else 20
            
            # Size columns (E, F, G) - Center
            ws.cell(row=row_idx, column=5).value = item_map["Length"]
            ws.cell(row=row_idx, column=6).value = item_map["Width"]
            ws.cell(row=row_idx, column=7).value = item_map["Height"]
            
            # Quantity columns (H, I, J, K, L, M) - Center
            ws.cell(row=row_idx, column=8).value = item_map["Quantity"]
            ws.cell(row=row_idx, column=9).value = item_map["Crates__c"]
            
            m2_val = item_map["m2__c"]
            if m2_val is not None:
                ws.cell(row=row_idx, column=10).value = float(m2_val)
                ws.cell(row=row_idx, column=10).number_format = '0.00'
            
            m3_val = item_map["m3__c"]
            if m3_val is not None:
                ws.cell(row=row_idx, column=11).value = float(m3_val)
                ws.cell(row=row_idx, column=11).number_format = '0.00'

            ws.cell(row=row_idx, column=12).value = item_map["Tons__c"]
            ws.cell(row=row_idx, column=13).value = item_map["Cont__c"]
            
            for col in range(5, 14):
                ws.cell(row=row_idx, column=col).alignment = align_center

            packing_val = item_map["Packing__c"]
            if packing_val:
                try:
                    ws.cell(row=row_idx, column=14).value = float(packing_val)
                    ws.cell(row=row_idx, column=14).number_format = '0.0 "viên/kiện"'
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=14).value = f"{packing_val}\nviên/kiện"
                
                ws.cell(row=row_idx, column=14).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            del_date = item_map["Delivery_Date__c"]
            if del_date:
                try:
                    dt = datetime.datetime.strptime(del_date[:10], "%Y-%m-%d")
                    ws.cell(row=row_idx, column=15).value = dt.strftime("%d/%m/%Y")
                except:
                    ws.cell(row=row_idx, column=15).value = del_date
            ws.cell(row=row_idx, column=15).alignment = align_center

            # Apply borders
            for col in range(1, 16):
                ws.cell(row=row_idx, column=col).border = thin_border
    
    
    # ----------------------------------------------------
    # KHẮC PHỤC LỖI TÍNH TỔNG CỘNG BẰNG CÔNG THỨC EXCEL
    # ----------------------------------------------------
    if products_data and total_row_template_idx:
        first_data_row = table_start_row
        last_data_row = table_start_row + len(products_data) - 1
        
        # Dòng Tổng Cộng mới đã được tính ở trên: total_row
        
        # CRITICAL: Unmerge cells in Total row to ensure totals are visible
        # Check columns H (8) to M (13)
        for col in range(8, 14):
            cell = ws.cell(row=total_row, column=col)
            is_merged = False
            target_range = None
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    target_range = merged_range
                    break
            
            if is_merged and target_range:
                try:
                    ws.unmerge_cells(str(target_range))
                except KeyError:
                    pass

        # Cột H: Quantity (Viên)
        ws.cell(row=total_row, column=8).value = f"=SUM({get_column_letter(8)}{first_data_row}:{get_column_letter(8)}{last_data_row})"
        
        # Cột I: Crates (Kiện)
        ws.cell(row=total_row, column=9).value = f"=SUM({get_column_letter(9)}{first_data_row}:{get_column_letter(9)}{last_data_row})"

        # Cột J: M2 (m2__c)
        ws.cell(row=total_row, column=10).value = f"=SUM({get_column_letter(10)}{first_data_row}:{get_column_letter(10)}{last_data_row})"
        ws.cell(row=total_row, column=10).number_format = '0.00'
        
        # Cột K: M3 (m3__c)
        ws.cell(row=total_row, column=11).value = f"=SUM({get_column_letter(11)}{first_data_row}:{get_column_letter(11)}{last_data_row})"
        ws.cell(row=total_row, column=11).number_format = '0.00'
        
        # Cột L: Tons (Tấn)
        ws.cell(row=total_row, column=12).value = f"=SUM({get_column_letter(12)}{first_data_row}:{get_column_letter(12)}{last_data_row})"
        ws.cell(row=total_row, column=12).number_format = '0.00'
        
        # Cột M: Conts (Container)
        ws.cell(row=total_row, column=13).value = f"=SUM({get_column_letter(13)}{first_data_row}:{get_column_letter(13)}{last_data_row})"
        ws.cell(row=total_row, column=13).number_format = '0.00'
        
        # Căn chỉnh và định dạng cho các ô số
        for col in range(8, 14): 
            cell = ws.cell(row=total_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, name='Times New Roman', size=11)
            cell.border = thin_border

    # ----------------------------------------------------
    # MERGE I, J, K FOR ROWS WITH "Người soạn lệnh" OR "Ngọc Bích"
    # ----------------------------------------------------
    for r in range(1, ws.max_row + 1):
        found_keyword = False
        for cell in ws[r]:
            if cell.value and isinstance(cell.value, str):
                val_upper = str(cell.value).strip().upper()
                if "NGƯỜI SOẠN LỆNH" in val_upper or "NGƯỜI SOAN LỆNH" in val_upper or "NGỌC BÍCH" in val_upper:
                    found_keyword = True
                    break
        
        if found_keyword:
            # Merge columns I (9), J (10), K (11)
            try:
                ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
                ws.cell(row=r, column=9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except Exception as e:
                print(f"Error merging IJK at row {r}: {e}")

    # ----------------------------------------------------
    # KHẮC PHỤC LỖI MERGE CELL (Giữ nguyên logic merge)
    # ----------------------------------------------------

    # Hàm hỗ trợ lấy nội dung chuỗi (dùng để so sánh)
    def get_cell_content_for_comparison(cell):
        val = cell.value
        if isinstance(val, CellRichText):
            # Lấy nội dung chuỗi thuần túy từ CellRichText
            return str(val)
        return str(val).strip() if val is not None else ""


    # Merge duplicate "TÊN HÀNG" (Column D / 4)
    if products_data:
        start_row = table_start_row 
        end_row = table_start_row + len(products_data) - 1 
        
        merge_start_row = start_row
        current_val_str = get_cell_content_for_comparison(ws.cell(row=start_row, column=4))
        
        for r in range(start_row + 1, end_row + 2): 
            val_str = get_cell_content_for_comparison(ws.cell(row=r, column=4)) if r <= end_row else "SENTINEL" # Sử dụng giá trị phân biệt
            
            should_break = (val_str != current_val_str)
            
            if should_break:
                if r - 1 > merge_start_row: 
                    # Merge cells
                    ws.merge_cells(start_row=merge_start_row, start_column=4, end_row=r-1, end_column=4)
                    # Giữ nguyên căn chỉnh cho ô đầu tiên sau khi merge
                    ws.cell(row=merge_start_row, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                merge_start_row = r
                current_val_str = val_str
                
    # Merge duplicate "THỜI GIAN GIAO HÀNG" (Column O / 15)
    if products_data:
        start_row = table_start_row
        end_row = table_start_row + len(products_data) - 1
        
        merge_start_row = start_row
        current_val = ws.cell(row=start_row, column=15).value
        
        for r in range(start_row + 1, end_row + 2):
            val = ws.cell(row=r, column=15).value if r <= end_row else "SENTINEL"
            
            should_break = (val != current_val)
            
            if should_break:
                if r - 1 > merge_start_row:
                    # Merge cells
                    ws.merge_cells(start_row=merge_start_row, start_column=15, end_row=r-1, end_column=15)
                    # Giữ nguyên căn chỉnh cho ô đầu tiên sau khi merge
                    ws.cell(row=merge_start_row, column=15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                merge_start_row = r
                current_val = val

    wb.save(output_path)
    print(f"Filled template saved to: {output_path}")

@app.get("/generate-production-order/{contract_id}")
async def generate_production_order_endpoint(contract_id: str):
    try:
        template_path = os.getenv('PO_TEMPLATE_PATH', 'templates/production_order_template.xlsx')
        if not os.path.exists(template_path):
             # Fallback
             template_path = 'production_order_template.xlsx'
             
        # Call the UPDATED function directly
        result = generate_production_order_file(contract_id, template_path)
        return result

    except Exception as e:
        print(f"Error generating PO: {e}")
        raise HTTPException(status_code=500, detail=str(e))
        
@app.get("/num-to-words")
async def get_num_to_words(amount: float):
    try:
        # Chuyển đổi số sang chữ (tiếng Anh, định dạng USD)
        text_value = num2words(amount, to='currency', currency='USD')
        
        # Chỉ trả về text_raw
        return {"text_raw": text_value}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Base.vn Sync Logic ---

def format_date_for_base(iso_date):
    """Convert Salesforce ISO date to dd/mm/yyyy for Base"""
    if not iso_date:
        return ""
    try:
        dt_str = str(iso_date).split('.')[0]
        if 'T' in dt_str:
            dt_obj = datetime.datetime.strptime(dt_str, "%Y-%m-%dT%H:%M:%S")
        else:
             dt_obj = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
        return dt_obj.strftime("%d/%m/%Y")
    except Exception:
        return iso_date

def fetch_base_jobs_map(workflow_id, access_token):
    """Fetch all jobs from Base to map Name -> ID"""
    url_list = os.getenv('BASE_WORKFLOW_URL_LIST', "https://workflow.base.vn/extapi/v1/workflow/jobs")
    name_id_map = {}
    page_id = 0
    page_size = 50
    
    while True:
        payload = {
            "access_token": access_token,
            "id": workflow_id,
            "page_id": page_id,
            "page_size": page_size
        }
        try:
            resp = requests.post(url_list, data=payload, timeout=30)
            if resp.status_code != 200:
                print(f"Error fetching Base jobs: {resp.text}")
                break
            
            data = resp.json()
            jobs = data.get('jobs', [])
            if not jobs:
                break
                
            for job in jobs:
                job_name = job.get('name', '').strip()
                job_id = job.get('id')
                if job_name:
                    name_id_map[job_name] = job_id
            
            page_id += 1
            
        except Exception as e:
            print(f"Exception fetching Base jobs: {e}")
            break
            
    return name_id_map

@app.get("/sync-base-workflow")
async def sync_base_workflow(case_id: str = None):
    """
    Sync Salesforce Case to Base Workflow.
    """
    try:
        base_token = os.getenv('BASE_ACCESS_TOKEN')
        workflow_id = os.getenv('BASE_WORKFLOW_ID', '11180')
        url_create = os.getenv('BASE_WORKFLOW_URL_CREATE', "https://workflow.base.vn/extapi/v1/job/create")
        url_edit = os.getenv('BASE_WORKFLOW_URL_EDIT', "https://workflow.base.vn/extapi/v1/job/edit")
        creator = os.getenv('BASE_CREATOR_USERNAME', 'PhuongTran')
        followers_str = os.getenv('BASE_FOLLOWERS_LIST', "['bichnguyen', 'PhuongTran', 'tungpham']")
        
        followers = followers_str

        if not base_token:
            raise HTTPException(status_code=500, detail="BASE_ACCESS_TOKEN not set")

        sf = get_salesforce_connection()
        
        query_fields = "Id, CaseNumber, Subject, CreatedDate, So_LSX__c, Date_Export__c, Link_BM02__c, Number_Container__c, Customer_Complain_Content__c, Account.Account_Code__c"
        
        if case_id:
            query = f"SELECT {query_fields} FROM Case WHERE Id = '{case_id}'"
        else:
            query = f"SELECT {query_fields} FROM Case ORDER BY CreatedDate DESC LIMIT 1"
            
        result = sf.query_all(query)
        if not result['records']:
            return {"status": "error", "message": "No case found"}
            
        case_record = result['records'][0]
        
        subject = case_record.get('Subject')
        if not subject:
             return {"status": "error", "message": "Case has no Subject"}
        subject = subject.strip()
        
        account_code = ""
        if case_record.get('Account'):
             account_code = case_record['Account'].get('Account_Code__c', "")
             
        payload = {
            "access_token": base_token,
            "name": subject,
            "custom_ma_khach_hang": account_code,
            "custom_ngay_phan_anh": format_date_for_base(case_record.get("CreatedDate")),
            "custom_noi_dung_khieu_nai": case_record.get("Customer_Complain_Content__c", ""),
            "custom_so_container": case_record.get("Number_Container__c", ""),
            "custom_so_lenh_san_xuat": case_record.get("So_LSX__c", ""),
            "custom_chi_tiet_thong_tin_khieu_nai": case_record.get("Link_BM02__c", "")
        }
        
        base_jobs_map = fetch_base_jobs_map(workflow_id, base_token)
        
        if subject in base_jobs_map:
            job_id_base = base_jobs_map[subject]
            payload['id'] = job_id_base
            resp = requests.post(url_edit, data=payload)
            action = "UPDATE"
        else:
            payload['workflow_id'] = workflow_id
            payload['creator_username'] = creator
            payload['followers'] = followers
            resp = requests.post(url_create, data=payload)
            action = "CREATE"
            
        try:
            resp_json = resp.json()
        except:
            resp_json = resp.text

        if resp.status_code == 200:
            return {
                "status": "success", 
                "action": action, 
                "case_subject": subject,
                "base_response": resp_json
            }
        else:
            return {
                "status": "error",
                "action": action,
                "message": f"Base API Error: {resp.text}"
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def expand_case_items_table(ws, template_row, n):
    """Expand the case items table to accommodate n rows"""
    max_col = ws.max_column
    row_style = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=template_row, column=col)
        row_style.append(style_copy(cell._style) if cell.has_style else None)
    row_height = ws.row_dimensions[template_row].height
    add_rows = max(0, n - 1)
    
    # Handle merged cells
    merges_to_shift = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > template_row:
            merges_to_shift.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))
    
    for mr in merges_to_shift:
        rng = f"{get_column_letter(mr[2])}{mr[0]}:{get_column_letter(mr[3])}{mr[1]}"
        ws.unmerge_cells(rng)
    
    # Insert rows
    if add_rows > 0:
        ws.insert_rows(template_row + 1, amount=add_rows)
        for offset in range(1, add_rows + 1):
            r = template_row + offset
            for col in range(1, max_col + 1):
                dst = ws.cell(row=r, column=col)
                dst.value = None
                st = row_style[col - 1]
                if st is not None:
                    dst._style = style_copy(st)
            if row_height is not None:
                ws.row_dimensions[r].height = row_height
    
    # Re-merge shifted cells
    for mr in merges_to_shift:
        new_min_row = mr[0] + add_rows
        new_max_row = mr[1] + add_rows
        rng = f"{get_column_letter(mr[2])}{new_min_row}:{get_column_letter(mr[3])}{new_max_row}"
        ws.merge_cells(rng)


def generate_case_report(case_id: str, template_path: str):
    """Generate complaint case report"""
    sf = get_salesforce_connection()
    
    # 1. Get Case Data
    case_query = f"""
        SELECT
            Id, CaseNumber, Subject, CreatedDate,
            So_LSX__c, Date_Export__c, Link_BM02__c,
            Number_Container__c, Customer_Complain_Content__c,
            Account.Account_Code__c
        FROM Case
        WHERE Id = '{case_id}'
    """
    case_res = sf.query(case_query)
    if case_res['totalSize'] == 0:
        raise ValueError(f"Case not found: {case_id}")
    case_data = case_res['records'][0]
    
    # 2. Find Contract ID using So_LSX__c
    lsx_number = case_data.get('So_LSX__c')
    contract_id = None
    if lsx_number:
        contract_query = f"""
            SELECT Id 
            FROM Contract__c 
            WHERE Production_Order_Number__c = '{lsx_number}' 
            LIMIT 1
        """
        contract_res = sf.query(contract_query)
        if contract_res['totalSize'] > 0:
            contract_id = contract_res['records'][0]['Id']
    
    # 3. Get Products if Contract found
    products = []
    if contract_id:
        products_query = f"""
            SELECT Id, Name, 
                   Length__c, Width__c, Height__c, 
                   Vietnamese_Description__c, 
                   Line_number__c 
            FROM Order_Product__c 
            WHERE Contract_PI__r.Id = '{contract_id}' 
            ORDER BY Line_number__c ASC
        """
        prod_res = sf.query_all(products_query)
        products = prod_res['records']
        
    # 4. Load Template
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active # Assuming the template has only 1 sheet or active one is correct
    
    # helper for dates
    def fmt_date(d_str):
        if not d_str: return ""
        # Salesforce returns YYYY-MM-DDT... or YYYY-MM-DD
        return d_str.split('T')[0]

    # helper for dates
    def fmt_date(d_str):
        if not d_str: return ""
        # Salesforce returns YYYY-MM-DDT... or YYYY-MM-DD
        return d_str.split('T')[0]

    def html_to_richtext(content, base_font=None):
        if not content:
            return ""
        
        # Normalize newlines
        # Replace <p> with nothing (start) and </p> with \n (end)
        # Replace <br> with \n
        c = str(content)
        c = c.replace('<p>', '').replace('</p>', '\n')
        c = re.sub(r'<br\s*/?>', '\n', c, flags=re.IGNORECASE)
        
        # Split by bold tags
        # Capture <b>...</b> or <strong>...</strong>
        # Note: minimal regex, won't handle nested tags well but sufficient for basic Salesforce rich text
        parts = re.split(r'(<(?:b|strong)>.*?</(?:b|strong)>)', c, flags=re.IGNORECASE | re.DOTALL)
        
        rich_text = CellRichText()
        
        # Base font properties
        font_name = base_font.name if base_font else 'Calibri'
        font_size = base_font.sz if base_font else 11
        font_color = base_font.color if base_font else None
        
        normal_font = InlineFont(rFont=font_name, sz=font_size, color=font_color)
        bold_font = InlineFont(rFont=font_name, sz=font_size, color=font_color, b=True)
        
        for part in parts:
            if not part:
                continue
                
            # Check if bold
            bold_match = re.match(r'<(?:b|strong)>(.*?)</(?:b|strong)>', part, flags=re.IGNORECASE | re.DOTALL)
            if bold_match:
                text_content = bold_match.group(1)
                # Decode entities
                text_content = text_content.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"')
                rich_text.append(TextBlock(bold_font, text_content))
            else:
                # Normal text
                text_content = part
                # Strip HTML tags that might remain (like <div> etc if any, or just clean entities)
                text_content = re.sub(r'<[^>]+>', '', text_content)
                text_content = text_content.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"')
                rich_text.append(TextBlock(normal_font, text_content))
                
        # Trim trailing newlines from the last block if possible, or just string result
        # CellRichText doesn't support strip, so we rely on the split logic.
        return rich_text

    # 5. Scalar Replacements
    account = case_data.get('Account') or {}
    replacements = {
        '{{Account.Account_Code__c}}': account.get('Account_Code__c') or '',
        '{{Subject}}': case_data.get('Subject') or '',
        '{{CreatedDate}}': fmt_date(case_data.get('CreatedDate')),
        '{{So_LSX__c}}': case_data.get('So_LSX__c') or '',
        '{{Date_Export__c}}': fmt_date(case_data.get('Date_Export__c')),
        '{{Number_Container__c}}': case_data.get('Number_Container__c') or '',
        # Skip Customer_Complain_Content__c here to handle it separately
        '{{Customer_Complain_Content__c}}': '', 
    }
    
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                # Handle special RichText field
                if '{{Customer_Complain_Content__c}}' in cell.value:
                     # Get current font to preserve size/family
                     base_font = cell.font
                     richtext_val = html_to_richtext(case_data.get('Customer_Complain_Content__c'), base_font)
                     cell.value = richtext_val
                     cell.alignment = Alignment(wrap_text=True, vertical='top') # Ensure wrap
                     
                     # Approximate row height adjustment
                     # 1 line ~ 15 points. 
                     # Estimate lines based on text length and column width (approx 100 chars per line for wide merged cells)
                     # Better: count actual newlines + wrap
                     text_val = str(richtext_val).replace('\r', '') # richtext str conversion might be simple text
                     
                     # Simple estimation:
                     # Split by format-forced newlines
                     # Also estimate wrapping for long lines (e.g. > 100 chars)
                     lines = text_val.split('\n')
                     total_lines = 0
                     estimated_chars_per_line = 90 # Adjust based on column width in template
                     
                     for line in lines:
                        if len(line) > estimated_chars_per_line:
                            total_lines += (len(line) // estimated_chars_per_line) + 1
                        else:
                            total_lines += 1
                     
                     # Set height (minimum 15, add buffer)
                     new_height = max(1, total_lines) * 21 + 15 
                     ws.row_dimensions[cell.row].height = new_height
                     
                     continue

                for placeholder, value in replacements.items():
                    cell.value = cell.value.replace(placeholder, str(value))
    
    # Adjust height for "Supply Department Response" section (empty rows)
    # Finding rows under "PHẢN HỒI TỪ BỘ PHẬN CUNG ỨNG"
    # User feedback: "kéo rộng ra" -> likely means make rows taller.
    
    supply_header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "PHẢN HỒI TỪ BỘ PHẬN CUNG ỨNG" in cell.value:
                supply_header_row = cell.row
                break
        if supply_header_row:
            break
            
    if supply_header_row:
        # Expand next 3 rows (Nội dung, Nguyên Nhân, Kết luận)
        # Typically these are the next 3 rows.
        for i in range(1, 4):
            target_row = supply_header_row + i
            if target_row <= ws.max_row:
                ws.row_dimensions[target_row].height = 60 # Set to ~4 lines height

    
    # 6. Table Expansion for Products
    # Find table start
    table_start_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            if isinstance(cell.value, str) and '{{TableStart:ProPlanProduct}}' in cell.value:
                table_start_row = cell.row
                break
        if table_start_row:
            break
            
    if table_start_row:
        # Expand
        expand_case_items_table(ws, table_start_row, len(products) if products else 1)
        
        # Fill data
        for idx, item in enumerate(products):
            r = table_start_row + idx
            
            # Map columns
            # Column A (1): STT 
            # Column B (2): Description
            # Column C (3): Length
            # Column D (4): Width
            # Column E (5): Height
            
            # STT
            line_no = item.get('Line_number__c') 
            if line_no:
                try:
                    val_stt = int(float(line_no))
                except:
                    val_stt = idx + 1
            else:
                val_stt = idx + 1
            
            ws.cell(r, 1).value = val_stt
            
            # Description
            ws.cell(r, 2).value = item.get('Vietnamese_Description__c')
            
            # Helper to format decimals
            def fmt_dim(val):
                if val is None: return ""
                try:
                    v = float(val)
                    return f"{v:g}" # removes trailing zeros
                except:
                    return str(val)

            ws.cell(r, 3).value = fmt_dim(item.get('Length__c'))
            ws.cell(r, 4).value = fmt_dim(item.get('Width__c'))
            ws.cell(r, 5).value = fmt_dim(item.get('Height__c'))
            
    # Clean up markers
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if '{{TableStart:ProPlanProduct}}' in cell.value:
                    cell.value = cell.value.replace('{{TableStart:ProPlanProduct}}', '')
                if '{{TableEnd:ProPlanProduct}}' in cell.value:
                    cell.value = cell.value.replace('{{TableEnd:ProPlanProduct}}', '')
                if '{{Line_number__c}}' in cell.value:
                     cell.value = cell.value.replace('{{Line_number__c}}', '')

    # Save
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    sanitized_case_number = sanitize_filename(case_data.get('CaseNumber', 'Case'))
    file_name = f"Case_Report_{sanitized_case_number}_{timestamp}.xlsx"
    
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    wb.save(str(file_path))
    
    # Upload to Salesforce
    with open(file_path, "rb") as f:
        data = f.read()
    encoded = base64.b64encode(data).decode("utf-8")
    
    try:
        content_version = sf.ContentVersion.create({
            "Title": file_name.rsplit(".", 1)[0],
            "PathOnClient": file_name,
            "VersionData": encoded,
            "FirstPublishLocationId": case_id
        })
        cv_id = content_version['id']
    except Exception as e:
        print(f"Failed to upload to Salesforce: {e}")
        cv_id = None

    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": cv_id
    }









# --- Case Report Generation ---

def summarize_complaint_with_ai(complaint_text):
    """
    Summarize complaint text using Groq AI.
    Logic referenced from code.py
    """
    if not complaint_text or len(str(complaint_text).strip()) < 5:
        return "N/A"

    system_prompt = """
    Bạn là chuyên viên tóm tắt lỗi kỹ thuật. 
    Nhiệm vụ: Đọc khiếu nại và tóm tắt ngắn gọn lỗi sản phẩm thực tế.
    
    Quy tắc quan trọng:
    1. BỎ QUA hoàn toàn các đoạn về: dọa nạt, cảm xúc tức giận, yêu cầu bồi thường tiền, quy trình giấy tờ.
    2. CHỈ LẤY thông tin mô tả hiện trạng lỗi (mối mọt, gãy vỡ, sai kích thước) và bằng chứng cụ thể (số lượng).
    3. Output phải là một câu tiếng Việt ngắn gọn, trực diện.

    Ví dụ:
    Input: "Container MEDU6271240 bị côn trùng (tìm thấy một con bọ và bụi khoan trên 4 kiện gỗ). Khách hàng nhấn mạnh đây không còn là sự trùng hợp. Yêu cầu bồi thường chi phí dỡ hàng."
    Output: "Các thùng gỗ xuất hiện côn trùng, mối mọt. Tìm thấy một con bọ và bụi khoan trên 4 kiện gỗ."
    """

    try:
        completion = groq_client.chat.completions.create(
            model="openai/gpt-oss-20b", 
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": str(complaint_text)}
            ],
            temperature=0.2, 
        )
        return completion.choices[0].message.content.strip()
    except Exception as e:
        print(f"AI Summary Error: {e}")
        return f"N/A (AI Error)"

def html_to_richtext(html_str, base_font=None):
    """
    Convert HTML to OpenPyXL CellRichText.
    Supports <br>, <p>, </div> for newlines.
    Supports <b>, <strong> for bold.
    """
    if not html_str: return ""
    text = str(html_str)
    
    # Normalize Newlines
    text = re.sub(r'(?i)<br\s*/?>', '\n', text)
    text = re.sub(r'(?i)</p>', '\n', text)
    text = re.sub(r'(?i)</div>', '\n', text)
    
    # Strip unsupported tags
    text = re.sub(r'<[^>]+>', '', text)
    
    # Fix multiple newlines
    text = re.sub(r'\n+', '\n', text).strip()
    return text

def generate_case_report(case_id: str, template_path: str = "templates/case_template.xlsx"):
    print(f"--- Generating Case Report for Case ID: {case_id} ---")
    
    sf = get_salesforce_connection()
    
    # Query Case
    query = f"""
        SELECT 
            Id, CaseNumber, Subject, CreatedDate, 
            So_LSX__c, Date_Export__c, Link_BM02__c, 
            Number_Container__c, Customer_Complain_Content__c,
            Account.Name, Account.Account_Code__c
        FROM Case 
        WHERE Id = '{case_id}'
        LIMIT 1
    """
    
    try:
        result = sf.query_all(query)
    except Exception as e:
        raise Exception(f"Error querying Case: {e}")
        
    if not result['records']:
        raise Exception(f"Case with ID {case_id} not found.")
        
    case_data = result['records'][0]
    lsx_number = case_data.get('So_LSX__c')

    # Query Contract & Products
    products_data = []
    if lsx_number:
        # Check if LSX contains "-" (like 020725-02), use the base part for Contract query
        base_lsx = lsx_number.split('-')[0].strip()
        print(f"Found LSX: {lsx_number} (Base: {base_lsx}), searching for Contract...")
        try:
             contract_query = f"SELECT Id FROM Contract__c WHERE Production_Order_Number__c = '{base_lsx}' LIMIT 1"
             contract_res = sf.query(contract_query)
             if contract_res['totalSize'] > 0:
                 contract_id = contract_res['records'][0]['Id']
                 
                 prod_query = f"""
                    SELECT Id, Name, Length__c, Width__c, Height__c, 
                           Vietnamese_Description__c, Line_number__c 
                    FROM Order_Product__c 
                    WHERE Contract_PI__r.Id = '{contract_id}' 
                    ORDER BY Line_number__c ASC
                 """
                 records = sf.query_all(prod_query)['records']
                 
                 for i, item in enumerate(records):
                     products_data.append({
                         "Line_number__c": i + 1,
                         "Vietnamese_Description__c": item.get('Vietnamese_Description__c'),
                         "Length": item.get('Length__c'),
                         "Width": item.get('Width__c'),
                         "Height": item.get('Height__c')
                     })
        except Exception as e:
            print(f"Error querying Contract/Products: {e}")

    # Load Template
    print(f"Loading template: {template_path}")
    if not os.path.exists(template_path):
         if os.path.exists(os.path.basename(template_path)):
             template_path = os.path.basename(template_path)
         else:
             raise FileNotFoundError(f"Template not found: {template_path}")
             
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # Sanitize Template
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "{{" in cell.value:
                if '\n' in cell.value:
                     cell.value = cell.value.replace('\n', '')

    # AI Summary Logic
    raw_complaint = case_data.get('Customer_Complain_Content__c') or ""
    clean_complaint_text = html_to_richtext(raw_complaint)
    print("Calling AI for summary...")
    ai_summary = summarize_complaint_with_ai(clean_complaint_text)

    # Prepare Data
    data_map = {
        "{{CaseNumber}}": case_data.get('CaseNumber', ''),
        "{{Subject}}": case_data.get('Subject', ''),
        "{{Account.Account_Code__c}}": case_data.get('Account', {}).get('Account_Code__c') or "",
        "{{Account_Name}}": case_data.get('Account', {}).get('Name') or "",
        "{{CreatedDate}}": case_data.get('CreatedDate', '').split('.')[0],
        "{{So_LSX__c}}": case_data.get('So_LSX__c', ''),
        "{{Date_Export__c}}": case_data.get('Date_Export__c', ''),
        "{{Link_BM02}}": case_data.get('Link_BM02__c', ''),
        "{{Number_Container__c}}": case_data.get('Number_Container__c', ''),
        "{{summary}}": ai_summary
    }
    
    # Formatting Dates
    try:
        if data_map["{{CreatedDate}}"]:
            dt = datetime.datetime.strptime(data_map["{{CreatedDate}}"], "%Y-%m-%dT%H:%M:%S")
            data_map["{{CreatedDate}}"] = dt.strftime("%d/%m/%Y")
        if data_map["{{Date_Export__c}}"]:
            dt = datetime.datetime.strptime(data_map["{{Date_Export__c}}"], "%Y-%m-%d")
            data_map["{{Date_Export__c}}"] = dt.strftime("%d/%m/%Y")
    except: pass

    # Fill Table
    table_start_row = expand_table_by_tag(ws, "{{TableStart:ProPlanProduct}}", "{{TableEnd:ProPlanProduct}}", products_data)
    
    # Merging Logic for Product Table
    if table_start_row and products_data:
        num_products = len(products_data)
        for i in range(num_products):
            row_idx = table_start_row + i
            ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=8)
            
        # IMPORTANT: Fix overlap - DO NOT merge column 9 (Photos). 
        # Keeping it as individual cells allows better visibility of stacked images.
        # if num_products > 1:
        #     ws.merge_cells(start_row=table_start_row, start_column=9, 
        #                    end_row=table_start_row + num_products - 1, end_column=9)
        #     ws.cell(row=table_start_row, column=9).alignment = Alignment(horizontal='center', vertical='center')

        # --- PHOTO INTEGRATION (v13 IMPROVED) ---
        try:
            print(f"Fetching photos for Case {case_id}...")
            # Query ContentDocumentLinks to find actual attachments
            cdl_query = f"SELECT ContentDocumentId FROM ContentDocumentLink WHERE LinkedEntityId = '{case_id}'"
            cdl_res = sf.query_all(cdl_query)
            doc_ids = [r['ContentDocumentId'] for r in cdl_res['records']]
            
            if doc_ids:
                ids_str = "','".join(doc_ids)
                # Look for NEWEST 5 images, avoiding generating report files
                cv_query = f"""
                    SELECT Id, Title, FileExtension, VersionData, ContentDocumentId 
                    FROM ContentVersion 
                    WHERE ContentDocumentId IN ('{ids_str}') 
                    AND IsLatest = true 
                    AND FileExtension IN ('jpg','jpeg','png','gif') 
                    ORDER BY CreatedDate DESC 
                    LIMIT 5
                """
                cv_res = sf.query_all(cv_query)
                images = cv_res['records']
                print(f"Found {len(images)} photos to include.")
                
                if images:
                    from io import BytesIO
                    import requests
                    
                    col_letter = get_column_letter(9) # Photos column
                    
                    for idx, img_data in enumerate(images):
                        # Anchor to table_start_row + idx
                        # This places them in separate rows instead of a single merged area
                        target_row_idx = table_start_row + idx
                        
                        cv_id = img_data['Id']
                        img_url = f"{sf.base_url}sobjects/ContentVersion/{cv_id}/VersionData"
                        headers = {'Authorization': f'Bearer {sf.session_id}'}
                        
                        img_res = requests.get(img_url, headers=headers)
                        if img_res.status_code == 200:
                            img_stream = BytesIO(img_res.content)
                            try:
                                pil_img = OpenpyxlImage(img_stream)
                                
                                # Scale to fit width (~200 pixels)
                                original_w, original_h = pil_img.width, pil_img.height
                                target_w = 180 
                                scale = target_w / original_w
                                pil_img.width = target_w
                                pil_img.height = int(original_h * scale)
                                
                                # Anchor and add
                                anchor_cell = f"{col_letter}{target_row_idx}"
                                ws.add_image(pil_img, anchor_cell)
                                
                                # Set row height to fit the image
                                # 1 pixel ~= 0.75 points
                                target_h_points = pil_img.height * 0.75 + 10
                                current_h = ws.row_dimensions[target_row_idx].height or 15
                                ws.row_dimensions[target_row_idx].height = max(current_h, target_h_points)
                                print(f"Added photo {idx+1} to {anchor_cell}, set row height to {ws.row_dimensions[target_row_idx].height}")
                                
                            except Exception as e:
                                print(f"Error processing image {cv_id}: {e}")
            else:
                print("No attached files found for this Case.")
        except Exception as e:
            print(f"Error in Photo Integration: {e}")

    # Fill Placeholders
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                
                # Feedback Section Fixed Height
                val_lower = val.lower()
                if ("nội dung:" in val_lower or "nguyên nhân:" in val_lower or "kết luận và hướng xử lý:" in val_lower):
                     row_has_summary = any("{{summary}}" in str(c.value).lower() for c in ws[cell.row] if c.value)
                     if not row_has_summary:
                         ws.row_dimensions[cell.row].height = 100
    
                # SPECIAL: Customer Complain Content
                if "{{Customer_Complain_Content__c}}" in val:
                    raw_html = case_data.get('Customer_Complain_Content__c', '')
                    
                    if val.strip() == "{{Customer_Complain_Content__c}}":
                         clean_text = html_to_richtext(raw_html)
                         cell.value = clean_text
                         cell.alignment = Alignment(wrap_text=True, vertical='top')
                         
                         text_val = str(clean_text)
                         newlines = text_val.count('\n') + 1
                         estimated_lines = newlines
                         
                         wrapped_lines = sum(1 for line in text_val.split('\n') if len(line) > 90)
                         estimated_lines += wrapped_lines
                         
                         new_height = max(1, estimated_lines) * 21 + 15
                         ws.row_dimensions[cell.row].height = new_height
                    else:
                         try:
                             clean_text = html_to_richtext(raw_html)
                             cell.value = val.replace("{{Customer_Complain_Content__c}}", str(clean_text))
                         except:
                             cell.value = val.replace("{{Customer_Complain_Content__c}}", "")
                    continue
                
                # SPECIAL: AI Summary
                if "{{summary}}" in val:
                    new_val = val.replace("{{summary}}", ai_summary)
                    cell.value = new_val
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    continue

                # Generic Replacement
                for key, value in data_map.items():
                    if key in val:
                        new_val = val.replace(key, str(value) if value else "")
                        val = new_val 
                        cell.value = new_val
                        
    # Save File
    output_dir = get_output_directory()
    os.makedirs(output_dir, exist_ok=True)
    
    safe_case_number = sanitize_filename(case_data.get('CaseNumber', 'Unknown'))
    fn_ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    file_name = f"Case_{safe_case_number}_{fn_ts}.xlsx"
    file_path = output_dir / file_name
    
    print(f"Saving to local output: {file_path}")
    wb.save(str(file_path))

    # Upload to Salesforce
    print(f"Uploading to Salesforce for Case: {case_id}")
    with open(file_path, "rb") as f:
        file_data = f.read()
    encoded = base64.b64encode(file_data).decode("utf-8")
    
    try:
        content_version = sf.ContentVersion.create({
            "Title": file_name.rsplit(".", 1)[0],
            "PathOnClient": file_name,
            "VersionData": encoded,
            "FirstPublishLocationId": case_id
        })
        print(f"Upload Success! ContentVersion ID: {content_version['id']}")
        
        return {
            "status": "success",
            "file_path": str(file_path),
            "file_name": file_name,
            "salesforce_content_version_id": content_version["id"],
            "message": "Report generated and attached to Case successfully"
        }
    except Exception as e:
        print(f"Upload failed: {e}")
        return {
            "status": "partial_success",
            "file_path": str(file_path),
            "message": f"Generated file locally but failed to upload: {e}"
        }










@app.get("/generate-case-report/{case_id}")
async def generate_case_report_endpoint(case_id: str):
    """Generate Excel report for a Case"""
    try:
        template_path = os.getenv('CASE_TEMPLATE_PATH', 'templates/case_template.xlsx')
        if not os.path.exists(template_path):
             raise HTTPException(status_code=404, detail=f"Template not found: {template_path}")
             
        result = generate_case_report(case_id, template_path)
        return {
            "status": "success",
            "data": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
